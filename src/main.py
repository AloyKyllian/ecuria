from Planning import *
from Jour import *
import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
import os

path_cavalier_mercredi = "liste_cavalier_mercredi.txt"
path_cavalier_samedi = "liste_cavalier_samedi.txt"


def heure_precedant():
    """
    Définit l'heure précédente pour la cellule et appelle la fonction changer_heure.

    Cette fonction met à jour l'heure de la cellule en la définissant comme étant l'heure
    précédente par rapport à la liste des heures de travail du planning. Elle utilise la
    fonction changer_heure pour mettre à jour l'interface utilisateur.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    liste_heure = list(planning.liste_heure)[::-1]
    for i in range(1, len(liste_heure)):
        if liste_heure[i] == cellule.heure:
            cellule.set_heure(liste_heure[i-1])
            changer_heure()


def heure_suivant():
    """
    Définit l'heure suivante pour la cellule et appelle la fonction changer_heure.

    Cette fonction met à jour l'heure de la cellule en la définissant comme étant l'heure
    suivante par rapport à la liste des heures de travail du planning. Elle utilise la fonction
    changer_heure pour mettre à jour l'interface utilisateur.

    Args:
        Aucun.

    Returns:
        int: 0 si l'heure suivante est définie avec succès, sinon rien.
    """
    liste_heure = list(planning.liste_heure)[::-1]
    for i in range(0, len(liste_heure)-1):
        if liste_heure[i] == cellule.heure:
            cellule.set_heure(liste_heure[i+1])
            changer_heure()
            return 0


def ajouter():
    """
    Ajoute une cellule au planning avec des vérifications et met à jour l'interface.

    Cette fonction ajoute une cellule au planning en utilisant la méthode ajout de la classe
    Planning. Elle effectue des vérifications préliminaires et met à jour l'interface utilisateur
    en conséquence. Elle ajoute également une entrée à l'historique du planning en cas de succès.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    err = planning.ajout(cellule)
    if err == None or err == -4:
        if cellule.ind_eleve != -1:
            print(cellule.ind_eleve)
            colorier_eleve(cellule.ind_eleve)
        ajoutuncheval(cellule.cheval, cellule.ind_cheval)
        affichage_txt(jour, planning)
        label_enregistrer.config(fg="#f0f0f0")
        inserer_liste_de_travaille()
        ajout_historique(
            "ajout", (cellule.heure, cellule.cheval, cellule.eleve))
    elif err == -1:
        msgbox = tk.messagebox.showerror(
            title="creation de fichier", message="vous n'avez pas selectionné toutes les informations necessaire à l'ajout")
    elif err == -2:
        msgbox = tk.messagebox.showerror(
            title="creation de fichier", message="Ne peux etre ajouter car ce cheval travaille deja durant cette heure")
    elif err == -3:
        msgbox = tk.messagebox.showerror(
            title="creation de fichier", message="Ne peux etre ajouter car ce cheval travaille deja 4 heure dans la journée")
    elif err == -5:
        print(f"Erreur numéro {err}, ajout annulé.")


def supprimer():
    """
    Supprime une cellule du planning avec des vérifications et met à jour l'interface.

    Cette fonction supprime une cellule du planning en utilisant la méthode supprime de la classe
    Planning. Elle effectue des vérifications préliminaires et met à jour l'interface utilisateur
    en conséquence. Elle ajoute également une entrée à l'historique du planning en cas de succès.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    err = planning.supprime(cellule)
    if err == None:
        if cellule.ind_eleve != -1:
            print(cellule.ind_eleve)
            decolorier_eleve(cellule.ind_eleve)
        ajoutuncheval(cellule.cheval, cellule.ind_cheval)
        affichage_txt(jour, planning)
        label_enregistrer.config(fg="#f0f0f0")
        inserer_liste_de_travaille()
        ajout_historique("suppression", (cellule.heure,
                         cellule.cheval, cellule.eleve))
    else:
        msgbox = tk.messagebox.showerror(
            title="creation de fichier", message="suppression impossible vous voulez supprimez un element qui n'existe pas")


def inserer_liste_de_travaille():
    """
    Insère les heures de travail dans la listebox et met en évidence les heures travaillées.

    Cette fonction vide d'abord la listebox des heures de travail, puis insère les heures de
    travail associées au cheval de la cellule dans la listebox. Elle met également en évidence
    les heures travaillées dans la listebox.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    vider_listebox(heure_listebox)
    liste = planning.heure_travailler(cellule.cheval)
    for i in liste:
        heure_listebox.insert(tk.END, i)


def colorier_eleve(ind):
    """
    Change la couleur de fond de l'élément d'index donné dans la listebox des élèves en rouge.

    Cette fonction prend en argument l'index d'un élément dans la listebox des élèves et change
    la couleur de fond de cet élément en rouge.

    Args:
        ind (int): L'index de l'élément dans la listebox des élèves.

    Returns:
        Aucun.
    """
    eleve_listbox.itemconfig(ind, {'bg': 'red'})


def decolorier_eleve(ind):
    """
    Change la couleur de fond de l'élément d'index donné dans la listebox des élèves en blanc.

    Cette fonction prend en argument l'index d'un élément dans la listebox des élèves et change
    la couleur de fond de cet élément en blanc.

    Args:
        ind (int): L'index de l'élément dans la listebox des élèves.

    Returns:
        Aucun.
    """
    eleve_listbox.itemconfig(ind, {'bg': 'white'})


def ajout_historique(type, element):
    """
    Ajoute un élément à l'historique du planning et met à jour l'affichage dans la zone de texte.

    Cette fonction prend en argument un type d'action (ajout ou suppression) et un élément à ajouter
    à l'historique du planning. Elle met à jour l'affichage de l'historique dans la zone de texte
    correspondante.

    Args:
        type (str): Le type d'action (ajout ou suppression).
        element (tuple): L'élément à ajouter à l'historique.

    Returns:
        Aucun.
    """
    planning.append_historique(type, element)
    historique.delete('1.0', END)
    for i in planning.historique:
        historique.insert("1.0", f"{i}\r\n")


def affichage_txt(jour, planning):
    """
    Affiche le planning du jour dans la zone de texte correspondante.

    Cette fonction prend en argument un objet "jour" et un objet "planning", puis affiche le planning
    du jour dans la zone de texte correspondante.

    Args:
        jour (object): L'objet jour correspondant au jour de la semaine.
        planning (object): L'objet planning contenant les données du planning.

    Returns:
        Aucun.
    """
    visu_fichier.delete('1.0', END)
    visu_fichier.insert(END, planning.fichier(jour.j))


def vider_listebox(listebox):
    """
    Vide une listebox en supprimant tous ses éléments.

    Cette fonction prend en argument une listebox et supprime tous les éléments qu'elle contient.

    Args:
        listebox: La listebox à vider.

    Returns:
        Aucun.
    """
    if listebox.size() > 0:
        listebox.delete(0, listebox.size())


def inserer_listbox(i):
    """
    Insère un élément dans la listebox des élèves et le met en évidence s'il est présent dans le planning.

    Cette fonction prend en argument un élément "i" et l'insère dans la listebox des élèves. Si l'élément
    est présent dans le planning, il est mis en évidence en rouge.

    Args:
        i (str): L'élément à insérer dans la listebox des élèves.

    Returns:
        Aucun.
    """
    eleve_listbox.insert(tk.END, i)
    if len(planning.planning) != 0:
        present = any((cellule.heure, i) == (tup[0], tup[2])
                      for tup in planning.planning)
        if present:
            eleve_listbox.itemconfig(tk.END, {'bg': 'red'})


def ajouteleve():
    """
    Vide la listebox des élèves et insère les éléments correspondant à l'heure de la cellule.

    Cette fonction vide d'abord la listebox des élèves, puis insère les éléments correspondant à
    l'heure de la cellule. Elle met également en évidence les élèves qui sont présents dans le planning.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    vider_listebox(eleve_listbox)
    for i in planning.liste_eleve[cellule.heure]:
        inserer_listbox(i)


def ajoutuncheval(cheval, ind):
    """
    Supprime un élément de la listebox des chevaux et insère un nouvel élément à la position donnée.

    Cette fonction prend en argument le nom d'un cheval et sa position dans la listebox des chevaux.
    Elle supprime l'élément existant à cette position et insère un nouvel élément représentant le cheval.

    Args:
        cheval (str): Le nom du cheval.
        ind (int): La position dans la listebox.

    Returns:
        Aucun.
    """
    cheval_listbox.delete(ind)
    cheval_listbox.insert(
        ind, (planning.cheval[cheval][0], cheval))
    cheval_listbox.itemconfig(ind, {'bg': 'green'})


def ajoutcheval():
    """
    Vide la listebox des chevaux et insère les éléments correspondant à tous les chevaux du planning.

    Cette fonction vide d'abord la listebox des chevaux, puis insère les éléments correspondant à tous
    les chevaux du planning. Elle met en évidence les chevaux en vert s'ils sont associés à des élèves
    présents à l'heure de la cellule.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    cheval_listbox.delete(0, END)
    for i in planning.cheval:
        cheval_listbox.insert(tk.END, (planning.cheval[i][0], i))
    if cellule.heure != "heure":
        colorier()


def colorier():
    """
    Change la couleur de fond des éléments de la listebox des chevaux en fonction de leur disponibilité.

    Cette fonction met en évidence les chevaux en vert s'ils sont associés à des élèves présents à
    l'heure de la cellule. Les autres chevaux ont un fond blanc.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    for i in range(0, len(planning.cheval)):
        cheval_listbox.itemconfig(
            i, {'bg': 'white'})
    setcheval = set()
    for i in planning.liste_eleve[cellule.heure]:
        ancient = planning.ancient_cheval_de(i, cellule.heure)
        for y in ancient:
            setcheval.add(y[1])
    for i in setcheval:
        cheval_listbox.itemconfig(
            i, {'bg': 'green'})


def colorier_ancient_chevaux(ancient_cheval_eleve):
    """
    Change la couleur de fond des chevaux anciens en fonction de leur disponibilité.

    Cette fonction prend en argument une liste d'éléments représentant des chevaux anciens associés à
    un élève. Elle change la couleur de fond des chevaux en fonction de leur disponibilité, en utilisant
    des couleurs telles que jaune, orange et rouge en fonction du nombre d'anciens chevaux.

    Args:
        ancient_cheval_eleve (list): Une liste d'éléments représentant des chevaux anciens.

    Returns:
        Aucun.
    """
    if len(ancient_cheval_eleve) >= 3:
        cheval_listbox.itemconfig(
            ancient_cheval_eleve[2][1], {'bg': 'yellow'})
    if len(ancient_cheval_eleve) >= 2:
        cheval_listbox.itemconfig(
            ancient_cheval_eleve[1][1], {'bg': 'orange'})
    if len(ancient_cheval_eleve) >= 1:
        cheval_listbox.itemconfig(
            ancient_cheval_eleve[0][1], {'bg': 'red'})


def changer_heure():
    """
    Met à jour l'affichage en fonction de l'heure de la cellule.

    Cette fonction met à jour l'affichage en fonction de l'heure de la cellule, en vidant la listebox
    des élèves, en ajoutant les élèves correspondant à l'heure, en coloriant les chevaux disponibles,
    et en mettant à jour les variables d'interface utilisateur.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    vider_listebox(eleve_listbox)
    ajouteleve()
    colorier()
    varheure.set(f"{cellule.heure}")
    varajout.set(cellule.getCellule())


def changement_heure(i):
    """
    Change l'heure de la cellule et met à jour l'affichage.

    Cette fonction prend en argument une nouvelle heure "i", change l'heure de la cellule en conséquence
    et met à jour l'affichage.

    Args:
        i (str): La nouvelle heure pour la cellule.

    Returns:
        Aucun.
    """
    cellule.set_heure(i)
    changer_heure()


def cmp_dates(d):
    """
    Compare deux dates au format "jour-mois-année".

    Cette fonction prend en argument une date au format "jour-mois-année" et retourne un tuple (année, mois, jour)
    pour permettre une comparaison correcte des dates.

    Args:
        d (str): Une date au format "jour-mois-année".

    Returns:
        tuple: Un tuple (année, mois, jour).
    """
    j, m, a = d[0].split("-")
    return (int(a), int(m), int(j))


def recupperation_excel(listeself, name):
    """
    Récupère les données depuis un fichier Excel et les organise en listes et dictionnaires.

    Cette fonction prend en argument une liste "listeself" (non utilisée dans la fonction) et le nom d'un
    fichier Excel "name". Elle extrait les données de ce fichier, les organise en listes et dictionnaires
    (liste, dict_cheval, dict_heure) et les renvoie.

    Args:
        listeself: Liste (non utilisée dans la fonction).
        name (str): Le nom du fichier Excel à lire.

    Returns:
        tuple: Un tuple contenant trois éléments :
            - Une liste de tuples (heure, cheval, élève).
            - Un dictionnaire des chevaux avec leur indice et ligne.
            - Un dictionnaire des heures avec leur libellé.
    """
    workbook = load_workbook(name)
    sheet = workbook.active
    liste = []
    dict_heure = {}
    dict_cheval = {}
    Nb = 0
    for i in range(3, len(sheet["A"])):
        for j in list(range(1, len(sheet[3])))[::-1]:
            if sheet.cell(row=i, column=j).value != None:
                if i == 3:
                    dict_heure[sheet.cell(
                        row=i, column=j).value] = j
                elif j == 1:
                    dict_cheval[sheet.cell(row=i, column=j).value] = [
                        Nb, i]
                    Nb = 0
                elif j > 1 and sheet.cell(row=i, column=j).value != "MERCREDI" and sheet.cell(row=i, column=j).value != "SAMEDI" and sheet.cell(row=3, column=j).value != None and sheet.cell(row=i, column=1).value != None:
                    Nb += 1
                    liste.append(
                        (sheet.cell(row=3, column=j).value, sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=j).value))
    return liste, dict_cheval, dict_heure


def recup_donne():
    """
    Récupère les données depuis un fichier Excel et initialise le planning.

    Cette fonction affiche une boîte de dialogue pour sélectionner un fichier Excel, puis utilise la
    fonction "recupperation_excel" pour extraire les données du fichier, les initialise dans l'objet
    "planning" et met à jour l'affichage.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    planning.set_liste_eleve(lire_fichiers())
    msgbox = tk.messagebox.showinfo(
        title="Sélection de fichier", message="Veuillez sélectionner le fichier que vous souhaitez remplir")
    varjour.set(jour.j)
    chemin = tk.Tk()
    chemin.withdraw()                 # pour ne pas afficher la fenêtre Tk
    name = askopenfilename()
    planning.set_nom_fichier(name)
    planning.cheval.clear()
    planning.liste_heure.clear()
    dict_planning, cheval, heure = recupperation_excel("", name)
    planning.set_cheval(cheval)
    planning.set_heure(heure)
    planning.set_planning(dict_planning)
    ajoutcheval()
    visu_fichier.delete('1.0', END)
    visu_fichier.insert(END, planning.fichier(jour.j))
    folder = name.split('/')
    name = folder[-1]
    folder = folder[:-1]
    path = '/'.join(folder)
    liste = []
    for files in os.listdir(path):
        if files != name and ".xlsx" in files and '$' not in files:
            if jour.j == "Mercredi" and "mercredi" in files:
                liste.append((files[15:25], files))
            elif jour.j == "Samedi" and "samedi" in files:
                liste.append((files[13:23], files))
    liste = sorted(liste, key=cmp_dates, reverse=True)
    if len(liste) >= 1:
        ancient_planning, x, y = recupperation_excel(
            "ancien", path + '/' + liste[0][1])
        planning.set_ancien_planning(ancient_planning)
    if len(liste) >= 2:
        ancient_planning2, x, y = recupperation_excel(
            "ancien2", path + '/' + liste[1][1])
        planning.set_ancien_planning2(ancient_planning2)
    if len(liste) >= 3:
        ancient_planning3, x, y = recupperation_excel(
            "ancien3", path + '/' + liste[2][1])
        planning.set_ancien_planning3(ancient_planning3)
    msgbox = tk.messagebox.showinfo(
        title="Création de fichier", message="Tous les fichiers ont été récupérés")
    ajout_des_commande_lena()


def recup_donne_mercredi():
    """
    Définit le jour comme "Mercredi" et récupère les données depuis un fichier Excel.

    Cette fonction met le jour comme "Mercredi" en utilisant la méthode "setmercredi" de l'objet "jour"
    puis appelle la fonction "recup_donne" pour récupérer les données depuis un fichier Excel.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    jour.set_mercredi()
    recup_donne()


def recup_donne_samedi():
    """
    Définit le jour comme "Samedi" et récupère les données depuis un fichier Excel.

    Cette fonction met le jour comme "Samedi" en utilisant la méthode "setsamedi" de l'objet "jour"
    puis appelle la fonction "recup_donne" pour récupérer les données depuis un fichier Excel.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    jour.set_samedi()
    recup_donne()


def lire_fichiers():
    """
    Lit les fichiers correspondant au jour en cours et organise les données.

    Cette fonction lit les fichiers correspondant au jour en cours (Mercredi ou Samedi), organise les
    données des cavaliers et renvoie un dictionnaire avec les cavaliers et leurs listes d'élèves associées.

    Args:
        Aucun.

    Returns:
        dict: Un dictionnaire avec les noms des cavaliers en tant que clés et les listes d'élèves en tant que valeurs.
    """
    liste = []
    liste_eleve = {}
    if jour.j == "Mercredi":
        fichier_mer = open(path_cavalier_mercredi, "r")
        lignes = fichier_mer.read()
        lignes = lignes.split("\n")
        for ligne in lignes:
            if "H LENA" in ligne or "H30 LENA" in ligne or "H KARINE" in ligne:
                liste_eleve[ligne] = liste[:]
                liste.clear()
            else:
                liste.append(ligne)
        fichier_mer.close()
    elif jour.j == "Samedi":
        fichier_sam = open(path_cavalier_samedi, "r")
        lignes = fichier_sam.read()
        lignes = lignes.split("\n")
        for ligne in lignes:
            if "H00" in ligne or "30" in ligne:
                liste_eleve[ligne] = liste[:]
                liste.clear()
            else:
                liste.append(ligne)
        fichier_sam.close()
    return liste_eleve


def ecrire_fichier():
    """
    Enregistre les données dans un fichier Excel.

    Cette fonction ouvre le fichier Excel actuellement en cours de modification, efface les données des
    cellules correspondant au planning, puis insère les nouvelles données du planning. Elle enregistre
    ensuite le fichier Excel.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    workbook = load_workbook(planning.name_fichier)
    sheet = workbook.active
    for ligne in range(4, len(planning.cheval)+4):
        for colonne in range(2, len(planning.liste_heure)+2):
            sheet.cell(ligne, colonne).value = None
    for i in planning.planning:
        if sheet.cell(planning.cheval[i[1]][1], planning.liste_heure[i[0]]).value == None:
            sheet.cell(planning.cheval[i[1]][1],
                       planning.liste_heure[i[0]]).value = i[2]
    err = workbook.save(planning.name_fichier)
    if err == None:
        label_enregistrer.config(fg="#000000")


# Importation des modules
cellule = Cellule()  # Création d'une instance de la classe Cellule
planning = Planning()  # Création d'une instance de la classe Planning
jour = Jour()  # Création d'une instance de la classe Jour

# Création de l'interface utilisateur
window = tk.Tk()  # Création de la fenêtre principale
window.title("Planning")  # Titre de la fenêtre
window.attributes('-fullscreen', True)  # Affichage en mode plein écran
# Permet de quitter en appuyant sur la touche "Échap"
window.bind('<Escape>', lambda e: window.destroy())

# Création d'un cadre dans la fenêtre
frame = tk.Frame(master=window, width=300, height=100)
frame.pack()

# Définition des variables de contrôle
varheure = StringVar()
varjour = StringVar()
varcavalier = StringVar()
varcheval = StringVar()
varajout = StringVar()
varheure_cheval = StringVar()
varcavalier1 = StringVar()
varcavalier2 = StringVar()

# Éléments d'affichage dans la fenêtre
label_jour = tk.Label(window, textvariable=varjour)
label_jour.place(x=20, y=20)

label_heure = tk.Label(window, textvariable=varheure)
label_heure.place(x=70, y=20)

# Création d'une étiquette pour le titre
title_label = tk.Label(
    window, text="Gestion du Planning", font=("Helvetica", 24), bg="#f0f0f0")
title_label.place(x=400, y=20)

# Boutons pour avancer et reculer dans les heures
boutton_avancer_heure = tk.Button(
    window, width=2, text="<", command=heure_precedant)
boutton_avancer_heure.place(x=150, y=20)

boutton_reculer_heure = tk.Button(
    window, width=2, text=">", command=heure_suivant)
boutton_reculer_heure.place(x=180, y=20)

# Étiquettes pour afficher les informations du cavalier
label_cavalier = tk.Label(window, text="INFOS cavalier")
label_cavalier.place(x=450, y=70)

label_cavalier2 = tk.Label(
    window, text="la semaine dernière il/elle a monté : ")
label_cavalier2.place(x=450, y=100)

label_cavalier3 = tk.Label(
    window, textvariable=varcavalier)
label_cavalier3.place(x=650, y=100)

label_cavalier6 = tk.Label(
    window, text="il y a 2 semaines il/elle a monté : ")
label_cavalier6.place(x=450, y=120)

label_cavalier4 = tk.Label(
    window, textvariable=varcavalier1)
label_cavalier4.place(x=650, y=120)

label_cavalier7 = tk.Label(
    window, text="il y a 3 semaines il/elle a monté : ")
label_cavalier5 = tk.Label(
    window, textvariable=varcavalier2)
label_cavalier7.place(x=450, y=140)
label_cavalier5.place(x=650, y=140)

# Initialisation des variables de contrôle
varcavalier.set("cheval")
varcavalier1.set("cheval1")
varcavalier2.set("cheval2")

# Liste déroulante pour les élèves
eleve_listbox = tk.Listbox(window, yscrollcommand=True)
eleve_listbox.place(x=20, y=50)

# Fonction appelée lorsqu'un élément est sélectionné dans la liste des élèves


def items_selected(event):
    # Indices des éléments sélectionnés
    selected_indices = eleve_listbox.curselection()
    cellule.set_eleve(eleve_listbox.get(selected_indices), selected_indices[0])
    ancient_cheval = planning.ancient_cheval_de(
        cellule.eleve, cellule.heure)
    # Mise à jour des étiquettes des chevaux associés
    if len(ancient_cheval) >= 1:
        varcavalier.set(ancient_cheval[0][0])
    else:
        varcavalier.set("cheval")
    if len(ancient_cheval) >= 2:
        varcavalier1.set(ancient_cheval[1][0])
    else:
        varcavalier1.set("cheval1")
    if len(ancient_cheval) >= 3:
        varcavalier2.set(ancient_cheval[2][0])
    else:
        varcavalier2.set("cheval2")
    colorier()
    colorier_ancient_chevaux(ancient_cheval)
    for tup in planning.planning:
        if (cellule.heure, cellule.eleve) == (tup[0], tup[2]):
            cavalier = []
            cellule.set_cheval(tup[1], planning.index_cheval(tup[1]))
            ancient_cavalier = planning.ancient_eleve_de(cellule.cheval)
            for i in ancient_cavalier:
                cavalier.append(f"{i[0]} a {i[1]}")
            varheure_cheval.set(f"heure de travaille de : {cellule.cheval}")
            varcheval.set(cavalier)
            inserer_liste_de_travaille()
    varajout.set(cellule.getCellule())


# Association de la fonction à l'événement de relâchement du bouton de la souris
eleve_listbox.bind('<ButtonRelease-1>', items_selected)

# Étiquettes pour afficher les informations du cheval
label_cheval = tk.Label(window, text="INFOS cheval")
label_cheval.place(x=450, y=300)

label_cheval2 = tk.Label(
    window, text="la semaine dernière il/elle a été monté par : ")
label_cheval2.place(x=450, y=330)

label_cheval3 = tk.Label(
    window, textvariable=varcheval)
label_cheval3.place(x=700, y=330)

# Liste déroulante pour les chevaux
cheval_listbox = tk.Listbox(window, height=len(planning.cheval))
cheval_listbox.place(x=200, y=50)

# Fonction appelée lorsqu'un élément est sélectionné dans la liste des chevaux


def items_selected_cheval(event):
    # Indices des éléments sélectionnés
    selected_indices = cheval_listbox.curselection()
    cavalier = []
    cheval = cheval_listbox.get(selected_indices)
    cellule.set_cheval(cheval[1], selected_indices)
    ancient_cavalier = planning.ancient_eleve_de(
        cheval_listbox.get(selected_indices)[1])
    for i in ancient_cavalier:
        cavalier.append(f"{i[0]} a {i[1]}")
    varheure_cheval.set(f"heure de travaille de : {cellule.cheval}")
    inserer_liste_de_travaille()
    varcheval.set(cavalier)
    varajout.set(cellule.getCellule())


# Association de la fonction à l'événement de relâchement du bouton de la souris
cheval_listbox.bind('<ButtonRelease-1>', items_selected_cheval)

# Zone de texte pour afficher le planning
visu_fichier = tk.Text(window, width=100)
visu_fichier.place(x=700, y=360)

# Étiquette pour afficher des informations sur l'ajout
label_ajout = tk.Label(window, textvariable=varajout)
label_ajout.place(x=450, y=470)

# Bouton pour ajouter une entrée
boutton_ajouter = tk.Button(
    window, text="Ajouter", command=ajouter)
boutton_ajouter.place(x=450, y=500)

# Bouton pour supprimer une entrée
boutton_supprimer = tk.Button(
    window, text="supprimer", command=supprimer)
boutton_supprimer.place(x=520, y=500)

# Bouton pour enregistrer les modifications
boutton_enregistrer = tk.Button(
    window, text="enregistrer", command=ecrire_fichier, width=18)
boutton_enregistrer.place(x=450, y=530)

# Étiquette pour afficher un message après l'enregistrement
label_enregistrer = tk.Label(
    window, text="le fichier à bien été enregistré")
label_enregistrer.place(x=450, y=560)
label_enregistrer.config(fg="#f0f0f0")

# Étiquette pour afficher l'heure de travail du cheval
label_heure_cheval = tk.Label(
    window, textvariable=varheure_cheval)
label_heure_cheval.place(x=800, y=40)

# Liste déroulante pour les heures de travail
heure_listebox = tk.Listbox(window, width=25, height=5)
heure_listebox.place(x=800, y=70)

# Fonction appelée lorsqu'un élément est sélectionné dans la liste des heures de travail


def items_selected_heure_cheval(event):
    # Indices des éléments sélectionnés
    selected_indices = heure_listebox.curselection()
    (h, p) = heure_listebox.get(selected_indices)
    if h != cellule.heure:
        cellule.set_heure(h)
        cellule.set_eleve(p, -1)
    else:
        Nb = 0
        for i in range(0, eleve_listbox.size()):
            if p == eleve_listbox.get(i):
                cellule.set_eleve(p, Nb)
            Nb += 1
    varajout.set(cellule.getCellule())


# Association de la fonction à l'événement de relâchement du bouton de la souris
heure_listebox.bind('<ButtonRelease-1>', items_selected_heure_cheval)

# Étiquette pour afficher l'historique
label_historique = tk.Label(window, text="historique")
label_historique.place(x=1000, y=40)

# Zone de texte pour afficher l'historique
historique = tk.Text(window, width=60, height=13)
historique.place(x=1000, y=70)

# Création du menu
menubar = Menu(window)
# Création d'un sous-menu
filemenu = Menu(menubar, tearoff=0)
filemenu.add_command(label="Mercredi", command=recup_donne_mercredi)
filemenu.add_command(label="Samedi", command=recup_donne_samedi)

# Création d'une liste déroulante pour sélectionner l'heure
listeCombo = ttk.Combobox(window)
listeCombo.place(x=5, y=0)

# Fonction appelée lorsqu'un élément est sélectionné dans la liste déroulante


def action(event):
    select = listeCombo.get()  # Élément sélectionné dans la liste déroulante
    changement_heure(select)


listeCombo.bind("<<ComboboxSelected>>", action)

# Ajout des éléments au menu


def ajout_des_commande_lena():
    listeCombo.delete(0, "end")
    listeCombo['values'] = list(planning.liste_heure)[::-1]


# Ajout des éléments au menu
menubar.add_cascade(label="Jour", menu=filemenu)
menubar.add_command(label="Quitter!", command=window.quit)

# Affichage du menu dans la fenêtre
window.config(menu=menubar)

# Lancement de la boucle principale de l'application
window.mainloop()
