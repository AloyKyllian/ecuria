from Planning import *
from Ftp import *
from Jour import *
from Word import *
import Parametre as param
from Mail import *
from Zip import *
from Maj import *
from Log import setup_logger
import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter.filedialog import askopenfilename, askdirectory
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import os
from datetime import datetime, timedelta
from PIL import Image, ImageTk
import subprocess
from tkinter import messagebox

logger = setup_logger("ecuria", user="default")

path_parametre = "parametre/"
path_cavalier_mercredi = path_parametre + "liste_cavalier_mercredi.txt"
path_cavalier_samedi = path_parametre + "liste_cavalier_samedi.txt"
path_cavalier_semaine = path_parametre + "liste_cavalier_semaine.txt"
path_cheval = path_parametre + "liste_cheval.txt"
path_cheval_semaine = path_parametre + "liste_cheval_semaine.txt"
path_mail = path_parametre + "mail.txt"
path_user = "user.txt"
path_Mercredi = path_parametre + "Mercredi.xlsx"
path_Samedi = path_parametre + "Samedi.xlsx"
path_semaine = path_parametre + "Semaine.xlsx"
path_image = "image/"


def remplir_cheval(dict_chevaux):
    dict_cheval_temp = {}
    logger.debug("remplir_cheval    dict_chevaux %s", dict_chevaux)
    # # print("remplir_cheval    planning.planning",planning.planning)
    for cheval in dict_chevaux:
        dict_cheval_temp[cheval] = [dict_chevaux[cheval][0], planning.nb_heure(cheval)]
    return dict_cheval_temp


def get_personne():
    with open(path_user, "r") as file:
        return file.read()


def get_mail():
    moniteur = []
    tableau = []
    with open(path_mail, "r") as file:
        lines = file.readlines()
        tab = [line.strip() for line in lines]
    for val in tab:
        moniteur.append(val.split(":")[0])
        tableau.append(val.split(":")[1])
    logger.debug("tableau %s", tableau)
    logger.debug("moniteur %s", moniteur)
    # # print(tableau)
    # # print(moniteur)
    return tableau, moniteur


def unesessionmoins(eleve, heure):
    for i in range(len(planning.liste_eleve[heure])):
        if planning.liste_eleve[heure][i][0] == eleve:
            planning.liste_eleve[heure][i][1] -= 1
            if planning.liste_eleve[heure][i][1] == 0:
                planning.liste_eleve[heure][i][1] = 10
            break
    ajouteleve()


def unesessionplus(eleve, heure):
    for i in range(len(planning.liste_eleve[heure])):
        if planning.liste_eleve[heure][i][0] == eleve:
            planning.liste_eleve[heure][i][1] += 1
            if planning.liste_eleve[heure][i][1] > 10:
                planning.liste_eleve[heure][i][1] = 1
            break
    ajouteleve()


def ecrire_excel_ref(jour):
    global planning_theme
    # ouverture du fichier reference
    workbook = load_workbook(path_parametre + jour + ".xlsx")
    feuille = workbook.active

    # recuperer la mise en forme du fichier reference
    taillecellule = feuille.column_dimensions["B"].width
    hauteurcellule = feuille.row_dimensions[4].height
    taille_police = feuille.cell(4, 2).font.size

    # creation d'un nouveau fichier excel
    workbook = Workbook()
    sheet = workbook.active

    # # print("ecrire_excel_ref    dict_cheval[jour] ",dict_cheval[jour])
    logger.debug("ecrire_excel_ref    dict_cheval[jour] %s", dict_cheval[jour])
    liste_cheval = list(dict_cheval[jour].keys())
    heure_trier = list(dict_eleve[jour].keys())

    lignes = len(liste_cheval) + 4
    colonnes = len(heure_trier) + 2

    dico_numeros = {
        1: "A",
        2: "B",
        3: "C",
        4: "D",
        5: "E",
        6: "F",
        7: "G",
        8: "H",
        9: "I",
        10: "J",
        11: "K",
        12: "L",
        13: "M",
        14: "N",
        15: "O",
        16: "P",
        17: "Q",
        18: "R",
        19: "S",
        20: "T",
        21: "U",
        22: "V",
        23: "W",
        24: "X",
        25: "Y",
        26: "Z",
    }

    double = Side(border_style="thin", color="000000")

    ind_ligne = 0
    for ligne in range(3, lignes):
        ind_colonne = 0
        # mise en forme de la heuteur des lignes
        for colonne in range(1, colonnes):
            # mise en forme des cellules
            sheet.cell(ligne, colonne).font = Font(size=taille_police)
            sheet.cell(ligne, colonne).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            sheet.cell(ligne, colonne).border = Border(
                left=double, top=double, right=double, bottom=double
            )
            # mise en forme des heures
            if ligne == 3:
                # mise en forme de la largeur des lignes
                sheet.column_dimensions[dico_numeros[colonne]].width = taillecellule
                if colonne != 1:
                    # ajout des heures
                    sheet.cell(ligne, colonne).value = heure_trier[ind_colonne]
                    ind_colonne += 1
                if colonne % 2 == 0:
                    # creation des heures vertes
                    sheet.cell(ligne, colonne).fill = PatternFill(
                        start_color="70AD47", end_color="70AD47", fill_type="solid"
                    )
            # creation des lignes vertes
            if ligne % 2 == 0:
                sheet.cell(ligne, colonne).fill = PatternFill(
                    start_color="A9D08E", end_color="A9D08E", fill_type="solid"
                )
            # mise en forme des chevaux
            if colonne == 1:
                sheet.row_dimensions[ligne].height = hauteurcellule
                if ligne != 3:
                    # ajout des chevaux
                    sheet.cell(ligne, colonne).value = liste_cheval[ind_ligne]
                    ind_ligne += 1

    # ajout du titre theme
    sheet.cell(lignes, 1).value = "theme"

    # sauvegarde du fichier reference
    workbook.save(path_parametre + jour + ".xlsx")


def heure_precedant():
    """
    Définit l'heure précédente pour la cellule et appelle la fonction changer_heure.

    Cette fonction met à jour l'heure de la cellule en la définissant comme étant l'heure
    précédente par rapport à la liste des heures de travail du planning. Elle utilise la
    fonction changer_heure pour mettre à jour l'interface utilisateur.

    Args:
        Aucun.

    Returns:
        Aucun.
    """

    liste_heure = list(planning.liste_eleve)
    for i in range(1, len(liste_heure)):
        if liste_heure[i] == cellule.heure:
            cellule.set_heure(liste_heure[i - 1])
            changer_heure()


def heure_suivant():
    """
    Définit l'heure suivante pour la cellule et appelle la fonction changer_heure.

    Cette fonction met à jour l'heure de la cellule en la définissant comme étant l'heure
    suivante par rapport à la liste des heures de travail du planning. Elle utilise la fonction
    changer_heure pour mettre à jour l'interface utilisateur.

    Args:
        Aucun.

    Returns:
        int: 0 si l'heure suivante est définie avec succès, sinon rien.
    """
    liste_heure = list(planning.liste_eleve)
    for i in range(0, len(liste_heure) - 1):
        if liste_heure[i] == cellule.heure:
            cellule.set_heure(liste_heure[i + 1])
            changer_heure()
            return 0


def ajouter_rattrapage():
    planning.liste_eleve[cellule.heure].append([eleve_rattrapage.get().upper(), -2])
    ajouteleve()


def ecrire_word():
    err = False
    try:
        theme_t = [planning_theme, planning_theme1, planning_theme2, planning_theme3]
        eleves = lire_fichier_cavalier(jour.j)  # remplacer par planning.liste_eleve
        word(jour.j, nom_fichier, planning, theme_t, user, eleves)
    except Exception as e:
        err = True
        messagebox.showerror(
            "Erreur", f"Erreur lors de la création des fichiers Word : {e}"
        )
        logger.error(
            "Erreur lors de la création des fichiers Word : %s", e, exc_info=True
        )
    if not err:
        messagebox.showinfo("Information", "Les fichiers Word ont été créé avec succès")
        logger.debug("Les fichiers Word ont été créés avec succès.")


def ajouter_theme():
    global planning_theme
    theme.set(theme_entry.get())
    planning_theme[cellule.heure] = theme.get()


def ajouter():
    """
    Ajoute une cellule au planning avec des vérifications et met à jour l'interface.

    Cette fonction ajoute une cellule au planning en utilisant la méthode ajout de la classe
    Planning. Elle effectue des vérifications préliminaires et met à jour l'interface utilisateur
    en conséquence. Elle ajoute également une entrée à l'historique du planning en cas de succès.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    err = -2
    bg_color = cheval_listbox.itemcget(cellule.ind_cheval, "background")
    # # print(bg_color)
    logger.debug("bg_color %s", bg_color)
    if bg_color != "violet":
        err = planning.ajout(cellule)
    if err is None or err == -4:
        if elevecarte is True:
            unesessionmoins(cellule.eleve, cellule.heure)
        if cellule.ind_eleve != -1:
            colorier_eleve(cellule.ind_eleve)
        ajoutuncheval(cellule.cheval, cellule.ind_cheval)
        affichage_txt(jour, planning)
        label_enregistrer.config(fg="#b4b4b4")
        colorier_chevaux()
        inserer_liste_de_travaille()
        ajout_historique("ajout", (cellule.heure, cellule.cheval, cellule.eleve))
    elif err == -1:
        tk.messagebox.showerror(
            title="ajout",
            message="Vous n'avez pas sélectionné toutes les informations nécessaires à l'ajout.",
        )
        logger.warning(
            "Informations manquantes pour l'ajout du cheval %s à l'heure %s.",
            cellule.cheval,
            cellule.heure,
        )
    elif err == -2:
        tk.messagebox.showerror(
            title="ajout",
            message="La réservation ne peut être ajoutée, car ce cheval travaille déjà durant cette heure.",
        )
        logger.warning(
            "Le cheval %s travaille déjà durant cette heure.", cellule.cheval
        )
    elif err == -3:
        tk.messagebox.showerror(
            title="ajout",
            message="La réservation ne peut être ajoutée, car ce cheval a déjà travaillé 4 heures dans la journée.",
        )
        logger.warning(
            "Le cheval %s a déjà travaillé 4 heures dans la journée.", cellule.cheval
        )
    elif err == -5:
        # print(f"Erreur numéro {err} : ajout annulé.")
        logger.warning(f"Erreur numéro {err} : ajout annulé.")
        pass


def supprimer():
    """
    Supprime une cellule du planning avec des vérifications et met à jour l'interface.

    Cette fonction supprime une cellule du planning en utilisant la méthode supprime de la classe
    Planning. Elle effectue des vérifications préliminaires et met à jour l'interface utilisateur
    en conséquence. Elle ajoute également une entrée à l'historique du planning en cas de succès.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    err = planning.supprime(cellule)
    if err is None:
        if elevecarte is True:
            unesessionplus(cellule.eleve, cellule.heure)
        if cellule.ind_eleve != -1:
            decolorier_eleve(cellule.ind_eleve)
        ajoutuncheval(cellule.cheval, cellule.ind_cheval)
        affichage_txt(jour, planning)
        label_enregistrer.config(fg="#b4b4b4")
        colorier_chevaux()
        inserer_liste_de_travaille()
        ajout_historique("suppression", (cellule.heure, cellule.cheval, cellule.eleve))
    else:
        tk.messagebox.showerror(
            title="creation de fichier",
            message="Suppression impossible : vous voulez supprimer un élément qui n'existe pas.",
        )
        logger.warning(
            "Suppression impossible : l'élément n'existe pas dans le planning."
        )


def inserer_liste_de_travaille():
    """
    Insère les heures de travail dans la listebox et met en évidence les heures travaillées.

    Cette fonction vide d'abord la listebox des heures de travail, puis insère les heures de
    travail associées au cheval de la cellule dans la listebox. Elle met également en évidence
    les heures travaillées dans la listebox.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    vider_listebox(heure_listebox)
    liste = planning.heure_travailler(cellule.cheval)
    for i in liste:
        heure_listebox.insert(tk.END, i)


def colorier_eleve(ind):
    """
    Change la couleur de fond de l'élément d'index donné dans la listebox des élèves en rouge.

    Cette fonction prend en argument l'index d'un élément dans la listebox des élèves et change
    la couleur de fond de cet élément en rouge.

    Args:
        ind (int): L'index de l'élément dans la listebox des élèves.

    Returns:
        Aucun.
    """
    if ind in range(0, eleve_listbox.size()):
        eleve_listbox.itemconfig(ind, {"bg": "red"})


def decolorier_eleve(ind):
    """
    Change la couleur de fond de l'élément d'index donné dans la listebox des élèves en blanc.

    Cette fonction prend en argument l'index d'un élément dans la listebox des élèves et change
    la couleur de fond de cet élément en blanc.

    Args:
        ind (int): L'index de l'élément dans la listebox des élèves.

    Returns:
        Aucun.
    """
    if ind in range(0, eleve_listbox.size()):
        eleve_listbox.itemconfig(ind, {"bg": "white"})


def ajout_historique(type, element):
    """
    Ajoute un élément à l'historique du planning et met à jour l'affichage dans la zone de texte.

    Cette fonction prend en argument un type d'action (ajout ou suppression) et un élément à ajouter
    à l'historique du planning. Elle met à jour l'affichage de l'historique dans la zone de texte
    correspondante.

    Args:
        type (str): Le type d'action (ajout ou suppression).
        element (tuple): L'élément à ajouter à l'historique.

    Returns:
        Aucun.
    """
    historique.config(state="normal")
    planning.append_historique(type, element)
    historique.delete("1.0", END)
    for i in planning.historique:
        historique.insert("1.0", f"{i}\r\n")
    historique.config(state="disabled")


def affichage_txt(jour, planning):
    """
    Affiche le planning du jour dans la zone de texte correspondante.

    Cette fonction prend en argument un objet "jour" et un objet "planning", puis affiche le planning
    du jour dans la zone de texte correspondante.

    Args:
        jour (object): L'objet jour correspondant au jour de la semaine.
        planning (object): L'objet planning contenant les données du planning.

    Returns:
        Aucun.
    """
    visu_fichier.config(state="normal")
    visu_fichier.delete("1.0", END)
    visu_fichier.insert(END, planning.fichier(jour.j))
    visu_fichier.config(state="disabled")


def vider_listebox(listebox):
    """
    Vide une listebox en supprimant tous ses éléments.

    Cette fonction prend en argument une listebox et supprime tous les éléments qu'elle contient.

    Args:
        listebox: La listebox à vider.

    Returns:
        Aucun.
    """
    if listebox.size() > 0:
        listebox.delete(0, listebox.size())


def inserer_listbox(i):
    """
    Insère un élément dans la listebox des élèves et le met en évidence s'il est présent dans le planning.

    Cette fonction prend en argument un élément "i" et l'insère dans la listebox des élèves. Si l'élément
    est présent dans le planning, il est mis en évidence en rouge.

    Args:
        i (str): L'élément à insérer dans la listebox des élèves.

    Returns:
        Aucun.
    """
    if i[1] == -1:
        eleve_listbox.insert(tk.END, i[0])
    else:
        eleve_listbox.insert(tk.END, i)
    i = i[0]
    if len(planning.planning) != 0:
        present = any(
            (cellule.heure, i) == (tup[0], tup[2]) for tup in planning.planning
        )
        if present:
            eleve_listbox.itemconfig(tk.END, {"bg": "red"})


def ajouteleve():
    """
    Vide la listebox des élèves et insère les éléments correspondant à l'heure de la cellule.

    Cette fonction vide d'abord la listebox des élèves, puis insère les éléments correspondant à
    l'heure de la cellule. Elle met également en évidence les élèves qui sont présents dans le planning.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    if cellule.heure in planning.liste_eleve:
        vider_listebox(eleve_listbox)
        if cellule.heure != "heure":
            for i in planning.liste_eleve[cellule.heure]:
                inserer_listbox(i)


def ajoutuncheval(cheval, ind):
    """
    Supprime un élément de la listebox des chevaux et insère un nouvel élément à la position donnée.

    Cette fonction prend en argument le nom d'un cheval et sa position dans la listebox des chevaux.
    Elle supprime l'élément existant à cette position et insère un nouvel élément représentant le cheval.

    Args:
        cheval (str): Le nom du cheval.
        ind (int): La position dans la listebox.

    Returns:
        Aucun.
    """
    cheval_listbox.delete(ind)
    cheval_listbox.insert(ind, (planning.cheval[cheval][1], cheval))


def ajoutcheval():
    """
    Vide la listebox des chevaux et insère les éléments correspondant à tous les chevaux du planning.

    Cette fonction vide d'abord la listebox des chevaux, puis insère les éléments correspondant à tous
    les chevaux du planning. Elle met en évidence les chevaux en vert s'ils sont associés à des élèves
    présents à l'heure de la cellule.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    cheval_listbox.delete(0, END)
    for i in planning.cheval:
        cheval_listbox.insert(tk.END, (planning.cheval[i][1], i))
    if cellule.heure != "heure" and cellule.heure in planning.liste_heure:
        colorier()
        colorier_chevaux()


def colorier():
    """
    Change la couleur de fond des éléments de la listebox des chevaux en fonction de leur disponibilité.

    Cette fonction met en évidence les chevaux en vert s'ils sont associés à des élèves présents à
    l'heure de la cellule. Les autres chevaux ont un fond blanc.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    for i in range(0, len(planning.cheval)):
        cheval_listbox.itemconfig(i, {"bg": "white"})


def colorier_ancient_chevaux(ancient_cheval_eleve):
    """
    Change la couleur de fond des chevaux anciens en fonction de leur disponibilité.

    Cette fonction prend en argument une liste d'éléments représentant des chevaux anciens associés à
    un élève. Elle change la couleur de fond des chevaux en fonction de leur disponibilité, en utilisant
    des couleurs telles que jaune, orange et rouge en fonction du nombre d'anciens chevaux.

    Args:
        ancient_cheval_eleve (list): Une liste d'éléments représentant des chevaux anciens.

    Returns:
        Aucun.
    """

    if len(ancient_cheval_eleve) >= 3 and ancient_cheval_eleve[2][1] != "":
        cheval_listbox.itemconfig(ancient_cheval_eleve[2][1], {"bg": "yellow"})
    if len(ancient_cheval_eleve) >= 2 and ancient_cheval_eleve[1][1] != "":
        cheval_listbox.itemconfig(ancient_cheval_eleve[1][1], {"bg": "orange"})
    if len(ancient_cheval_eleve) >= 1 and ancient_cheval_eleve[0][1] != "":
        cheval_listbox.itemconfig(ancient_cheval_eleve[0][1], {"bg": "red"})


def colorier_chevaux():
    if jour.j != "Semaine":
        for cell in planning.planning:
            if cell[0][0:2] in cellule.heure and cell[1] in planning.cheval:
                cheval_listbox.itemconfig(planning.cheval[cell[1]][0], {"bg": "violet"})
    elif jour.j == "Semaine":
        for cell in planning.planning:
            # print( cell[0][-4:-1] , cellule.heure)
            logger.debug(
                "cell[0][-4:-1] %s cellule.heure %s", cell[0][-4:-1], cellule.heure
            )
            if (
                cell[0][0:2] in cellule.heure
                and cell[1] in planning.cheval
                and cell[0][-4:-1] in cellule.heure
            ):
                cheval_listbox.itemconfig(planning.cheval[cell[1]][0], {"bg": "violet"})


def changer_heure():
    """
    Met à jour l'affichage en fonction de l'heure de la cellule.

    Cette fonction met à jour l'affichage en fonction de l'heure de la cellule, en vidant la listebox
    des élèves, en ajoutant les élèves correspondant à l'heure, en coloriant les chevaux disponibles,
    et en mettant à jour les variables d'interface utilisateur.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    ajouteleve()
    colorier()
    colorier_chevaux()
    varheure.set(f"{cellule.heure}")
    varajout.set(cellule.getCellule())
    if cellule.heure in planning_theme:
        theme.set(planning_theme[cellule.heure])
    else:
        theme.set("thème")
    if cellule.heure in planning_theme1:
        theme1.set(planning_theme1[cellule.heure])
    else:
        theme1.set("thème1")
    if cellule.heure in planning_theme2:
        theme2.set(planning_theme2[cellule.heure])
    else:
        theme2.set("thème2")
    if cellule.heure in planning_theme3:
        theme3.set(planning_theme3[cellule.heure])
    else:
        theme3.set("thème3")


def changement_heure(i):
    """
    Change l'heure de la cellule et met à jour l'affichage.

    Cette fonction prend en argument une nouvelle heure "i", change l'heure de la cellule en conséquence
    et met à jour l'affichage.

    Args:
        i (str): La nouvelle heure pour la cellule.

    Returns:
        Aucun.
    """
    cellule.set_heure(i)
    changer_heure()


def cmp_dates(d):
    """
    Compare deux dates au format "jour-mois-année".

    Cette fonction prend en argument une date au format "jour-mois-année" et retourne un tuple (année, mois, jour)
    pour permettre une comparaison correcte des dates.

    Args:
        d (str): Une date au format "jour-mois-année".

    Returns:
        tuple: Un tuple (année, mois, jour).
    """

    j, m, a = d[0].split("-")

    return (int(a), int(m), int(j))


def recupperation_excel(name):
    """
    Récupère les données depuis un fichier Excel et les organise en listes et dictionnaires.

    Cette fonction prend en argument une liste "listeself" (non utilisée dans la fonction) et le nom d'un
    fichier Excel "name". Elle extrait les données de ce fichier, les organise en listes et dictionnaires
    (liste, dict_cheval, dict_heure) et les renvoie.

    Args:
        listeself: Liste (non utilisée dans la fonction).
        name (str): Le nom du fichier Excel à lire.

    Returns:
        tuple: Un tuple contenant trois éléments :
            - Une liste de tuples (heure, cheval, élève).
            - Un dictionnaire des chevaux avec leur indice et ligne.
            - Un dictionnaire des heures avec leur libellé.
    """
    workbook = load_workbook(name)
    sheet = workbook.active
    liste = []
    planning_theme = {}
    dict_heure = {}
    dict_cheval = {}
    heure_temp = {}
    Nb = 0
    for i in range(3, len(sheet["A"]) + 1):
        for j in list(range(1, len(sheet[3]) + 1))[::-1]:
            valeur_case = str(sheet.cell(row=i, column=j).value).strip()

            if valeur_case != "None":
                if i == 3:
                    dict_heure[valeur_case.upper()] = j
                    heure_temp[j] = valeur_case
                elif j == 1:
                    dict_cheval[valeur_case] = [Nb, i]
                    Nb = 0
                elif (
                    j > 1
                    and valeur_case != "MERCREDI"
                    and valeur_case != "SAMEDI"
                    and sheet.cell(row=3, column=j).value is not None
                    and sheet.cell(row=i, column=1).value is not None
                    and str(sheet.cell(row=i, column=1).value).strip() != "theme"
                ):
                    Nb += 1
                    liste.append(
                        (
                            sheet.cell(row=3, column=j).value.strip(),
                            sheet.cell(row=i, column=1).value.strip(),
                            valeur_case,
                        )
                    )
                elif (
                    j > 1 and str(sheet.cell(row=i, column=1).value).strip() == "theme"
                ):
                    planning_theme[heure_temp[j].upper()] = valeur_case
    return liste, dict_cheval, dict_heure, planning_theme


def sort_files_by_date(files):
    # Trie les fichiers par date (les plus récents en premier)
    sorted_files = sorted(files, key=lambda x: x[1], reverse=True)
    return sorted_files


def extract_date_from_filename(filename):
    # Extrait la date du nom du fichier
    date_str = filename.split()[-1].split(".")[0]
    file_date = datetime.strptime(date_str, "%d-%m-%Y")
    return file_date


def recup_donne():
    """
    Récupère les données depuis un fichier Excel et initialise le planning.

    Cette fonction affiche une boîte de dialogue pour sélectionner un fichier Excel, puis utilise la
    fonction "recupperation_excel" pour extraire les données du fichier, les initialise dans l'objet
    "planning" et met à jour l'affichage.

    Args:
        Aucun.

    Returns:
        Aucun.
    """

    global \
        ancient_nom, \
        nom_fichier, \
        planning_theme, \
        planning_theme1, \
        planning_theme2, \
        planning_theme3
    interface_default()
    tk.messagebox.showinfo(
        title="Sélection de fichier",
        message="Veuillez sélectionner le fichier que vous souhaitez compléter.",
    )
    chemin = tk.Tk()
    chemin.withdraw()  # pour ne pas afficher la fenêtre Tk
    path = askopenfilename()
    nom_fichier = []
    folder = path.split("/")
    name = folder[-1]
    nom_fichier.append(name)
    # print(name)
    logger.debug("name %s", name)
    if "mercredi" in name.lower():
        jour.set_mercredi()
    elif "samedi" in name.lower():
        jour.set_samedi()
    elif "semaine" in name.lower():
        jour.set_semaine()
    else:
        tk.messagebox.showerror(
            title="Erreur de fichier",
            message="Le fichier sélectionné n'est pas un fichier de planning.",
        )
        logger.warning("Le fichier %s n'est pas un fichier de planning.", name)
        return
    planning_theme = {}
    planning_theme1 = {}
    planning_theme2 = {}
    planning_theme3 = {}
    planning.cheval.clear()
    planning.liste_heure.clear()

    date = extract_date_from_filename(name)
    # print("recup_donne    date",date)
    logger.debug("date %s", date)
    planning.set_liste_eleve(dict_eleve[jour.j].copy())
    # print("recup_donne    planning.liste_eleve",planning.liste_eleve)
    logger.debug("planning.liste_eleve %s", planning.liste_eleve)
    varjour.set(jour.j + " " + date.strftime("%d-%m-%y"))

    planning.set_nom_fichier(path)

    dict_planning, cheval, heure, planning_theme = recupperation_excel(path)

    planning.set_heure(heure)
    planning.set_planning(dict_planning)
    planning.set_cheval(remplir_cheval(dict_cheval[jour.j]))
    affichage_txt(jour, planning)

    liste = []
    path = path.replace(name, "")
    files = []
    for file in os.listdir(path):
        if (
            jour.j.lower() in file.lower()
            and "~$" not in file.lower()
            and ".xlsx" in file.lower()
        ):
            files.append((file, extract_date_from_filename(file)))
    liste = sort_files_by_date(files)
    # print("recup_donne    liste",liste)
    logger.debug("liste %s", liste)
    for i in range(len(liste)):
        if liste[i][0] == name:
            selected_ind = i

    # print("recup_donne    selected_ind",selected_ind)
    logger.debug("selected_ind %s", selected_ind)
    if len(liste) < 4:
        nb_fichier = len(liste) - selected_ind
    else:
        nb_fichier = 4
    # print("recup_donne    nb_fichier",nb_fichier)
    logger.debug("nb_fichier %s", nb_fichier)
    liste = [item[0] for item in liste[selected_ind + 1 : selected_ind + nb_fichier]]
    # print("recup_donne    liste",liste)
    logger.debug("liste %s", liste)
    if len(liste) > 0:
        semaine = liste[0].replace(".xlsx", "")
        varsemaine1.set(semaine.replace("liste", "s-1"))
        ancient_nom = path + liste[0]
        nom_fichier.append(liste[0])
        ancient_planning, x, y, planning_theme1 = recupperation_excel(path + liste[0])
        planning.set_ancien_planning(ancient_planning)
    else:
        planning.set_ancien_planning([])
        varsemaine1.set("Aucun fichier")
    if len(liste) > 1:
        semaine = liste[1].replace(".xlsx", "")
        varsemaine2.set(semaine.replace("liste", "s-2"))
        nom_fichier.append(liste[1])
        ancient_planning2, x, y, planning_theme2 = recupperation_excel(path + liste[1])
        planning.set_ancien_planning2(ancient_planning2)
    else:
        varsemaine2.set("Aucun fichier")
        planning.set_ancien_planning2([])
    if len(liste) > 2:
        semaine = liste[2].replace(".xlsx", "")
        varsemaine3.set(semaine.replace("liste", "s-3"))
        nom_fichier.append(liste[2])
        ancient_planning3, x, y, planning_theme3 = recupperation_excel(path + liste[2])
        planning.set_ancien_planning3(ancient_planning3)
    else:
        varsemaine3.set("Aucun fichier")
        planning.set_ancien_planning3([])
    tk.messagebox.showinfo(
        title="Création de fichier", message="Tous les fichiers ont été récupérés"
    )
    logger.debug("Fichier %s chargé avec succès.", name)
    remplir_listbox_heure()
    liste_heure = list(planning.liste_eleve.keys())
    ajoutcheval()
    changement_heure(liste_heure[0])


def ecrire_fichier():
    """
    Enregistre les données dans un fichier Excel.

    Cette fonction ouvre le fichier Excel actuellement en cours de modification, efface les données des
    cellules correspondant au planning, puis insère les nouvelles données du planning. Elle enregistre
    ensuite le fichier Excel.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    global planning_theme

    workbook = load_workbook(path_parametre + jour.j + ".xlsx")

    # Accéder à la feuille de calcul souhaitée
    sheet = workbook.active

    liste_cheval = list(dict_cheval[jour.j].keys())
    heure_trier = list(dict_eleve[jour.j].keys())

    # print("ecrire_fichier    plannning.liste_heure",planning.liste_heure)
    logger.debug("plannning.liste_heure %s", planning.liste_heure)
    dict_heure = {}
    Nb = 0
    for heure in heure_trier:
        dict_heure[heure] = Nb + 2
        Nb += 1
    planning.set_heure(dict_heure)
    # print("ecrire_fichier    plannning.liste_heure",planning.liste_heure)
    logger.debug("plannning.liste_heure %s", planning.liste_heure)
    for cellule in planning.planning:
        if cellule[1] in planning.cheval and cellule[0] in planning.liste_heure:
            sheet.cell(
                planning.cheval[cellule[1]][0] + 4, planning.liste_heure[cellule[0]]
            ).value = cellule[2]
    for heure in planning_theme:
        sheet.cell(len(liste_cheval) + 4, dict_heure[heure]).value = planning_theme[
            heure
        ]
    sheet.cell(len(liste_cheval) + 4, 1).value = "theme"

    err = workbook.save(planning.name_fichier)
    if err is None:
        label_enregistrer.config(fg="#ffffff")
        cavalier = lire_fichier_cavalier(jour.j)
        fichier = open(path_parametre + "liste_cavalier_" + jour.j + ".txt", "w")
        fichier.write(ecrire_fichier_cavalier(cavalier, carte=True))
        fichier.close()


def add_heure():
    global dict_eleve
    try:
        # ajout de l'heure dans le dictionnaire
        dict_eleve[parajour][para_input_heure.get().upper()] = []
        # tri des heures
        heure_trier = tri_heure(dict_eleve[parajour], parajour)
        # creation d'un dictionnaire temporaire afin d'appliquer le tri a dict_eleve
        dictionnaire_eleve = {}
        for heure in heure_trier:
            dictionnaire_eleve[heure] = dict_eleve[parajour][heure]
        dict_eleve[parajour] = dictionnaire_eleve.copy()
        del dictionnaire_eleve
        # mise a jour de la listebox des heures
        para_inserer_listebox(dict_eleve[parajour])
        # mise a jour de l'affichage
        visualiser_fichier_cavalier(dict_eleve[parajour])
    except:
        messagebox.showerror(
            "Erreur",
            "L'heure existe déjà ou n'est pas valide (la forme correcte doit commencer par un nombre).",
        )
        logger.error(
            "Tentative d'ajout d'une heure invalide : %s", para_input_heure.get()
        )


def suppr_heure(dict_eleve, heure):
    dict_eleve[parajour] = param.suppr_heure(dict_eleve[parajour], heure)
    para_inserer_listebox(dict_eleve[parajour])
    visualiser_fichier_cavalier(dict_eleve[parajour])


def add_eleve():
    global dict_eleve
    global heure
    try:
        nb_carte = -1
        if v.get() == 1:
            nb_carte = para_nbcarte.get()
        dict_eleve[parajour][heure].append(
            [para_input_eleve.get().upper().strip(), int(nb_carte)]
        )
        remplirlisteboxeleve()
        visualiser_fichier_cavalier(dict_eleve[parajour])
    except:
        messagebox.showerror("Erreur", "L'élève existe déjà ou n'est pas valide.")
        logger.error(
            "Tentative d'ajout d'un élève invalide : %s", para_input_eleve.get()
        )


def add_cheval():
    liste_cheval = []
    for cheval in dict_cheval[parajour]:
        liste_cheval.append([dict_cheval[parajour][cheval][0], cheval])

    if not para_input_ind_chevaux.get():
        ind_cheval = len(liste_cheval)
    else:
        ind_cheval = int(para_input_ind_chevaux.get())

    present = False
    if para_input_chevaux.get() in dict_cheval[parajour]:
        messagebox.showerror("Erreur", "Le cheval est déjà présent dans la liste.")
        logger.error(
            "Le cheval %s est déjà présent dans la liste.", para_input_chevaux.get()
        )
        present = True

    # print(dict_cheval[parajour])
    # print(liste_cheval,ind_cheval)
    # print(present)
    logger.debug("dict_cheval[parajour] %s", dict_cheval[parajour])
    logger.debug("liste_cheval %s", liste_cheval)
    logger.debug("ind_cheval %s", ind_cheval)
    logger.debug("present %s", present)
    if not present and len(liste_cheval) >= ind_cheval:
        for cheval in liste_cheval:
            if cheval[0] >= ind_cheval:
                cheval[0] += 1
        liste_cheval.append([ind_cheval, para_input_chevaux.get().upper()])
        liste_cheval.sort()
        # print(liste_cheval)
        logger.debug("liste_cheval %s", liste_cheval)
        dict_cheval_temp = {}
        for i in liste_cheval:
            if i[1] in dict_cheval[parajour]:
                dict_cheval_temp[i[1]] = [i[0], dict_cheval[parajour][i[1]][1]]
            else:
                dict_cheval_temp[i[1]] = [i[0], planning.nb_heure(i[1])]
        if parajour != "Semaine":
            dict_cheval["Mercredi"] = dict_cheval_temp
            dict_cheval["Samedi"] = dict_cheval_temp
        else:
            dict_cheval["Semaine"] = dict_cheval_temp
        remplirlisteboxcheval(liste_cheval)


def suppr_cheval():
    # print(cheval)
    logger.debug("cheval %s", cheval)
    liste_cheval = []
    for chevali in dict_cheval[parajour]:
        liste_cheval.append([dict_cheval[parajour][chevali][0], chevali])
    # print(dict_cheval[parajour])
    logger.debug("dict_cheval[parajour] %s", dict_cheval[parajour])
    if cheval in liste_cheval:
        liste_cheval.remove(cheval)
        for che in liste_cheval:
            if che[0] >= cheval[0]:
                che[0] -= 1
        liste_cheval.sort()
        # print(liste_cheval)
        logger.debug("liste_cheval %s", liste_cheval)
        for i in liste_cheval:
            dict_cheval[parajour][i[1]] = [i[0], dict_cheval[parajour][i[1]][1]]
        if parajour != "Semaine":
            del dict_cheval["Mercredi"][cheval[1]]
        else:
            del dict_cheval[parajour][cheval[1]]
        # print(dict_cheval[parajour])
        logger.debug("dict_cheval[parajour] %s", dict_cheval[parajour])
        remplirlisteboxcheval(liste_cheval)


def suppr_eleve():
    liste = []
    for eleves in dict_eleve[parajour][heure]:
        # print(eleves, eleve)
        logger.debug("eleves %s eleve %s", eleves, eleve)
        if (eleves[0], eleves[1]) != (eleve[0], eleve[1]):
            # print(eleves, eleve)
            logger.debug("eleves %s eleve %s", eleves, eleve)
            liste.append(eleves)
    dict_eleve[parajour][heure] = liste[:]
    remplirlisteboxeleve()
    visualiser_fichier_cavalier(dict_eleve[parajour])


def add_moniteur():
    global moniteur, mail
    if para_input_moniteur.get() not in moniteur and para_input_moniteur.get() != "":
        if para_input_mail.get() != "" and para_input_mail.get() not in mail:
            mail.append(para_input_mail.get())
            moniteur.append(para_input_moniteur.get())

            remplirlisteboxmoniteur(moniteur, mail)


def suppr_moniteur():
    global moniteur, mail
    if moniteur:
        if mail:
            mail.remove(mail[para_listebox_moniteur.curselection()[0]])
            moniteur.remove(moniteur[para_listebox_moniteur.curselection()[0]])
            remplirlisteboxmoniteur(moniteur, mail)


def remplirlisteboxmoniteur(moniteurs, mail):
    para_listebox_moniteur.delete(0, END)
    for i in range(len(moniteurs)):
        para_listebox_moniteur.insert(END, moniteurs[i] + ":" + mail[i])


def ecrire_fichier_cavalier(liste_eleve, carte=False):
    txt = ""

    heure_trier = list(liste_eleve.keys())
    eleves = liste_eleve

    # print("ecrire_fichier_cavalier  heure_trier",heure_trier)
    logger.debug("heure_trier %s", heure_trier)
    for heure in heure_trier:
        ind = 0
        for eleve in eleves[heure]:
            if carte and ((eleve[1] >= 0 and eleve[1] <= 10) or eleve[1] == -1):
                txt += (
                    planning.liste_eleve[heure][ind][0]
                    + "/"
                    + str(planning.liste_eleve[heure][ind][1])
                    + "\r"
                )
            elif eleve[1] != -2:
                txt += eleve[0] + "/" + str(eleve[1]) + "\r"
            ind += 1
        txt += "\\heure/" + heure + "\r"
    txt += "\\Fin fichier/"
    return txt


def ecrire_fichier_cheval(dict_chevaux):
    liste_cheval = []
    for cheval in dict_chevaux:
        liste_cheval.append([dict_chevaux[cheval][0], cheval])
    txt = ""
    for cheval in liste_cheval:
        txt += cheval[1] + "\r"
    return txt


def para_enregistrer():
    global parajour, user, dict_cheval, dict_eleve
    err = False
    try:
        if para_listeCombo_user.get() in moniteur:
            with open(path_user, "w") as fichier:
                fichier.write(para_listeCombo_user.get())
                user = para_listeCombo_user.get()
                user_var.set(user)

        if parajour != "" and dict_eleve[parajour]:
            fichier = open(path_parametre + "liste_cavalier_" + parajour + ".txt", "w")
            fichier.write(ecrire_fichier_cavalier(dict_eleve[parajour]))
            fichier.close()
        if dict_cheval[parajour] and parajour != "Semaine":
            fichier = open(path_cheval, "w")
            fichier.write(ecrire_fichier_cheval(dict_cheval[parajour]))
            fichier.close()
        elif dict_cheval[parajour]:
            fichier = open(path_cheval_semaine, "w")
            fichier.write(ecrire_fichier_cheval(dict_cheval[parajour]))
            fichier.close()
        visualiser_fichier_cavalier(dict_eleve[parajour])
        # print("enregistrer excel ref")
        logger.debug("enregistrer excel ref")
        ecrire_excel_ref("Mercredi")
        ecrire_excel_ref("Samedi")
        ecrire_excel_ref("Semaine")
        sauvegarder_mail()

    except Exception as e:
        err = True
        messagebox.showerror(
            "Erreur", f"Erreur lors de l'enregistrement des paramètres : {e}."
        )
        logger.error(
            "Erreur lors de l'enregistrement des paramètres : %s", e, exc_info=True
        )
    if not err:
        messagebox.showinfo(
            "Enregistrement", "Les paramètres ont été enregistrés avec succès !"
        )
        logger.debug("Les paramètres ont été enregistrés avec succès.")


def remplirlisteboxcheval(chevaux):
    para_listebox_chevaux.delete(0, END)
    for cheval in chevaux:
        para_listebox_chevaux.insert(END, cheval)


def remplirlisteboxeleve():
    if heure != "":
        para_listebox_eleve.delete(0, END)
        for eleve in dict_eleve[parajour][heure]:
            para_listebox_eleve.insert(END, eleve)


def tri_heure(dictionnaire_eleve, jour):
    liste_heure_non_trier = list(dictionnaire_eleve.keys())
    # print("interface_default liste_heure_non_trier",liste_heure_non_trier)
    logger.debug("liste_heure_non_trier %s", liste_heure_non_trier)
    if jour != "Semaine":
        liste_heure_trier = sorted(liste_heure_non_trier, key=cmp_heure)
    elif jour == "Semaine":
        liste_heure_trier = sorted(liste_heure_non_trier, key=cmp_heure_semaine)
    return liste_heure_trier


def interface_default():
    global parajour
    for widget in widgets_parametre:
        widget.place_forget()

    image_label.destroy()

    # Réajuster les placements en utilisant les nouvelles proportions
    label_jour.place(x=int(135 * proportion_x), y=int(70 * proportion_y))
    label_heure.place(x=int(150 * proportion_x), y=int(145 * proportion_y))
    title_label.place(x=int(60 * proportion_x), y=int(35 * proportion_y))
    boutton_avancer_heure.place(x=int(65 * proportion_x), y=int(140 * proportion_y))
    boutton_reculer_heure.place(x=int(260 * proportion_x), y=int(140 * proportion_y))
    label_cavalier.place(x=int(470 * proportion_x), y=int(70 * proportion_y))
    label_cavalier2.place(x=int(470 * proportion_x), y=int(100 * proportion_y))
    label_cavalier3.place(x=int(660 * proportion_x), y=int(100 * proportion_y))
    label_cavalier6.place(x=int(470 * proportion_x), y=int(150 * proportion_y))
    label_cavalier4.place(x=int(660 * proportion_x), y=int(150 * proportion_y))
    label_cavalier7.place(x=int(470 * proportion_x), y=int(200 * proportion_y))
    label_cavalier5.place(x=int(660 * proportion_x), y=int(200 * proportion_y))
    boutton_absent.place(x=int(755 * proportion_x), y=int(100 * proportion_y))
    boutton_correction.place(x=int(810 * proportion_x), y=int(100 * proportion_y))
    eleve_listbox.place(x=int(133 * proportion_x), y=int(170 * proportion_y))
    eleve_rattrapage.place(x=int(133 * proportion_x), y=int(390 * proportion_y))
    label_eleve_rattrapage.place(x=int(137 * proportion_x), y=int(360 * proportion_y))
    boutton_eleve_rattrapage.place(x=int(160 * proportion_x), y=int(420 * proportion_y))
    cheval_listbox.place(x=int(330 * proportion_x), y=int(35 * proportion_y))
    visu_fichier.place(x=int(900 * proportion_x), y=int(395 * proportion_y))
    label_visu_fichier.place(x=int(900 * proportion_x), y=int(365 * proportion_y))
    label_ajout.place(x=int(470 * proportion_x), y=int(400 * proportion_y))
    boutton_ajouter.place(x=int(570 * proportion_x), y=int(480 * proportion_y))
    boutton_supprimer.place(x=int(670 * proportion_x), y=int(480 * proportion_y))
    boutton_enregistrer.place(x=int(570 * proportion_x), y=int(530 * proportion_y))
    label_enregistrer.place(x=int(560 * proportion_x), y=int(585 * proportion_y))
    label_heure_cheval.place(x=int(470 * proportion_x), y=int(250 * proportion_y))
    heure_listebox.place(x=int(470 * proportion_x), y=int(280 * proportion_y))
    historique.place(x=int(900 * proportion_x), y=int(70 * proportion_y))
    label_historique.place(x=int(900 * proportion_x), y=int(40 * proportion_y))
    label_user.place(x=int(60 * proportion_x), y=int(70 * proportion_y))
    listeCombo.place(x=int(65 * proportion_x), y=int(100 * proportion_y))
    image1.place(x=int(535 * proportion_x), y=int(606 * proportion_y))
    image2.place(x=int(70 * proportion_x), y=int(600 * proportion_y))
    image3.place(x=int(680 * proportion_x), y=int(220 * proportion_y))
    bouton_ouvrir_excel.place(x=int(1400 * proportion_x), y=int(60 * proportion_y))
    bouton_rafraichir.place(x=int(1400 * proportion_x), y=int(100 * proportion_y))
    label_theme.place(x=int(133 * proportion_x), y=int(460 * proportion_y))
    theme_entry.place(x=int(133 * proportion_x), y=int(490 * proportion_y))
    boutton_theme.place(x=int(140 * proportion_x), y=int(520 * proportion_y))
    label_theme_actuelle.place(x=int(160 * proportion_x), y=int(550 * proportion_y))
    label_theme_avant1.place(x=int(660 * proportion_x), y=int(125 * proportion_y))
    label_theme_avant2.place(x=int(660 * proportion_x), y=int(175 * proportion_y))
    label_theme_avant3.place(x=int(660 * proportion_x), y=int(225 * proportion_y))
    bouton_word.place(x=int(1400 * proportion_x), y=int(140 * proportion_y))
    bouton_mail.place(x=int(1400 * proportion_x), y=int(180 * proportion_y))
    bouton_fusion.place(x=int(1400 * proportion_x), y=int(220 * proportion_y))
    info.place(x=int(460 * proportion_x), y=int(35 * proportion_y))
    info2.place(x=int(1400 * proportion_x), y=int(260 * proportion_y))

    if jour.j != "":
        planning.set_liste_eleve(dict_eleve[jour.j].copy())
        planning.set_cheval(remplir_cheval(dict_cheval[jour.j].copy()))
        ajoutcheval()
        ajouteleve()

    # print("interface_default planning.liste_eleve",planning.liste_eleve)
    # print("interface_default planning.cheval",planning.cheval)
    logger.debug("planning.liste_eleve %s", planning.liste_eleve)
    logger.debug("planning.cheval %s", planning.cheval)

    remplir_listbox_heure()


def interface_paramete():
    global dict_cheval, dict_eleve
    for widget in widgets_principaux:
        widget.place_forget()

    image_label.destroy()

    # Réajustement pour les autres widgets...
    para_image1.place(relx=0.48, rely=0.6, anchor=tk.CENTER)
    para_visu_fichier.place(x=int(900 * proportion_x), y=int(395 * proportion_y))
    para_listebox_heure.place(x=int(400 * proportion_x), y=int(70 * proportion_y))
    para_listebox_eleve.place(x=int(730 * proportion_x), y=int(70 * proportion_y))
    para_listeCombo.place(x=int(65 * proportion_x), y=int(40 * proportion_y))
    para_listeCombo_user.place(x=int(170 * proportion_x), y=int(40 * proportion_y))
    para_input_heure.place(x=int(560 * proportion_x), y=int(140 * proportion_y))
    para_add_heure.place(x=int(560 * proportion_x), y=int(170 * proportion_y))
    para_suppr_heure.place(x=int(560 * proportion_x), y=int(200 * proportion_y))
    para_input_eleve.place(x=int(890 * proportion_x), y=int(135 * proportion_y))
    para_add_eleve.place(x=int(890 * proportion_x), y=int(165 * proportion_y))
    para_suppr_eleve.place(x=int(890 * proportion_x), y=int(195 * proportion_y))
    para_boutton_enregistrer.place(x=int(635 * proportion_x), y=int(680 * proportion_y))

    para_listebox_chevaux.place(x=int(60 * proportion_x), y=int(70 * proportion_y))
    para_input_chevaux.place(x=int(220 * proportion_x), y=int(140 * proportion_y))
    para_add_chevaux.place(x=int(220 * proportion_x), y=int(170 * proportion_y))
    para_suppr_chevaux.place(x=int(220 * proportion_x), y=int(200 * proportion_y))
    para_input_ind_chevaux.place(x=int(360 * proportion_x), y=int(140 * proportion_y))
    para_case.place(x=int(890 * proportion_x), y=int(225 * proportion_y))
    para_nbcarte.place(x=int(890 * proportion_x), y=int(245 * proportion_y))
    image3.place(x=int(170 * proportion_x), y=int(500 * proportion_y))
    image4.place(x=int(1070 * proportion_x), y=int(70 * proportion_y))

    if user.upper() == "LENA" or user.upper() == "MANON":
        posymail = 280 * proportion_y
        para_listebox_moniteur.place(x=int(1000 * proportion_x), y=int(posymail))

        para_add_moniteur.place(x=int(830 * proportion_x), y=int(posymail + 55))
        para_suppr_moniteur.place(x=int(830 * proportion_x), y=int(posymail + 85))

        para_label_moniteur.place(x=int(730 * proportion_x), y=int(posymail))
        para_input_moniteur.place(x=int(820 * proportion_x), y=int(posymail))
        para_label_mail.place(x=int(730 * proportion_x), y=int(posymail + 30))
        para_input_mail.place(x=int(820 * proportion_x), y=int(posymail + 30))

    para_bouton_importer_param.place(
        x=int(1300 * proportion_x), y=int(290 * proportion_y)
    )
    para_bouton_exporter_param.place(
        x=int(1300 * proportion_x), y=int(325 * proportion_y)
    )
    para_bouton_ouvrir_excel.place(
        x=int(1295 * proportion_x), y=int(360 * proportion_y)
    )

    dict_eleve, dict_cheval = lire_parametre()
    remplirlisteboxmoniteur(moniteur, mail)


def lire_fichier_chevaux(path):
    liste = {}
    fichier = open(path, "r")
    lignes = fichier.read()
    fichier.close()
    lignes = lignes.split("\n")
    i = 0
    for ligne in lignes:
        if ligne != "":
            if ligne[2] == "\t":
                cheval = ligne[3:].strip()
            elif ligne[1] == "\t":
                cheval = ligne[2:].strip()
            else:
                cheval = ligne.strip()
            liste[cheval] = [i, planning.nb_heure(cheval)]
            i += 1
    return liste


def lire_fichier_cavalier(jour):
    liste = []
    liste_eleve = {}
    fichier = open(path_parametre + "liste_cavalier_" + jour + ".txt", "r")
    lignes = fichier.read()
    fichier.close()

    lignes = lignes.split("\n")
    for ligne in lignes:
        if ligne != "":
            if "\\Fin fichier/" in ligne:
                return liste_eleve
            if "\\heure/" in ligne:
                liste_eleve[ligne[7:].strip()] = liste[:]
                liste.clear()
            else:
                nom, numero = ligne.strip().split("/")
                liste.append([nom, int(numero)])


def para_inserer_listebox(data):
    heure_trier = list(data.keys())
    para_listebox_heure.delete(0, END)
    for i in heure_trier:
        para_listebox_heure.insert(END, i)


def mode_parametre():
    interface_paramete()


def mode_default():
    interface_default()


def visualiser_fichier_cavalier(data):
    heure_trier = list(data.keys())
    txt = ""
    for heure in heure_trier:
        for eleve in data[heure]:
            txt += eleve[0] + "/" + str(eleve[1]) + "\r\n"
        txt += "\\heure/" + heure + "\r\n"
    txt += "\\Fin fichier/"
    para_visu_fichier.config(state="normal")
    para_visu_fichier.delete("1.0", END)
    para_visu_fichier.insert(END, txt)
    para_visu_fichier.config(state="disabled")


def cmp_heure(d):
    if d[1].isdigit():
        heure = d[:2]
        if d[3] == "3":
            heure += "5"
        else:
            heure += "0"
    elif d[1] == "H":
        heure = d[0]
        if d[2] == "3":
            heure += "5"
        else:
            heure += "0"
    else:
        heure = d[0]
    return int(heure)


def cmp_heure_semaine(d):
    heure, jour = d.split()
    # print(heure, jour)
    logger.debug("heure %s jour %s", heure, jour)
    jours = {
        "LUNDI": 1,
        "MARDI": 2,
        "MERCREDI": 3,
        "JEUDI": 4,
        "VENDREDI": 5,
        "SAMEDI": 6,
        "DIMANCHE": 7,
    }
    # print(jours[jour], heure)
    logger.debug("jours[jour] %s heure %s", jours[jour], heure)
    return int(jours[jour]), heure


def prochain_jour(semaine, jour_actuel):
    jours_de_la_semaine = [
        "lundi",
        "mardi",
        "mercredi",
        "jeudi",
        "vendredi",
        "samedi",
        "dimanche",
    ]
    jour_cible = jours_de_la_semaine.index(semaine)

    jours_jusquau_prochain = (jour_cible - jour_actuel.weekday() + 7) % 7
    if jours_jusquau_prochain == 0:
        jours_jusquau_prochain = 7

    return jour_actuel + timedelta(days=jours_jusquau_prochain)


def add_centered_image(root, image_path, width, height):
    # Charger l'image
    original_image = Image.open(image_path)

    # Redimensionner l'image
    resized_image = original_image.resize((width, height), Image.NEAREST)
    photo = ImageTk.PhotoImage(resized_image)

    # Configurer l'image redimensionnée comme arrière-plan de la fenêtre
    image_label = tk.Label(
        root, borderwidth=0, image=photo, highlightthickness=0, bg="#b4b4b4"
    )
    image_label.image = photo
    image_label.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
    return image_label


def set_background(root, image_path):
    # Charger l'image
    original_image = Image.open(image_path)

    # Obtenir la taille de l'écran
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # Redimensionner l'image pour s'adapter à l'écran
    resized_image = original_image.resize((screen_width, screen_height), Image.NEAREST)

    photo = ImageTk.PhotoImage(resized_image)

    # Configurer l'image redimensionnée comme arrière-plan de la fenêtre
    background_label = tk.Label(root, image=photo, bg="#b4b4b4")
    background_label.image = photo
    background_label.place(x=0, y=0, relwidth=1, relheight=1)
    return background_label, photo


def nouveau_fichier():
    global pop
    pop = Toplevel(window)
    pop.title("Creation de excel")
    pop.geometry("350x150")

    date_actuelle = datetime.now()
    v = StringVar()
    date = StringVar()

    # Vérifier si aujourd'hui est déjà un mercredi
    if (
        date_actuelle.weekday() == 1
    ):  # Le code de semaine pour mercredi est 2 (0 pour lundi, 1 pour mardi, etc.)
        prochains_mardis = [date_actuelle]
        prochain_mardi = date_actuelle
    else:
        # Trouver le prochain mardi
        prochain_mardi = prochain_jour("mardi", date_actuelle)

        # Ajouter les trois prochains mardis au tableau
        prochains_mardis = [prochain_mardi]
    for _ in range(2):
        prochain_mardi = prochain_jour("mardi", prochain_mardi)
        prochains_mardis.append(prochain_mardi)

    # Vérifier si aujourd'hui est déjà un mercredi
    if (
        date_actuelle.weekday() == 2
    ):  # Le code de semaine pour mercredi est 2 (0 pour lundi, 1 pour mardi, etc.)
        prochains_mercredis = [date_actuelle]
        prochain_mercredi = date_actuelle
    else:
        # Trouver le prochain mercredi
        prochain_mercredi = prochain_jour("mercredi", date_actuelle)

        # Ajouter les trois prochains mercredis au tableau
        prochains_mercredis = [prochain_mercredi]
    for _ in range(2):
        prochain_mercredi = prochain_jour("mercredi", prochain_mercredi)
        prochains_mercredis.append(prochain_mercredi)

    if (
        date_actuelle.weekday() == 5
    ):  # Le code de semaine pour mercredi est 2 (0 pour lundi, 1 pour mardi, etc.)
        prochains_samedis = [date_actuelle]
        prochain_samedi = date_actuelle
    else:
        # Trouver le prochain mercredi
        prochain_samedi = prochain_jour("samedi", date_actuelle)

        # Ajouter les trois prochains mercredis au tableau
        prochains_samedis = [prochain_samedi]
    for _ in range(2):
        prochain_samedi = prochain_jour("samedi", prochain_samedi)
        prochains_samedis.append(prochain_samedi)

    pop_r1_1 = Radiobutton(
        pop,
        text=prochains_samedis[0].strftime("%d-%m-%Y"),
        variable=date,
        value=prochains_samedis[0].strftime("%d-%m-%Y"),
    )
    pop_r1_2 = Radiobutton(
        pop,
        text=prochains_samedis[1].strftime("%d-%m-%Y"),
        variable=date,
        value=prochains_samedis[1].strftime("%d-%m-%Y"),
    )
    pop_r1_3 = Radiobutton(
        pop,
        text=prochains_samedis[2].strftime("%d-%m-%Y"),
        variable=date,
        value=prochains_samedis[2].strftime("%d-%m-%Y"),
    )

    pop_r2_1 = Radiobutton(
        pop,
        text=prochains_mercredis[0].strftime("%d-%m-%Y"),
        variable=date,
        value=prochains_mercredis[0].strftime("%d-%m-%Y"),
    )
    pop_r2_2 = Radiobutton(
        pop,
        text=prochains_mercredis[1].strftime("%d-%m-%Y"),
        variable=date,
        value=prochains_mercredis[1].strftime("%d-%m-%Y"),
    )
    pop_r2_3 = Radiobutton(
        pop,
        text=prochains_mercredis[2].strftime("%d-%m-%Y"),
        variable=date,
        value=prochains_mercredis[2].strftime("%d-%m-%Y"),
    )

    pop_r3_1 = Radiobutton(
        pop,
        text=prochains_mardis[0].strftime("%d-%m-%Y"),
        variable=date,
        value=prochains_mardis[0].strftime("%d-%m-%Y"),
    )
    pop_r3_2 = Radiobutton(
        pop,
        text=prochains_mardis[1].strftime("%d-%m-%Y"),
        variable=date,
        value=prochains_mardis[1].strftime("%d-%m-%Y"),
    )
    pop_r3_3 = Radiobutton(
        pop,
        text=prochains_mardis[2].strftime("%d-%m-%Y"),
        variable=date,
        value=prochains_mardis[2].strftime("%d-%m-%Y"),
    )
    pop_r3_2 = Radiobutton(
        pop,
        text=prochains_mardis[1].strftime("%d-%m-%Y"),
        variable=date,
        value=prochains_mardis[1].strftime("%d-%m-%Y"),
    )
    pop_r3_3 = Radiobutton(
        pop,
        text=prochains_mardis[2].strftime("%d-%m-%Y"),
        variable=date,
        value=prochains_mardis[2].strftime("%d-%m-%Y"),
    )

    def affichejoursam():
        pop_r2_1.place_forget()
        pop_r2_2.place_forget()
        pop_r2_3.place_forget()
        pop_r3_1.place_forget()
        pop_r3_2.place_forget()
        pop_r3_3.place_forget()

        pop_r1_1.place(y=40, x=80)
        pop_r1_2.place(y=60, x=80)
        pop_r1_3.place(y=80, x=80)

    def affichejoursmer():
        pop_r1_1.place_forget()
        pop_r1_2.place_forget()
        pop_r1_3.place_forget()
        pop_r3_1.place_forget()
        pop_r3_2.place_forget()
        pop_r3_3.place_forget()
        # Vérifier si aujourd'hui est déjà un mercredi

        pop_r2_1.place(y=40, x=80)
        pop_r2_2.place(y=60, x=80)
        pop_r2_3.place(y=80, x=80)

    def affichejourmar():
        pop_r2_1.place_forget()
        pop_r2_2.place_forget()
        pop_r2_3.place_forget()
        pop_r1_1.place_forget()
        pop_r1_2.place_forget()
        pop_r1_3.place_forget()

        pop_r3_1.place(y=40, x=80)
        pop_r3_2.place(y=60, x=80)
        pop_r3_3.place(y=80, x=80)

    pop_label = Label(
        pop, text="Voulez-vous créer un fichier pour mercredi ou samedi ?"
    )
    pop_label.place(x=40, y=20)

    v.set("Mercredi")  # initialiser
    pop_r1 = Radiobutton(
        pop, text="Mercredi", variable=v, value="mercredi", command=affichejoursmer
    )
    pop_r1.place(x=10, y=40)
    pop_r2 = Radiobutton(
        pop, text="Samedi", variable=v, value="samedi", command=affichejoursam
    )
    pop_r2.place(x=10, y=60)

    pop_r3 = Radiobutton(
        pop, text="Semaine", variable=v, value="semaine", command=affichejourmar
    )
    pop_r3.place(x=10, y=80)

    def choix_date():
        pop.destroy()
        workbook = Workbook()
        name = askdirectory()
        workbook.save(name + "/liste " + v.get() + " " + date.get() + ".xlsx")

        recup_donne()

    pop_valider = Button(pop, text="Valider", command=choix_date)
    pop_valider.place(x=300, y=120)


def image(root, image_path, width, height):
    original_image = Image.open(image_path)

    # Redimensionner l'image
    resized_image = original_image.resize((width, height), Image.NEAREST)
    photo = ImageTk.PhotoImage(resized_image)

    # Configurer l'image redimensionnée comme arrière-plan de la fenêtre
    image_label = tk.Label(
        root, borderwidth=0, image=photo, highlightthickness=0, bg="#b4b4b4"
    )
    image_label.image = photo
    return image_label


def lire_parametre():
    dict_eleve = {
        "Semaine": lire_fichier_cavalier("semaine"),
        "Mercredi": lire_fichier_cavalier("mercredi"),
        "Samedi": lire_fichier_cavalier("samedi"),
    }
    dict_cheval = {
        "Semaine": lire_fichier_chevaux(path_cheval_semaine),
        "Samedi": lire_fichier_chevaux(path_cheval),
    }
    dict_cheval["Mercredi"] = dict_cheval["Samedi"]
    # print("eleve : ",dict_eleve)
    # print("\n\n\n\n")
    # print("cheval : ",dict_cheval)
    logger.debug("dict_eleve %s", dict_eleve)
    logger.debug("dict_cheval %s", dict_cheval)

    return dict_eleve, dict_cheval


def mettre_a_jour():
    global dict_eleve, dict_cheval, mail, moniteur
    dict_eleve, dict_cheval = lire_parametre()

    mail, moniteur = get_mail()
    if len(mail) > 0:
        para_input_moniteur.delete(0, END)
        para_input_moniteur.insert(0, mail[0])
    if len(mail) > 1:
        para_input_mail.delete(0, END)
        para_input_mail.insert(0, mail[1])


def importer_param():
    err = False
    try:
        chemin = askopenfilename()
        dezipper(chemin, path_parametre, suppr_rep_destination=False)
        mettre_a_jour()
    except Exception as e:
        err = True
        messagebox.showerror(
            "Erreur", f"Erreur lors de l'importation des paramètres : {e}."
        )
        logger.error(
            "Erreur lors de l'importation des paramètres : %s", e, exc_info=True
        )
    if not err:
        if parajour != "":
            liste_cheval = []
            for chevali in dict_cheval[parajour]:
                liste_cheval.append([dict_cheval[parajour][chevali][0], chevali])
            remplirlisteboxcheval(liste_cheval)
            remplirlisteboxeleve()
            para_inserer_listebox(dict_eleve[parajour])
            visualiser_fichier_cavalier(dict_eleve[parajour])
        messagebox.showinfo(
            "Importation de paramètres", "Les paramètres ont été importés avec succès !"
        )


def exporter_param():
    def verif_case_valid():
        # Définir le chemin du dossier de destination
        dossier_export = "parametre_exporter/"

        # Créer le dossier s'il n'existe pas
        os.makedirs(dossier_export, exist_ok=True)
        # Get the selected checkboxes
        selected_checkboxes = [
            files[i] for i, var in enumerate(vars_checkboxes2) if var.get() == 1
        ]

        for checkbox in selected_checkboxes:
            fichier = "parametre/" + checkbox
            # print(checkbox)
            logger.debug("checkbox %s", checkbox)
            if os.path.exists(fichier):  # Vérifier si le fichier existe
                shutil.copy(fichier, dossier_export)
        pop.destroy()

        def envoyer_param():
            selected = [
                mail[i] for i, var in enumerate(vars_checkboxes) if var.get() == 1
            ]
            # print("Moniteurs sélectionnés :", selected)  # Remplace par l'action souhaitée
            logger.debug("Moniteurs sélectionnés %s", selected)
            for moniteur in selected:
                try:
                    nom_zip = "parametre.zip"
                    chemin = zip_fichiers(dossier_export, nom_zip)
                    envoyer_email(
                        moniteur, chemin, nom_zip, "exportation des parametres", mail
                    )
                except Exception as e:
                    messagebox.showerror(
                        "Erreur", f"Erreur lors de l'envoi du planning : {e}"
                    )
                    logger.error(
                        "Erreur lors de l'envoi du planning : %s", e, exc_info=True
                    )
            shutil.rmtree(dossier_export)
            pop2.destroy()

        vars_checkboxes = []
        try:
            pop2 = Toplevel(window)
            pop2.title("Sélection des moniteurs destinataires du mail")
            pop2.geometry("550x120")

            # Create buttons for each file
            # Create checkboxes for each file
            checkboxes = []
            for i, file in enumerate(moniteur):
                var = tk.IntVar()
                checkbox = tk.Checkbutton(pop2, text=file, variable=var)
                checkboxes.append(checkbox)
                vars_checkboxes.append(var)

            # Place the checkboxes in a grid
            numx = 0
            numy = 0
            num = 0
            for i, checkbox in enumerate(checkboxes):
                checkbox.place(x=0 + numx, y=0 + numy)
                numy += 20
                num += 1
                if num > 4:
                    num = 0
                    numy = 0
                    numx += 200
            # Create buttons for "Annuler" and "Valider"
            button_annuler = tk.Button(pop2, text="Annuler", command=pop2.destroy)
            button_valider = tk.Button(pop2, text="Valider", command=envoyer_param)

            # Place the buttons at the bottom right of the window
            button_annuler.place(x=450, y=90)
            button_valider.place(x=500, y=90)
            window.wait_window(pop2)

        except Exception as e:
            messagebox.showerror(
                "Erreur", f"Erreur lors de l'exportation des paramètres : {e}."
            )
            logger.error(
                "Erreur lors de l'exportation des paramètres : %s", e, exc_info=True
            )

    err = False
    try:
        files = os.listdir(path_parametre)
        pop = Toplevel(window)
        pop.title("Choix des paramètres à exporter.")
        pop.geometry("550x120")

        # Create buttons for each file
        # Create checkboxes for each file
        vars_checkboxes2 = []
        checkboxes = []
        for i, file in enumerate(files):
            var = tk.IntVar()
            checkbox = tk.Checkbutton(pop, text=file, variable=var)
            checkboxes.append(checkbox)
            vars_checkboxes2.append(var)

        # Place the checkboxes in a grid
        numx = 0
        numy = 0
        num = 0
        for i, checkbox in enumerate(checkboxes):
            checkbox.place(x=0 + numx, y=0 + numy)
            numy += 20
            num += 1
            if num > 4:
                num = 0
                numy = 0
                numx += 200
        # Create buttons for "Annuler" and "Valider"
        button_annuler = tk.Button(pop, text="Annuler", command=pop.destroy)
        button_valider = tk.Button(pop, text="Valider", command=verif_case_valid)

        # Place the buttons at the bottom right of the window
        button_annuler.place(x=450, y=90)
        button_valider.place(x=500, y=90)
        window.wait_window(pop)
    except Exception as e:
        err = True
        messagebox.showerror(
            "Erreur", f"Erreur lors de l'exportation des paramètres : {e}."
        )
        logger.error(
            "Erreur lors de l'exportation des paramètres : %s", e, exc_info=True
        )
    if not err:
        messagebox.showinfo(
            "Exportation de paramètres", "Les paramètres ont été exportés avec succès !"
        )
        logger.debug("Exportation de paramètres réussie.")


def on_enter_pressed(event):
    if str(window.focus_get()) == ".!entry":
        ajouter_rattrapage()
    elif str(window.focus_get()) == ".!entry2":
        ajouter_theme()
    elif str(window.focus_get()) == ".!entry3" or str(window.focus_get()) == ".!entry4":
        add_cheval()
    elif str(window.focus_get()) == ".!entry5":
        add_heure()
    elif str(window.focus_get()) == ".!entry6" or str(window.focus_get()) == ".!entry9":
        add_eleve()
    elif title_label.place_info():
        ajouter()

    focused_widget = window.focus_get()
    # print(str(focused_widget)[1:])
    logger.debug("focused_widget %s", str(focused_widget)[1:])


def on_delete_pressed(event):
    if str(window.focus_get())[1:] == "para_listebox_chevaux":
        suppr_cheval()
    elif str(window.focus_get())[1:] == "para_listebox_eleve":
        suppr_eleve()
    elif str(window.focus_get())[1:] == "para_listebox_heure":
        suppr_heure(dict_eleve, heure)
    elif title_label.place_info():
        supprimer()


def on_right_pressed(event):
    heure_suivant()


def on_left_pressed(event):
    heure_precedant()


def on_crtls_pressed(event):
    if title_label.place_info():
        ecrire_fichier()
    elif para_image1.place_info():
        para_enregistrer()


def fusion():
    path = askopenfilename()
    dict_planning, chevaux, heure, planning_theme_fusion = recupperation_excel(path)
    plan = planning.planning + dict_planning
    # plan = []
    # print("plan",plan)
    logger.debug("plan %s", plan)
    for cellule in planning.planning:
        for cell in dict_planning:
            if cellule[0] == cell[0] and cellule[1] == cell[1]:
                plan.remove(cellule)
                break

    planning.set_planning(plan)
    ecrire_fichier()
    rafraichir()
    # ajoutcheval()
    # changer_heure()
    # affichage_txt(jour, planning)
    # plan.sort
    # # print("plantriée",plan)


# Importation des modules
cellule = Cellule()  # Création d'une instance de la classe Cellule
planning = Planning()  # Création d'une instance de la classe Planning
jour = Jour()  # Création d'une instance de la classe Jour
dict_eleve, dict_cheval = lire_parametre()

version = 1.84  # Version actuelle du programme
user = get_personne()
logger.user_filter.set_user(user)
# print(user)
logger.debug("user %s", user)
mail, moniteur = get_mail()

bg_button = "#8abd45"  # Couleur de fond des boutons
bg_titre = "#568A03"  # Couleur de fond des titres

fg_button = "#000000"  # Couleur du texte des boutons
fg_titre = "#FFFFFF"  # Couleur du texte des titres


# Création de l'interface utilisateur
window = tk.Tk()  # Création de la fenêtre principale
window.title("Planning")  # Titre de la fenêtre
window.attributes("-fullscreen", True)  # Affichage en mode plein écran
window.update()
wx = window.winfo_width()
wy = window.winfo_height()
# Calculer les proportions pour la nouvelle taille de l'écran
proportion_x = wx / 1536
proportion_y = wy / 864
# print("taille de fenetre", wx, wy, proportion_x, proportion_y)
logger.debug("taille de fenetre %s %s", wx, wy)
window.bind("<Return>", on_enter_pressed)
window.bind("<Right>", on_right_pressed)
window.bind("<Left>", on_left_pressed)
window.bind("<Control-s>", on_crtls_pressed)
window.bind("<Control-S>", on_crtls_pressed)
window.bind("<Delete>", on_delete_pressed)
set_background(window, path_image + "image_fond.png")

widgets_principaux = []

widgets_parametre = []

ancient_nom = ""


image_label = add_centered_image(window, path_image + "logo.png", 169 * 4, 166 * 4)

para_image1 = image(
    window,
    path_image + "image1.png",
    int(2388 / 5 * proportion_x),
    int(1668 / 5 * proportion_y),
)
image1 = image(
    window,
    path_image + "image1.png",
    int(2388 / 8.5 * proportion_x),
    int(1668 / 8.5 * proportion_y),
)
image2 = image(
    window,
    path_image + "image2.png",
    int(2388 / 8.5 * proportion_x),
    int(1668 / 8.5 * proportion_y),
)
image3 = image(
    window,
    path_image + "image3.png",
    int(2388 / 8.5 * proportion_x),
    int(1668 / 8.5 * proportion_y),
)
image4 = image(
    window, path_image + "image4.png", int(2388 / 7 * proportion_x), int(1668 / 7)
)

info = image(
    window,
    path_image + "info2.png",
    int(300 / 10 * proportion_x),
    int(300 / 10 * proportion_y),
)

info2 = image(
    window,
    path_image + "info2.png",
    int(300 / 10 * proportion_x),
    int(300 / 10 * proportion_y),
)
label_version = tk.Label(window, text="Version " + str(version), bg="#b4b4b4")
label_version.place(x=int(1395 * proportion_x), y=int(780 * proportion_y))


def on_info_clicked(event):
    pop = Toplevel(window)
    pop.title("information couleurs de chevaux")
    pop.geometry("550x120")

    label1 = tk.Label(
        pop,
        text="KABILE",
        font=("Corbel", int(13 * proportion_x)),
        bg="violet",
        justify=LEFT,
    )
    label2 = tk.Label(
        pop,
        text="VIOLETTE",
        font=("Corbel", int(13 * proportion_x)),
        bg="red",
        justify=LEFT,
    )
    label3 = tk.Label(
        pop,
        text="KID",
        font=("Corbel", int(13 * proportion_x)),
        bg="orange",
        justify=LEFT,
    )
    label4 = tk.Label(
        pop,
        text="SAMOURAI",
        font=("Corbel", int(13 * proportion_x)),
        bg="yellow",
        justify=LEFT,
    )

    label1_1 = tk.Label(
        pop,
        text=" Violet : ce cheval est déjà pris durant cette heure de la journée.",
        font=("Corbel", int(13 * proportion_x)),
        justify=LEFT,
    )
    label2_1 = tk.Label(
        pop,
        text=" Rouge : ce cheval a été monté par cet élève la semaine dernière.",
        font=("Corbel", int(13 * proportion_x)),
        justify=LEFT,
    )
    label3_1 = tk.Label(
        pop,
        text=" Orange : ce cheval a été monté par cet élève il y a deux semaines.",
        font=("Corbel", int(13 * proportion_x)),
        justify=LEFT,
    )
    label4_1 = tk.Label(
        pop,
        text=" Jaune : ce cheval a été monté par cet élève il y a trois semaines.",
        font=("Corbel", int(13 * proportion_x)),
        justify=LEFT,
    )

    # Create 8 labels in a 2x4 configuration
    labels = [label1, label2, label3, label4, label1_1, label2_1, label3_1, label4_1]

    # Place the labels in a 2x4 grid
    for i in range(4):
        for j in range(2):
            labels[i + j * 4].grid(row=i, column=j)


def on_info2_clicked(event):
    pop = Toplevel(window)
    pop.title("Information sur les couleurs des chevaux.")
    pop.geometry("500x150")

    label1 = tk.Label(
        pop,
        text="Rafraîchir permet de prendre en compte les ajouts faits sur l'Excel directement.",
        font=("Corbel", int(13 * proportion_x)),
    )
    label2 = tk.Label(
        pop,
        text="Word permet de créer un document Word contenant le récapitulatif des cours.",
        font=("Corbel", int(13 * proportion_x)),
    )
    label3 = tk.Label(
        pop,
        text="Mail permet d'envoyer la liste aux autres moniteurs.",
        font=("Corbel", int(13 * proportion_x)),
    )
    label4 = tk.Label(
        pop,
        text="Fusion permet de fusionner deux fichiers Excel sans prendre en compte les erreurs possibles.",
        font=("Corbel", int(13 * proportion_x)),
    )

    # Create 8 labels in a 2x4 configuration
    labels = [label1, label2, label3, label4]

    # Place the labels in a 2x4 grid
    for i in range(4):
        labels[i].grid(row=i, column=0)


info.bind("<Button-1>", on_info_clicked)
info2.bind("<Button-1>", on_info2_clicked)
# Définition des variables de contrôle
varheure = StringVar()
varjour = StringVar()
varcavalier = StringVar()
varcheval = StringVar()
varajout = StringVar()
varheure_cheval = StringVar()
varcavalier1 = StringVar()
varcavalier2 = StringVar()
varsemaine1 = StringVar()
varsemaine2 = StringVar()
varsemaine3 = StringVar()
theme1 = StringVar()
theme2 = StringVar()
theme3 = StringVar()
theme = StringVar()
v = IntVar()
user_var = StringVar()
user_var.set(user)

label_jour = tk.Label(
    window,
    textvariable=varjour,
    bg="#b4b4b4",
    font=("Comic Sans MS", int(15 * proportion_x)),
)

label_heure = tk.Label(window, textvariable=varheure, bg="#b4b4b4")

label_user = tk.Label(
    window,
    textvariable=user_var,
    font=("Comic Sans MS", int(15 * proportion_x)),
    bg="#b4b4b4",
)

# Création d'une étiquette pour le titre
title_label = tk.Label(
    window,
    text="GESTION PLANNING",
    font=("Comic Sans MS", int(17 * proportion_x)),
    bg="#b4b4b4",
)

# Boutons pour avancer et reculer dans les heures
boutton_avancer_heure = tk.Button(
    window,
    width=8,
    fg=fg_button,
    bg=bg_button,
    text="precedent",
    command=heure_precedant,
)

boutton_reculer_heure = tk.Button(
    window, width=8, fg=fg_button, bg=bg_button, text="suivant", command=heure_suivant
)

# Étiquettes pour afficher les informations du cavalier
label_cavalier = tk.Label(
    window,
    text="INFOS CAVALIER",
    font=("Corbel", int(14 * proportion_x)),
    bg=bg_titre,
    fg=fg_titre,
)

label_cavalier2 = tk.Label(
    window,
    textvariable=varsemaine1,
    font=("ComicsansMS", int(13 * proportion_x)),
    bg="#b4b4b4",
)

label_cavalier3 = tk.Label(
    window,
    textvariable=varcavalier,
    font=("Corbel", int(13 * proportion_x)),
    bg="#b4b4b4",
)

label_cavalier6 = tk.Label(
    window,
    textvariable=varsemaine2,
    font=("ComicsansMS", int(13 * proportion_x)),
    bg="#b4b4b4",
)

label_cavalier4 = tk.Label(
    window,
    textvariable=varcavalier1,
    font=("Corbel", int(13 * proportion_x)),
    bg="#b4b4b4",
)

label_cavalier7 = tk.Label(
    window,
    textvariable=varsemaine3,
    font=("ComicsansMS", int(13 * proportion_x)),
    bg="#b4b4b4",
)
label_cavalier5 = tk.Label(
    window,
    textvariable=varcavalier2,
    font=("Corbel", int(13 * proportion_x)),
    bg="#b4b4b4",
)


def correction():
    global dernier_cheval
    plan = []
    plan.append((cellule.heure, cellule.cheval, cellule.eleve))
    for cel in planning.ancien_planning:
        if cel[0] == cellule.heure and cel[2] == cellule.eleve:
            pass
        else:
            plan.append(cel)
    planning.set_ancien_planning(plan)
    # print(planning.ancien_planning)
    # print(planning.liste_heure)
    logger.debug("planning.ancien_planning %s", planning.ancien_planning)
    logger.debug("planning.liste_heure %s", planning.liste_heure)

    workbook = load_workbook(ancient_nom)
    sheet = workbook.active
    if elevecarte is True and varcavalier.get() == "cheval":
        unesessionmoins(cellule.eleve, cellule.heure)

    heure_non_trier = list(dict_eleve[jour.j].keys())
    colonnes = len(heure_non_trier) + 2

    if cellule.heure not in planning.liste_heure:
        planning.set_liste_eleve(dict_eleve[jour.j].copy())

    for ind in range(1, len(sheet["A"]) + 1):
        for colonne in range(1, colonnes):
            if sheet.cell(3, colonne).value == cellule.heure:
                if sheet.cell(ind, 1).value == dernier_cheval:
                    sheet.cell(ind, colonne).value = None
                if sheet.cell(ind, 1).value == cellule.cheval:
                    # print(ind,colonne,cellule.eleve)
                    sheet.cell(ind, colonne).value = cellule.eleve
                    dernier_cheval = cellule.cheval

    err = workbook.save(ancient_nom)
    if err is None:
        ajout_historique("correction", (cellule.heure, cellule.cheval, cellule.eleve))
        varcavalier.set(cellule.cheval)


def absent():
    global dernier_cheval
    workbook = load_workbook(ancient_nom)
    sheet = workbook.active
    err = True
    for cel in planning.ancien_planning:
        if cel[0] == cellule.heure and cel[2] == cellule.eleve:
            planning.ancien_planning.remove(cel)
            err = False
    if not err:
        # print(planning.ancien_planning)
        logger.debug("planning.ancien_planning %s", planning.ancien_planning)
        heure_non_trier = list(dict_eleve[jour.j].keys())
        colonnes = len(heure_non_trier) + 2
        # print("dercheval",dernier_cheval)
        logger.debug("dernier_cheval %s", dernier_cheval)
        if elevecarte is True:
            unesessionplus(cellule.eleve, cellule.heure)
        for ind in range(1, len(sheet["A"]) + 1):
            if sheet.cell(ind, 1).value == dernier_cheval:
                for colonne in range(1, colonnes):
                    if sheet.cell(3, colonne).value == cellule.heure:
                        sheet.cell(ind, colonne).value = None
                        dernier_cheval = ""
        err = workbook.save(ancient_nom)
        if err is None:
            ajout_historique("absence", (cellule.heure, cellule.eleve))
            varcavalier.set("cheval")
    else:
        messagebox.showerror(
            "Erreur",
            "Cet élève avec ce cheval à cette heure-ci n'est pas dans la liste du"
            + ancient_nom,
        )
        logger.warning(
            "Cet élève avec ce cheval à cette heure-ci n'est pas dans la liste du %s",
            ancient_nom,
        )


dernier_cheval = ""

boutton_absent = tk.Button(
    window,
    fg=fg_button,
    bg=bg_button,
    height=1,
    width=int(4 * proportion_x),
    text="ABS",
    command=absent,
    borderwidth=2,
)
boutton_correction = tk.Button(
    window, fg=fg_button, bg=bg_button, height=1, text="correction", command=correction
)
# Initialisation des variables de contrôle
varcavalier.set("cheval")
varcavalier1.set("cheval1")
varcavalier2.set("cheval2")

# Liste déroulante pour les élèves
eleve_listbox = tk.Listbox(window, name="eleve_listbox", yscrollcommand=True)

eleve_rattrapage = tk.Entry(window)


label_eleve_rattrapage = tk.Label(
    window, text="Ajouter un nom", font=("Corbel", int(13 * proportion_x)), bg="#b4b4b4"
)
boutton_eleve_rattrapage = tk.Button(
    window,
    width=int(8 * proportion_x),
    fg=fg_button,
    bg=bg_button,
    text="rattrapage",
    command=ajouter_rattrapage,
)

theme_entry = tk.Entry(window)

label_theme = tk.Label(
    window,
    text="Ajouter un theme",
    font=("Corbel", int(13 * proportion_x)),
    bg="#b4b4b4",
)
boutton_theme = tk.Button(
    window,
    width=int(14 * proportion_x),
    fg=fg_button,
    bg=bg_button,
    text="ajout du theme",
    command=ajouter_theme,
)
label_theme_actuelle = tk.Label(
    window, textvariable=theme, font=("Corbel", int(13 * proportion_x)), bg="#b4b4b4"
)
label_theme_avant1 = tk.Label(
    window, textvariable=theme1, font=("Corbel", int(13 * proportion_x)), bg="#b4b4b4"
)
label_theme_avant2 = tk.Label(
    window, textvariable=theme2, font=("Corbel", int(13 * proportion_x)), bg="#b4b4b4"
)
label_theme_avant3 = tk.Label(
    window, textvariable=theme3, font=("Corbel", int(13 * proportion_x)), bg="#b4b4b4"
)
theme1.set("theme1")
theme2.set("theme2")
theme3.set("theme3")


# Fonction appelée lorsqu'un élément est sélectionné dans la liste des élèves
def items_selected(event):
    global dernier_cheval, elevecarte
    # Indices des éléments sélectionnés
    selected_indices = eleve_listbox.curselection()
    if len(selected_indices) > 0:
        eleve = eleve_listbox.get(selected_indices)

        cellule.set_eleve(eleve, selected_indices[0])
        elevecarte = False
        if isinstance(cellule.eleve[1], int):
            cellule.eleve = cellule.eleve[0]
            elevecarte = True
        # print(cellule.eleve)
        logger.debug("cellule.eleve %s", cellule.eleve)
        ancient_cheval = planning.ancient_cheval_de(cellule.eleve, cellule.heure)

        # Mise à jour des étiquettes des chevaux associés
        if ancient_cheval[0][1] != "":
            varcavalier.set(ancient_cheval[0][0])
            dernier_cheval = ancient_cheval[0][0]
        else:
            varcavalier.set("cheval")
        if len(ancient_cheval) >= 2:
            varcavalier1.set(ancient_cheval[1][0])
        else:
            varcavalier1.set("cheval1")
        if len(ancient_cheval) >= 3:
            varcavalier2.set(ancient_cheval[2][0])
        else:
            varcavalier2.set("cheval2")
        colorier()
        colorier_ancient_chevaux(ancient_cheval)
        colorier_chevaux()
        for tup in planning.planning:
            if (cellule.heure, cellule.eleve) == (tup[0], tup[2]) and tup[
                1
            ] in planning.cheval:
                cellule.set_cheval(tup[1], planning.index_cheval(tup[1]))
                varheure_cheval.set(f"HEURE DE TRAVAIL DE: {cellule.cheval}")
                inserer_liste_de_travaille()
        varajout.set(cellule.getCellule())


# Association de la fonction à l'événement de relâchement du bouton de la souris
eleve_listbox.bind("<<ListboxSelect>>", items_selected)

# Liste déroulante pour les chevaux
cheval_listbox = tk.Listbox(
    window, name="cheval_listbox", height=int(47 * proportion_y)
)

# Fonction appelée lorsqu'un élément est sélectionné dans la liste des chevaux


def items_selected_cheval(event):
    # Indices des éléments sélectionnés
    selected_indices = cheval_listbox.curselection()
    if len(selected_indices) > 0:
        cheval = cheval_listbox.get(selected_indices)
        cellule.set_cheval(cheval[1], selected_indices)
        varheure_cheval.set(f"HEURE DE TRAVAIL DE: {cellule.cheval}")
        inserer_liste_de_travaille()
        varajout.set(cellule.getCellule())


# Association de la fonction à l'événement de relâchement du bouton de la souris
cheval_listbox.bind("<<ListboxSelect>>", items_selected_cheval)

# Zone de texte pour afficher le planning
visu_fichier = tk.Text(
    window, width=int(70 * proportion_x), height=int(24 * proportion_y)
)
visu_fichier.config(state="disabled")

label_visu_fichier = tk.Label(
    window,
    text="PREVISUALISATION",
    font=("Corbel", int(14 * proportion_x)),
    bg=bg_titre,
    fg=fg_titre,
)


# Étiquette pour afficher des informations sur l'ajout
label_ajout = tk.Label(
    window, textvariable=varajout, font=int(30 * proportion_x), bg="#ffffff"
)


# Bouton pour ajouter une entrée
boutton_ajouter = tk.Button(
    window,
    text="Ajouter",
    command=ajouter,
    width=int(11 * proportion_x),
    height=int(2 * proportion_y),
    fg=fg_button,
    bg=bg_button,
)


# Bouton pour supprimer une entrée
boutton_supprimer = tk.Button(
    window,
    text="Supprimer",
    command=supprimer,
    width=int(11 * proportion_x),
    height=int(2 * proportion_y),
    fg=fg_button,
    bg=bg_button,
)


# Bouton pour enregistrer les modifications
boutton_enregistrer = tk.Button(
    window,
    text="ENREGISTRER",
    command=ecrire_fichier,
    width=12,
    font=("Helvetica", 18, "bold"),
    bg="#000000",
    fg="#ffffff",
)


# Étiquette pour afficher un message après l'enregistrement
label_enregistrer = tk.Label(
    window,
    text="Le fichier a bien été enregistré.",
    font=("Corbel", int(13 * proportion_x)),
    bg="#b4b4b4",
)  # le fichier à bien été enregistré
label_enregistrer.config(fg="#b4b4b4")

# Étiquette pour afficher l'heure de travail du cheval
label_heure_cheval = tk.Label(
    window,
    textvariable=varheure_cheval,
    font=("Corbel", int(13 * proportion_x)),
    bg=bg_titre,
    fg=fg_titre,
)


# Liste déroulante pour les heures de travail
heure_listebox = tk.Listbox(
    window,
    name="heure_listebox",
    width=int(25 * proportion_x),
    height=int(5 * proportion_y),
)


def rafraichir():
    global planning_theme
    dict_planning, cheval, heures, planning_theme = recupperation_excel(
        planning.name_fichier
    )
    elements_ajout = [
        element for element in dict_planning if element not in planning.planning
    ]
    # print(elements_ajout)
    logger.debug("elements_ajout %s", elements_ajout)
    elements_suppr = [
        element for element in planning.planning if element not in dict_planning
    ]
    # print(elements_suppr)
    logger.debug("elements_suppr %s", elements_suppr)
    eleves_carte = []
    # print("heure",heures)
    logger.debug("heure %s", heures)
    for heure in planning.liste_eleve:
        for eleve in planning.liste_eleve[heure]:
            if eleve[1] != -1:
                eleves_carte.append((heure, eleve[0]))
    # print(eleves_carte)
    logger.debug("eleves_carte %s", eleves_carte)
    for cell in elements_suppr:
        if (cell[0], cell[2]) in eleves_carte:
            # print('cessionplus',cell[2])
            logger.debug("cessionplus %s", cell[2])
            unesessionplus(cell[2], cell[0])
    for cell in elements_ajout:
        if (cell[0], cell[2]) in eleves_carte:
            # print('cessionMOINS',cell[2])
            logger.debug("cessionMOINS %s", cell[2])
            unesessionmoins(cell[2], cell[0])

    planning.set_planning(dict_planning)
    planning.set_cheval(remplir_cheval(dict_cheval[jour.j]))
    planning.set_heure(heures)
    ajoutcheval()
    changer_heure()
    affichage_txt(jour, planning)


def ouvrir_excel():
    subprocess.Popen(["start", "excel", planning.name_fichier], shell=True)


def ecrire_mail():
    def envoyer_mail():
        selected = [mail[i] for i, var in enumerate(vars_checkboxes) if var.get() == 1]
        # print("Moniteurs sélectionnés :", selected)  # Remplace par l'action souhaitée
        logger.debug("Moniteurs sélectionnés %s", selected)
        for moniteur in selected:
            try:
                envoyer_email(
                    moniteur,
                    planning.name_fichier,
                    nom_fichier[0],
                    "envoie du planning du "
                    + str(extract_date_from_filename(planning.name_fichier))[0:11],
                    mail,
                )
            except Exception as e:
                messagebox.showerror(
                    "Erreur", f"Erreur lors de l'envoi du planning : {e}"
                )
                logger.error(
                    "Erreur lors de l'envoi du planning : %s", e, exc_info=True
                )
        pop.destroy()

    err = False
    vars_checkboxes = []
    try:
        pop = Toplevel(window)
        pop.title("Sélection des moniteurs destinataires du mail")
        pop.geometry("550x120")

        # Create buttons for each file
        # Create checkboxes for each file
        checkboxes = []
        for i, file in enumerate(moniteur):
            var = tk.IntVar()
            checkbox = tk.Checkbutton(pop, text=file, variable=var)
            checkboxes.append(checkbox)
            vars_checkboxes.append(var)

        # Place the checkboxes in a grid
        numx = 0
        numy = 0
        num = 0
        for i, checkbox in enumerate(checkboxes):
            checkbox.place(x=0 + numx, y=0 + numy)
            numy += 20
            num += 1
            if num > 4:
                num = 0
                numy = 0
                numx += 200
        # Create buttons for "Annuler" and "Valider"
        button_annuler = tk.Button(pop, text="Annuler", command=pop.destroy)
        button_valider = tk.Button(pop, text="Valider", command=envoyer_mail)

        # Place the buttons at the bottom right of the window
        button_annuler.place(x=450, y=90)
        button_valider.place(x=500, y=90)
        window.wait_window(pop)

    except Exception as e:
        err = True
        messagebox.showerror("Erreur", f"Erreur lors de l'envoi du planning : {e}")
        logger.error("Erreur lors de l'envoi du planning : %s", e, exc_info=True)
    if not err:
        messagebox.showinfo(
            "Envoi du planning", "Le planning a été envoyé avec succès !"
        )
        logger.debug("Envoi du planning réussi.")


bouton_ouvrir_excel = tk.Button(
    window, text="ouvrir", fg=fg_button, bg=bg_button, command=ouvrir_excel
)

bouton_rafraichir = tk.Button(
    window, text="rafraichir", fg=fg_button, bg=bg_button, command=rafraichir
)

bouton_word = tk.Button(
    window, text="word", fg=fg_button, bg=bg_button, command=ecrire_word
)

bouton_mail = tk.Button(
    window, text="mail", fg=fg_button, bg=bg_button, command=ecrire_mail
)

bouton_fusion = tk.Button(
    window, text="fusion", fg=fg_button, bg=bg_button, command=fusion
)

# Fonction appelée lorsqu'un élément est sélectionné dans la liste des heures de travail


def items_selected_heure_cheval(event):
    # Indices des éléments sélectionnés
    selected_indices = heure_listebox.curselection()
    if len(selected_indices) > 0:
        (h, p) = heure_listebox.get(selected_indices)
        if h != cellule.heure:
            cellule.set_heure(h)
            changer_heure()
            cellule.set_eleve(p, -1)
        else:
            Nb = 0
            for i in range(0, eleve_listbox.size()):
                eleve = eleve_listbox.get(i)
                if isinstance(eleve[1], int):
                    eleve = eleve[0]
                if p == eleve:
                    cellule.set_eleve(p, Nb)
                Nb += 1
        varajout.set(cellule.getCellule())


# Association de la fonction à l'événement de relâchement du bouton de la souris
heure_listebox.bind("<<ListboxSelect>>", items_selected_heure_cheval)

# Étiquette pour afficher l'historique
label_historique = tk.Label(
    window,
    text="HISTORIQUE",
    font=("Corbel", int(13 * proportion_x)),
    bg=bg_titre,
    fg=fg_titre,
)

# Zone de texte pour afficher l'historique
historique = tk.Text(
    window, width=int(60 * proportion_x), height=int(13 * proportion_y)
)
historique.config(state="disabled")

# Création du menu
menubar = Menu(window)

# Création d'une liste déroulante pour sélectionner l'heure
listeCombo = ttk.Combobox(
    window, height=int(10 * proportion_y), width=int(40 * proportion_x)
)


# Fonction appelée lorsqu'un élément est sélectionné dans la liste déroulante


def action(event):
    select = listeCombo.get()  # Élément sélectionné dans la liste déroulante
    changement_heure(select)


listeCombo.bind("<<ComboboxSelected>>", action)


def remplir_listbox_heure():
    listeCombo.delete(0, "end")
    listeCombo["values"] = list(planning.liste_eleve)


menubar = Menu(window)

# Ajout des éléments au menu
sousmenu = Menu(menubar, tearoff=0)
sousmenu.add_command(label="parametre", command=mode_parametre)
sousmenu.add_command(label="principal", command=mode_default)


# Ajout des éléments au menu
menubar.add_command(label="Nouveau", command=nouveau_fichier)
menubar.add_command(label="Jour", command=recup_donne)
menubar.add_cascade(label="Mode", menu=sousmenu)
menubar.add_command(label="Quitter!", command=window.quit)

# Affichage du menu dans la fenêtre
window.config(menu=menubar)


# interface_paramete()

para_visu_fichier = tk.Text(
    window, width=int(70 * proportion_x), height=int(24 * proportion_y)
)
para_visu_fichier.config(state="disabled")

# Étiquette pour afficher l'historique
para_label_historique = tk.Label(
    window, text="historique", font=("Corbel", int(13 * proportion_x)), bg="#b4b4b4"
)

# Zone de texte pour afficher l'historique
para_historique = tk.Text(
    window, width=int(60 * proportion_x), height=int(13 * proportion_y)
)
para_historique.config(state="disabled")

para_listebox_heure = tk.Listbox(
    window,
    name="para_listebox_heure",
    width=int(25 * proportion_x),
    height=int(45 * proportion_y),
)

heure = ""


def items_selected_heure(event):
    global heure
    # Indices des éléments sélectionnés
    selected_indices = para_listebox_heure.curselection()
    if len(selected_indices) > 0:
        heure = para_listebox_heure.get(selected_indices)
        remplirlisteboxeleve()


# Association de la fonction à l'événement de relâchement du bouton de la souris
para_listebox_heure.bind("<<ListboxSelect>>", items_selected_heure)


para_listebox_eleve = tk.Listbox(
    window,
    name="para_listebox_eleve",
    width=int(25 * proportion_x),
    height=int(12 * proportion_y),
)


eleve = ""


def items_selected_eleve(event):
    # Indices des éléments sélectionnés
    global eleve
    selected_indices = para_listebox_eleve.curselection()
    if len(selected_indices) > 0:
        eleve = para_listebox_eleve.get(selected_indices)


# Association de la fonction à l'événement de relâchement du bouton de la souris
para_listebox_eleve.bind("<<ListboxSelect>>", items_selected_eleve)


para_listebox_chevaux = tk.Listbox(
    window,
    name="para_listebox_chevaux",
    width=int(25 * proportion_x),
    height=int(45 * proportion_y),
)

cheval = ""


def items_selected_cheval(event):
    global cheval
    # Indices des éléments sélectionnés
    selected_indices = para_listebox_chevaux.curselection()
    if len(selected_indices) > 0:
        cheval = list(para_listebox_chevaux.get(selected_indices))


para_listebox_moniteur = tk.Listbox(
    window,
    name="para_listebox_moniteur",
    width=int(40 * proportion_x),
    height=int(5 * proportion_y),
)

para_add_moniteur = tk.Button(
    window,
    text="ajouter moniteur",
    command=add_moniteur,
    width=int(18 * proportion_x),
    fg=fg_button,
    bg=bg_button,
)
para_suppr_moniteur = tk.Button(
    window,
    text="supprimer moniteur",
    command=suppr_moniteur,
    width=int(18 * proportion_x),
    fg=fg_button,
    bg=bg_button,
)

para_label_moniteur = tk.Label(
    window, text="moniteur", font=("Corbel", 13), bg="#b4b4b4"
)
para_input_moniteur = tk.Entry(window, width=int(25 * proportion_x))

para_label_mail = tk.Label(window, text="mail", font=("Corbel", 13), bg="#b4b4b4")
para_input_mail = tk.Entry(window, width=int(25 * proportion_x))


# Association de la fonction à l'événement de relâchement du bouton de la souris
para_listebox_chevaux.bind("<<ListboxSelect>>", items_selected_cheval)


def action_base():
    global parajour
    parajour = para_listeCombo.get()  # Élément sélectionné dans la liste déroulante
    liste_cheval = []
    for cheval in dict_cheval[parajour]:
        liste_cheval.append([dict_cheval[parajour][cheval][0], cheval])
    remplirlisteboxcheval(liste_cheval)
    para_inserer_listebox(dict_eleve[parajour])
    visualiser_fichier_cavalier(dict_eleve[parajour])


# Création d'une liste déroulante pour sélectionner l'heure
para_listeCombo = ttk.Combobox(window, width=int(10 * proportion_x))
para_listeCombo["values"] = ["Mercredi", "Samedi", "Semaine"]
para_listeCombo.current(0)
parajour = ""
action_base()

para_listeCombo_user = ttk.Combobox(window, width=int(10 * proportion_x))
para_listeCombo_user["values"] = moniteur
para_listeCombo_user.current(0)


def items_selected_moniteur(event):
    selected_indices = para_listebox_moniteur.curselection()
    para_input_moniteur.delete(0, tk.END)
    para_input_mail.delete(0, tk.END)
    if len(selected_indices) > 0:
        para_input_moniteur.insert(
            0, moniteur[para_listebox_moniteur.curselection()[0]]
        )
        para_input_mail.insert(0, mail[para_listebox_moniteur.curselection()[0]])


para_listebox_moniteur.bind("<<ListboxSelect>>", items_selected_moniteur)


def action(event):
    action_base()


para_input_chevaux = tk.Entry(window)
para_input_ind_chevaux = tk.Entry(window, width=3)
para_add_chevaux = tk.Button(
    window,
    text="ajouter cheval",
    command=add_cheval,
    width=int(18 * proportion_x),
    fg=fg_button,
    bg=bg_button,
)
para_suppr_chevaux = tk.Button(
    window,
    text="supprimer cheval",
    command=suppr_cheval,
    width=int(18 * proportion_x),
    fg=fg_button,
    bg=bg_button,
)

para_listeCombo.bind("<<ComboboxSelected>>", action)

para_input_heure = tk.Entry(window)

para_add_heure = tk.Button(
    window,
    text="creer heure",
    command=add_heure,
    width=int(18 * proportion_x),
    fg=fg_button,
    bg=bg_button,
)

para_suppr_heure = tk.Button(
    window,
    text="supprimer heure",
    command=lambda: suppr_heure(dict_eleve, heure),
    width=int(18 * proportion_x),
    fg=fg_button,
    bg=bg_button,
)


para_input_eleve = tk.Entry(window)

para_add_eleve = tk.Button(
    window,
    text="creer eleve",
    command=add_eleve,
    width=int(18 * proportion_x),
    fg=fg_button,
    bg=bg_button,
)

para_suppr_eleve = tk.Button(
    window,
    text="supprimer eleve",
    command=suppr_eleve,
    width=int(18 * proportion_x),
    fg=fg_button,
    bg=bg_button,
)

para_boutton_enregistrer = tk.Button(
    window,
    text="ENREGISTRER",
    command=para_enregistrer,
    width=12,
    font=("Helvetica", 18, "bold"),
    bg="#000000",
    fg="#ffffff",
)


def sauvegarder_mail():
    # if para_input_moniteur.get() and len(mail) > 0:
    #     mail[0] = para_input_moniteur.get()
    # elif para_input_moniteur.get() and len(mail) == 0:
    #     mail.append(para_input_moniteur.get())
    # if para_input_mail.get() and len(mail) > 1:
    #     mail[1] = para_input_mail.get()
    # elif para_input_mail.get() and len(mail) == 1:
    #     mail.append(para_input_mail.get())

    with open(path_mail, "w") as f:
        for i in range(len(moniteur)):
            f.write(moniteur[i] + ":" + mail[i] + "\n")


def ouvrir_excel():
    # print("parajour",parajour)
    logger.debug("parajour %s", parajour)
    if parajour == "Mercredi":
        subprocess.Popen(["start", "excel", path_Mercredi], shell=True)
    elif parajour == "Samedi":
        subprocess.Popen(["start", "excel", path_Samedi], shell=True)
    elif parajour == "Semaine":
        subprocess.Popen(["start", "excel", path_semaine], shell=True)


para_bouton_ouvrir_excel = tk.Button(
    window,
    text="ouvrir excel reference",
    fg=fg_button,
    bg=bg_button,
    command=ouvrir_excel,
)


def toggle_entry_nbcarte():
    if v.get() == 1:
        para_nbcarte.config(state=NORMAL)
    else:
        para_nbcarte.config(state=DISABLED)


para_case = Checkbutton(
    variable=v, bg="#b4b4b4", text="eleve à la carte", command=toggle_entry_nbcarte
)

para_nbcarte = Entry(window)
para_nbcarte.insert(0, "nombre de seances")
para_nbcarte.config(state=DISABLED)


def on_para_nbcarte_click(event):
    if para_nbcarte.get() == "nombre de seances":
        para_nbcarte.delete(0, tk.END)


para_nbcarte.bind("<FocusIn>", on_para_nbcarte_click)

para_bouton_importer_param = tk.Button(
    window,
    text="importer parametre",
    width=15,
    fg=fg_button,
    bg=bg_button,
    command=importer_param,
)

para_bouton_exporter_param = tk.Button(
    window,
    text="exporter parametre",
    width=15,
    fg=fg_button,
    bg=bg_button,
    command=exporter_param,
)


widgets_parametre.extend(
    [
        para_visu_fichier,
        para_bouton_importer_param,
        para_input_moniteur,
        para_bouton_exporter_param,
        para_label_moniteur,
        para_add_moniteur,
        para_suppr_moniteur,
        para_listebox_moniteur,
        para_bouton_ouvrir_excel,
        para_case,
        para_label_mail,
        para_input_mail,
        para_nbcarte,
        para_listeCombo_user,
        para_label_historique,
        para_historique,
        para_listebox_eleve,
        para_listebox_heure,
        para_listeCombo,
        para_input_heure,
        para_input_eleve,
        para_boutton_enregistrer,
        para_suppr_eleve,
        para_add_eleve,
        para_suppr_heure,
        para_add_heure,
        para_listebox_chevaux,
        para_input_chevaux,
        para_add_chevaux,
        para_suppr_chevaux,
        para_input_ind_chevaux,
        image3,
        image4,
        para_image1,
    ]
)

widgets_principaux.extend(
    [
        label_jour,
        label_heure,
        title_label,
        boutton_avancer_heure,
        boutton_reculer_heure,
        label_cavalier,
        label_cavalier2,
        label_cavalier3,
        label_cavalier6,
        label_cavalier4,
        label_cavalier5,
        label_cavalier7,
        eleve_listbox,
        cheval_listbox,
        label_ajout,
        boutton_ajouter,
        boutton_supprimer,
        visu_fichier,
        label_visu_fichier,
        boutton_enregistrer,
        bouton_rafraichir,
        bouton_ouvrir_excel,
        label_enregistrer,
        image1,
        image2,
        label_heure_cheval,
        heure_listebox,
        label_historique,
        historique,
        listeCombo,
        boutton_absent,
        boutton_correction,
        eleve_rattrapage,
        label_eleve_rattrapage,
        boutton_eleve_rattrapage,
        theme_entry,
        info,
        info2,
        bouton_fusion,
        bouton_mail,
        label_user,
        bouton_word,
        label_theme,
        boutton_theme,
        label_theme_actuelle,
        label_theme_avant1,
        label_theme_avant2,
        label_theme_avant3,
    ]
)

installateur(window, user, version)

# Lancement de la boucle principale de l'application
window.mainloop()
