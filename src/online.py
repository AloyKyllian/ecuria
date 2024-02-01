from Planning import *
from Ftp import *
from Jour import *
import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter.filedialog import askopenfilename, askdirectory
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import os
from datetime import datetime, timedelta
from PIL import Image,ImageTk
from PIL.ImageTk import PhotoImage
import subprocess


adresse_serveur = "83.113.54.154"
nom_utilisateur = "lena"
mot_de_passe = "1234"

path_cavalier_mercredi = "liste_cavalier_mercredi.txt"
path_cavalier_samedi = "liste_cavalier_samedi.txt"

def heure_precedant():
    """
    Définit l'heure précédente pour la cellule et appelle la fonction changer_heure.

    Cette fonction met à jour l'heure de la cellule en la définissant comme étant l'heure
    précédente par rapport à la liste des heures de travail du planning. Elle utilise la
    fonction changer_heure pour mettre à jour l'interface utilisateur.

    Args:
        Aucun.

    Returns:
        Aucun.
    """

    liste_heure = list(planning.liste_eleve)

    for i in range(1, len(liste_heure)):
        if liste_heure[i] == cellule.heure:
            cellule.set_heure(liste_heure[i-1])
            changer_heure()


def heure_suivant():
    """
    Définit l'heure suivante pour la cellule et appelle la fonction changer_heure.

    Cette fonction met à jour l'heure de la cellule en la définissant comme étant l'heure
    suivante par rapport à la liste des heures de travail du planning. Elle utilise la fonction
    changer_heure pour mettre à jour l'interface utilisateur.

    Args:
        Aucun.

    Returns:
        int: 0 si l'heure suivante est définie avec succès, sinon rien.
    """
    liste_heure = list(planning.liste_eleve)
    for i in range(0, len(liste_heure)-1):
        if liste_heure[i] == cellule.heure:
            cellule.set_heure(liste_heure[i+1])
            changer_heure()
            return 0


def ajouter_rattrapage():
    planning.liste_eleve[cellule.heure].append(eleve_rattrapage.get().upper())
    ajouteleve()
    
def unesessionmoins(eleve):
    for i in range(len(planning.liste_eleve[cellule.heure])):
        if planning.liste_eleve[cellule.heure][i][0] == eleve:
            print(cellule.eleve)
            planning.liste_eleve[cellule.heure][i][1] -= 1
            if planning.liste_eleve[cellule.heure][i][1] == 0:
                planning.liste_eleve[cellule.heure][i][1] = 10
            break
    ajouteleve()
    
def unesessionplus(eleve):
    for i in range(len(planning.liste_eleve[cellule.heure])):
        print(cellule.heure,i)
        if planning.liste_eleve[cellule.heure][i][0] == eleve:
            print("ici", eleve)
            planning.liste_eleve[cellule.heure][i][1] += 1
            if planning.liste_eleve[cellule.heure][i][1] > 10:
                planning.liste_eleve[cellule.heure][i][1] = 1
            break
    ajouteleve()


def ajouter():
    """
    Ajoute une cellule au planning avec des vérifications et met à jour l'interface.

    Cette fonction ajoute une cellule au planning en utilisant la méthode ajout de la classe
    Planning. Elle effectue des vérifications préliminaires et met à jour l'interface utilisateur
    en conséquence. Elle ajoute également une entrée à l'historique du planning en cas de succès.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    err = planning.ajout(cellule)
    if err == None or err == -4:
        if elevecarte == True:
            unesessionmoins(cellule.eleve)
        # print(planning.liste_eleve)
        if cellule.ind_eleve != -1:
            colorier_eleve(cellule.ind_eleve)
        ajoutuncheval(cellule.cheval, cellule.ind_cheval)
        affichage_txt(jour, planning)
        label_enregistrer.config(fg="#b4b4b4")
        inserer_liste_de_travaille()
        ajout_historique(
            "ajout", (cellule.heure, cellule.cheval, cellule.eleve))
    elif err == -1:
        msgbox = tk.messagebox.showerror(
            title="creation de fichier", message="vous n'avez pas selectionné toutes les informations necessaire à l'ajout")
    elif err == -2:
        msgbox = tk.messagebox.showerror(
            title="creation de fichier", message="Ne peux etre ajouter car ce cheval travaille deja durant cette heure")
    elif err == -3:
        msgbox = tk.messagebox.showerror(
            title="creation de fichier", message="Ne peux etre ajouter car ce cheval travaille deja 4 heure dans la journée")
    elif err == -5:
        print(f"Erreur numéro {err}, ajout annulé.")


def supprimer():
    """
    Supprime une cellule du planning avec des vérifications et met à jour l'interface.

    Cette fonction supprime une cellule du planning en utilisant la méthode supprime de la classe
    Planning. Elle effectue des vérifications préliminaires et met à jour l'interface utilisateur
    en conséquence. Elle ajoute également une entrée à l'historique du planning en cas de succès.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    err = planning.supprime(cellule)
    if err == None:
        if elevecarte == True:
            unesessionplus(cellule.eleve)
        if cellule.ind_eleve != -1:
            decolorier_eleve(cellule.ind_eleve)
        ajoutuncheval(cellule.cheval, cellule.ind_cheval)
        affichage_txt(jour, planning)
        label_enregistrer.config(fg="#b4b4b4")
        inserer_liste_de_travaille()
        ajout_historique("suppression", (cellule.heure,
                         cellule.cheval, cellule.eleve))
    else:
        msgbox = tk.messagebox.showerror(
            title="creation de fichier", message="suppression impossible vous voulez supprimez un element qui n'existe pas")


def inserer_liste_de_travaille():
    """
    Insère les heures de travail dans la listebox et met en évidence les heures travaillées.

    Cette fonction vide d'abord la listebox des heures de travail, puis insère les heures de
    travail associées au cheval de la cellule dans la listebox. Elle met également en évidence
    les heures travaillées dans la listebox.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    vider_listebox(heure_listebox)
    liste = planning.heure_travailler(cellule.cheval)
    for i in liste:
        heure_listebox.insert(tk.END, i)


def colorier_eleve(ind):
    """
    Change la couleur de fond de l'élément d'index donné dans la listebox des élèves en rouge.

    Cette fonction prend en argument l'index d'un élément dans la listebox des élèves et change
    la couleur de fond de cet élément en rouge.

    Args:
        ind (int): L'index de l'élément dans la listebox des élèves.

    Returns:
        Aucun.
    """
    eleve_listbox.itemconfig(ind, {'bg': 'red'})


def decolorier_eleve(ind):
    """
    Change la couleur de fond de l'élément d'index donné dans la listebox des élèves en blanc.

    Cette fonction prend en argument l'index d'un élément dans la listebox des élèves et change
    la couleur de fond de cet élément en blanc.

    Args:
        ind (int): L'index de l'élément dans la listebox des élèves.

    Returns:
        Aucun.
    """
    eleve_listbox.itemconfig(ind, {'bg': 'white'})


def ajout_historique(type, element):
    """
    Ajoute un élément à l'historique du planning et met à jour l'affichage dans la zone de texte.

    Cette fonction prend en argument un type d'action (ajout ou suppression) et un élément à ajouter
    à l'historique du planning. Elle met à jour l'affichage de l'historique dans la zone de texte
    correspondante.

    Args:
        type (str): Le type d'action (ajout ou suppression).
        element (tuple): L'élément à ajouter à l'historique.

    Returns:
        Aucun.
    """
    planning.append_historique(type, element)
    historique.delete('1.0', END)
    for i in planning.historique:
        historique.insert("1.0", f"{i}\r\n")


def affichage_txt(jour, planning):
    """
    Affiche le planning du jour dans la zone de texte correspondante.

    Cette fonction prend en argument un objet "jour" et un objet "planning", puis affiche le planning
    du jour dans la zone de texte correspondante.

    Args:
        jour (object): L'objet jour correspondant au jour de la semaine.
        planning (object): L'objet planning contenant les données du planning.

    Returns:
        Aucun.
    """
    visu_fichier.delete('1.0', END)
    visu_fichier.insert(END, planning.fichier(jour.j))


def vider_listebox(listebox):
    """
    Vide une listebox en supprimant tous ses éléments.

    Cette fonction prend en argument une listebox et supprime tous les éléments qu'elle contient.

    Args:
        listebox: La listebox à vider.

    Returns:
        Aucun.
    """
    if listebox.size() > 0:
        listebox.delete(0, listebox.size())


def inserer_listbox(i):
    """
    Insère un élément dans la listebox des élèves et le met en évidence s'il est présent dans le planning.

    Cette fonction prend en argument un élément "i" et l'insère dans la listebox des élèves. Si l'élément
    est présent dans le planning, il est mis en évidence en rouge.

    Args:
        i (str): L'élément à insérer dans la listebox des élèves.

    Returns:
        Aucun.
    """
    if i[1] == -1:
        i = i[0]
    eleve_listbox.insert(tk.END, i)
    try:
        entier_valeur = int(i[1])
        i = i[0]
    except ValueError:
        pass
    # if int(i[1]) <= 10 and int(i[1]) >= 0:
    #     i = i[0]
    if len(planning.planning) != 0:
        present = any((cellule.heure, i) == (tup[0], tup[2])
                      for tup in planning.planning)
        if present:
            eleve_listbox.itemconfig(tk.END, {'bg': 'red'})


def ajouteleve():
    """
    Vide la listebox des élèves et insère les éléments correspondant à l'heure de la cellule.

    Cette fonction vide d'abord la listebox des élèves, puis insère les éléments correspondant à
    l'heure de la cellule. Elle met également en évidence les élèves qui sont présents dans le planning.

    Args:
        Aucun.

    Returns:
        Aucun.
    """

    vider_listebox(eleve_listbox)
    if cellule.heure != 'heure':
        if cellule.heure in planning.liste_eleve:
            for i in planning.liste_eleve[cellule.heure]:
                inserer_listbox(i)


def ajoutuncheval(cheval, ind):
    """
    Supprime un élément de la listebox des chevaux et insère un nouvel élément à la position donnée.

    Cette fonction prend en argument le nom d'un cheval et sa position dans la listebox des chevaux.
    Elle supprime l'élément existant à cette position et insère un nouvel élément représentant le cheval.

    Args:
        cheval (str): Le nom du cheval.
        ind (int): La position dans la listebox.

    Returns:
        Aucun.
    """
    cheval_listbox.delete(ind)
    cheval_listbox.insert(
        ind, (planning.cheval[cheval][0], cheval))
    cheval_listbox.itemconfig(ind, {'bg': 'green'})


def ajoutcheval():
    """
    Vide la listebox des chevaux et insère les éléments correspondant à tous les chevaux du planning.

    Cette fonction vide d'abord la listebox des chevaux, puis insère les éléments correspondant à tous
    les chevaux du planning. Elle met en évidence les chevaux en vert s'ils sont associés à des élèves
    présents à l'heure de la cellule.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    cheval_listbox.delete(0, END)
    for i in planning.cheval:
        cheval_listbox.insert(tk.END, (planning.cheval[i][0], i))
    if cellule.heure != "heure" and cellule.heure in planning.liste_heure:
        colorier()


def colorier():
    """
    Change la couleur de fond des éléments de la listebox des chevaux en fonction de leur disponibilité.

    Cette fonction met en évidence les chevaux en vert s'ils sont associés à des élèves présents à
    l'heure de la cellule. Les autres chevaux ont un fond blanc.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    for i in range(0, len(planning.cheval)):
        cheval_listbox.itemconfig(
            i, {'bg': 'white'})
    setcheval = set()
    for i in planning.liste_eleve[cellule.heure]:
        print(i)
        if len(i)==2:
            ancient = planning.ancient_cheval_de(i[0], cellule.heure)
        else:
            ancient = planning.ancient_cheval_de(i, cellule.heure)
        for y in ancient:
            if y[1] != "":
                setcheval.add(y[1])
    for i in setcheval:
        cheval_listbox.itemconfig(
            i, {'bg': 'green'})


def colorier_ancient_chevaux(ancient_cheval_eleve):
    """
    Change la couleur de fond des chevaux anciens en fonction de leur disponibilité.

    Cette fonction prend en argument une liste d'éléments représentant des chevaux anciens associés à
    un élève. Elle change la couleur de fond des chevaux en fonction de leur disponibilité, en utilisant
    des couleurs telles que jaune, orange et rouge en fonction du nombre d'anciens chevaux.

    Args:
        ancient_cheval_eleve (list): Une liste d'éléments représentant des chevaux anciens.

    Returns:
        Aucun.
    """

    if len(ancient_cheval_eleve) >= 3 and ancient_cheval_eleve[2][1] != "":
        cheval_listbox.itemconfig(
            ancient_cheval_eleve[2][1], {'bg': 'yellow'})
    if len(ancient_cheval_eleve) >= 2 and ancient_cheval_eleve[1][1] != "":
        cheval_listbox.itemconfig(
            ancient_cheval_eleve[1][1], {'bg': 'orange'})
    if len(ancient_cheval_eleve) >= 1 and ancient_cheval_eleve[0][1] != "":
        cheval_listbox.itemconfig(
            ancient_cheval_eleve[0][1], {'bg': 'red'})


def changer_heure():
    """
    Met à jour l'affichage en fonction de l'heure de la cellule.

    Cette fonction met à jour l'affichage en fonction de l'heure de la cellule, en vidant la listebox
    des élèves, en ajoutant les élèves correspondant à l'heure, en coloriant les chevaux disponibles,
    et en mettant à jour les variables d'interface utilisateur.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    ajouteleve()
    colorier()
    varheure.set(f"{cellule.heure}")
    varajout.set(cellule.getCellule())


def changement_heure(i):
    """
    Change l'heure de la cellule et met à jour l'affichage.

    Cette fonction prend en argument une nouvelle heure "i", change l'heure de la cellule en conséquence
    et met à jour l'affichage.

    Args:
        i (str): La nouvelle heure pour la cellule.

    Returns:
        Aucun.
    """
    cellule.set_heure(i)
    changer_heure()


def cmp_dates(d):
    """
    Compare deux dates au format "jour-mois-année".

    Cette fonction prend en argument une date au format "jour-mois-année" et retourne un tuple (année, mois, jour)
    pour permettre une comparaison correcte des dates.

    Args:
        d (str): Une date au format "jour-mois-année".

    Returns:
        tuple: Un tuple (année, mois, jour).
    """

    j, m, a = d[0].split("-")

    return (int(a), int(m), int(j))


def recupperation_excel(listeself, name):
    """
    Récupère les données depuis un fichier Excel et les organise en listes et dictionnaires.

    Cette fonction prend en argument une liste "listeself" (non utilisée dans la fonction) et le nom d'un
    fichier Excel "name". Elle extrait les données de ce fichier, les organise en listes et dictionnaires
    (liste, dict_cheval, dict_heure) et les renvoie.

    Args:
        listeself: Liste (non utilisée dans la fonction).
        name (str): Le nom du fichier Excel à lire.

    Returns:
        tuple: Un tuple contenant trois éléments :
            - Une liste de tuples (heure, cheval, élève).
            - Un dictionnaire des chevaux avec leur indice et ligne.
            - Un dictionnaire des heures avec leur libellé.
    """
    workbook = load_workbook(name)
    sheet = workbook.active
    liste = []
    dict_heure = {}
    dict_cheval = {}
    Nb = 0
    for i in range(3, len(sheet["A"])+1):
        for j in list(range(1, len(sheet[3])+1))[::-1]:
            valeur_case = str(sheet.cell(row=i, column=j).value).strip()
            if valeur_case != 'None':
                if i == 3:
                    dict_heure[valeur_case] = j
                elif j == 1:
                    dict_cheval[valeur_case] = [
                        Nb, i]
                    Nb = 0
                elif j > 1 and valeur_case != "MERCREDI" and valeur_case != "SAMEDI" and sheet.cell(row=3, column=j).value != None and sheet.cell(row=i, column=1).value != None:
                    Nb += 1
                    liste.append(
                        (sheet.cell(row=3, column=j).value.strip(), sheet.cell(row=i, column=1).value.strip(), valeur_case))
    # print("liste : ", liste, "dict_cheval :",dict_cheval,"dict_heure :",dict_heure)
    return liste, dict_cheval, dict_heure



def sort_files_by_date(files):
    # Trie les fichiers par date (les plus récents en premier)
    sorted_files = sorted(files, key=lambda x: x[1], reverse=True)
    return sorted_files


def extract_date_from_filename(filename):
    # Extrait la date du nom du fichier
    date_str = filename.split()[-1].split('.')[0]
    file_date = datetime.strptime(date_str, "%d-%m-%Y")
    return file_date


def recup_donne2(name):
    """
    Récupère les données depuis un fichier Excel et initialise le planning.

    Cette fonction affiche une boîte de dialogue pour sélectionner un fichier Excel, puis utilise la
    fonction "recupperation_excel" pour extraire les données du fichier, les initialise dans l'objet
    "planning" et met à jour l'affichage.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    global ancient_nom, cpt_heuresuppr, cpt_chevalsuppr, ancient_nom2, ancient_nom3
    cpt_chevalsuppr = 0
    cpt_heuresuppr = 0
    planning.cheval.clear()
    planning.liste_heure.clear()
    interface_default()
    dict_cheval = {}
    for i in lire_fichier_chevaux():
        dict_cheval[i[1]] = [0, i[0]+3]
    planning.set_cheval(dict_cheval)

    if "mercredi" in name[0].lower():
        jour.set_mercredi()
    else:
        jour.set_samedi()
    planning.set_liste_eleve(lire_fichier_cavalier(jour.j))
    varjour.set(jour.j)
    planning.set_nom_fichier(name[0])

    dict_planning, cheval, heure = recupperation_excel("", name[0])
    for i in planning.cheval.keys():
        if i in cheval:
            dict_cheval[i] = cheval[i]

    planning.set_cheval(dict_cheval)
    planning.set_heure(heure)
    planning.set_planning(dict_planning)
    ajoutcheval()
    visu_fichier.delete('1.0', END)
    visu_fichier.insert(END, planning.fichier(jour.j))

    liste = name

    if len(liste) > 1:
        varsemaine1.set(name[1].replace(".xlsx", ""))
        ancient_nom = name[1]
        ancient_planning, x, y = recupperation_excel(
            "ancien", name[1])
        planning.set_ancien_planning(ancient_planning)
    if len(liste) > 2:

        varsemaine2.set(name[2].replace(".xlsx", ""))
        ancient_nom2 = name[2]
        ancient_planning2, x, y = recupperation_excel(
            "ancien2", name[2])
        planning.set_ancien_planning2(ancient_planning2)
    if len(liste) > 3:
        ancient_nom3 = name[3]
        varsemaine3.set(ancient_nom3.replace(".xlsx", ""))
        ancient_planning3, x, y = recupperation_excel(
            "ancien3", name[3])
        planning.set_ancien_planning3(ancient_planning3)
    msgbox = tk.messagebox.showinfo(
        title="Création de fichier", message="Tous les fichiers ont été récupérés")
    remplir_listecombo_heure()
    # print(planning.planning)
    # print(planning.liste_eleve)


def ecrire_fichier():
    """
    Enregistre les données dans un fichier Excel.

    Cette fonction ouvre le fichier Excel actuellement en cours de modification, efface les données des
    cellules correspondant au planning, puis insère les nouvelles données du planning. Elle enregistre
    ensuite le fichier Excel.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    global cpt_chevalsuppr, cpt_heuresuppr
    workbook = load_workbook(planning.name_fichier)
    sheet = workbook.active
    liste_cheval = list(planning.cheval.keys())
    data2 = lire_fichier_cavalier(jour.j)
    cle2 = list(data2.keys())
    cle2 = sorted(cle2, key=cmp_heure)
    dico_numeros = {
        1: 'A',
        2: 'B',
        3: 'C',
        4: 'D',
        5: 'E',
        6: 'F',
        7: 'G',
        8: 'H',
        9: 'I',
        10: 'J',
        11: 'K',
        12: 'L',
        13: 'M',
        14: 'N',
        15: 'O',
        16: 'P',
        17: 'Q',
        18: 'R',
        19: 'S',
        20: 'T',
        21: 'U',
        22: 'V',
        23: 'W',
        24: 'X',
        25: 'Y',
        26: 'Z'
    }
    # effacer excel
    # vert ligne = #A9D08E
    # vert heure = #70AD47
    # {'thin', 'slantDashDot', 'dotted', 'mediumDashDotDot', 'double', 'medium', 'thick', 'mediumDashed', 'dashed', 'dashDotDot', 'hair', 'mediumDashDot', 'dashDot'}
    if jour.j == "Mercredi":
        taillecellule = 107
    else:
        taillecellule = 75
    if jour.j == "Mercredi":
        hauteurcellule = 94
    else:
        hauteurcellule = 90
    double = Side(border_style="thin", color="000000")
    sans = Side(border_style=None, color='FF000000')
    if cpt_heuresuppr < 0:
        cpt_heuresuppr = 0
    if cpt_chevalsuppr < 0:
        cpt_chevalsuppr = 0
    for ligne in range(3, len(liste_cheval)+4+cpt_chevalsuppr):
        for colonne in range(1, len(cle2)+2+cpt_heuresuppr):
            sheet.cell(ligne, colonne).border = Border(
                left=sans, top=sans, right=sans, bottom=sans)
            sheet.cell(ligne, colonne).value = None
            sheet.cell(ligne, colonne).fill = PatternFill(
                start_color='FFFFFFFF', end_color='FF000000', fill_type=None)
            sheet.cell(ligne, colonne).font = Font(size="72")
            if colonne-2 < len(cle2) and ligne-4 < len(liste_cheval):
                sheet.cell(ligne, colonne).border = Border(
                    left=double, top=double, right=double, bottom=double)
            sheet.cell(
                ligne, colonne).alignment = Alignment(horizontal='center', vertical='center')
            sheet.row_dimensions[ligne].height = hauteurcellule
            sheet.column_dimensions[dico_numeros[colonne]
                                    ].width = taillecellule
            # sheet.cell(ligne, colonne).width = 55
            if ligne % 2 == 0 and colonne-2 < len(cle2) and ligne-4 < len(liste_cheval):
                sheet.cell(ligne, colonne).fill = PatternFill(
                    start_color='A9D08E', end_color='A9D08E', fill_type='solid')
            if ligne == 3 and colonne % 2 == 0 and colonne-2 < len(cle2):
                sheet.cell(ligne, colonne).fill = PatternFill(
                    start_color='70AD47', end_color='70AD47', fill_type='solid')
            if ligne == 3 and colonne >= 2 and colonne-2 < len(cle2):
                sheet.cell(
                    ligne, colonne).value = cle2[colonne-2]
                sheet.cell(ligne, colonne).font = Font(size="72")
            elif colonne == 1 and ligne >= 4 and ligne-4 < len(liste_cheval):
                sheet.cell(
                    ligne, colonne).value = liste_cheval[ligne-4]

    dict_heure = {}
    Nb = 0
    for i in cle2:
        dict_heure[i] = Nb+2
        Nb += 1
    planning.set_heure(dict_heure)

    for i in planning.planning:
        if i[1] in planning.cheval and i[0] in planning.liste_heure:
            sheet.cell(planning.cheval[i[1]][1],
                       planning.liste_heure[i[0]]).value = i[2]
    err = workbook.save(planning.name_fichier)
    if err == None:
        label_enregistrer.config(fg="#ffffff")
        sauvegarder_liste_eleve()
        upload_les_excel()


def add_heure():
    global cpt_heuresuppr
    data[para_input_heure.get().upper()] = []
    para_inserer_listebox(data)
    visualiser_fichier_cavalier(data)
    cpt_heuresuppr -= 1


def suppr_heure():
    global data, cpt_heuresuppr
    del data[heure]
    para_inserer_listebox(data)
    visualiser_fichier_cavalier(data)
    cpt_heuresuppr += 1


def add_eleve():
    global data
    global heure
    data[heure].append(para_input_eleve.get().upper().strip())
    remplirlisteboxeleve()
    visualiser_fichier_cavalier(data)


def add_cheval():
    global cpt_chevalsuppr
    present = any(para_input_chevaux.get().upper() == tup[1]
                  for tup in chevaux)
    if not present and len(chevaux)+1 >= int(para_input_ind_chevaux.get()):

        for cheval in chevaux:
            if cheval[0] >= int(para_input_ind_chevaux.get()):
                cheval[0] += 1
        chevaux.append([int(para_input_ind_chevaux.get()),
                        para_input_chevaux.get().upper()])
        chevaux.sort()
        remplirlisteboxcheval(chevaux)
        cpt_chevalsuppr -= 1


def suppr_cheval():
    global cpt_chevalsuppr
    if cheval in chevaux:
        chevaux.remove(cheval)
        for che in chevaux:
            if che[0] >= cheval[0]:
                che[0] -= 1
        chevaux.sort()
        remplirlisteboxcheval(chevaux)
        cpt_chevalsuppr += 1


def ecrire_fichier_cheval(chevaux):
    txt = ""
    for cheval in chevaux:
        txt += str(cheval[0]) + '\t' + cheval[1] + "\r"
    return txt


def suppr_eleve():
    global parajour
    liste = []
    for eleves in data[heure]:
        if eleves != eleve:
            liste.append(eleves)
    data[heure] = liste[:]
    remplirlisteboxeleve()
    visualiser_fichier_cavalier(data)


def para_enregistrer():
    global parajour
    if parajour != '' and data:
        fichier = open("liste_cavalier_" + parajour + ".txt", "w")
        fichier.write(ecrire_fichier_cavalier(data))
        fichier.close()
    fichier = open("liste_cheval.txt", "w")
    fichier.write(ecrire_fichier_cheval(chevaux))
    fichier.close()


def remplirlisteboxcheval(chevaux):
    para_listebox_chevaux.delete(0, END)
    for cheval in chevaux:
        para_listebox_chevaux.insert(END, cheval)


def remplirlisteboxeleve():
    para_listebox_eleve.delete(0, END)
    for eleve in data[heure]:
        para_listebox_eleve.insert(END, eleve)


def interface_default():
    global parajour
    for widget in widgets_parametre:
        widget.place_forget()

    image_label.destroy()

    label_jour.place(x=133, y=140)
    label_heure.place(x=183, y=140)
    title_label.place(x=60, y=35)
    boutton_avancer_heure.place(x=65, y=140)
    boutton_reculer_heure.place(x=260, y=140)
    label_cavalier.place(x=460, y=70)
    label_cavalier2.place(x=460, y=100)
    label_cavalier3.place(x=650, y=100)
    label_cavalier6.place(x=460, y=150)
    label_cavalier4.place(x=650, y=150)
    label_cavalier7.place(x=460, y=200)
    label_cavalier5.place(x=650, y=200)
    boutton_absent.place(x=755, y=100)
    boutton_correction.place(x=810, y=100)
    eleve_listbox.place(x=133, y=160)
    eleve_rattrapage.place(x=133, y=360)
    label_eleve_rattrapage.place(x=137, y=330)
    boutton_eleve_rattrapage.place(x=160, y=390)
    cheval_listbox.place(x=330, y=35)
    visu_fichier.place(x=900, y=400)
    label_visu_fichier.place(x=900, y=370)
    label_ajout.place(x=460, y=400)
    boutton_ajouter.place(x=570, y=480)
    boutton_supprimer.place(x=670, y=480)
    boutton_enregistrer.place(x=570, y=530)
    label_enregistrer.place(x=560, y=585)
    label_heure_cheval.place(x=460, y=250)
    heure_listebox.place(x=460, y=280)
    historique.place(x=900, y=70)
    label_historique.place(x=900, y=40)
    listeCombo.place(x=65, y=100)
    image1.place(x=535, y=606)
    image2.place(x=70, y=606)
    image3.place(x=680, y=220)
    bouton_ouvrir_excel.place(x=1400, y=60)
    bouton_rafraichir.place(x=1400, y=100)

    dict_cheval = {}

    if parajour != "":
        for i in chevaux:
            if i[1] in planning.cheval:
                if i[0] == planning.cheval[i[1]][1]-3:
                    dict_cheval[i[1]] = planning.cheval[i[1]]
                else:
                    dict_cheval[i[1]] = [planning.nb_heure(i[1]), i[0]+3]
            else:

                dict_cheval[i[1]] = [planning.nb_heure(i[1]), i[0]+3]

        planning.set_cheval(dict_cheval)

    if parajour == jour.j:
        planning.set_heure(data)
        planning.set_liste_eleve(data)
    elif jour.j != '':
        planning.set_liste_eleve(lire_fichier_cavalier(jour.j))
    else:
        planning.set_liste_eleve({})
    if planning.liste_eleve != {}:
        ajouteleve()
    if planning.cheval != []:

        ajoutcheval()

    remplir_listecombo_heure()


def interface_paramete():
    for widget in widgets_principaux:
        widget.place_forget()

    image_label.destroy()

    para_image1.place(relx=0.48, rely=0.6, anchor=tk.CENTER)
    para_visu_fichier.place(x=900, y=400)
    # para_label_historique.place(x=1000, y=40)
    # para_historique.place(x=1000, y=70)
    para_listebox_heure.place(x=400, y=70)
    para_listebox_eleve.place(x=730, y=70)
    para_listeCombo.place(x=65, y=40)
    para_input_heure.place(x=560, y=140)
    para_add_heure.place(x=560, y=170)
    para_suppr_heure.place(x=560, y=200)
    para_input_eleve.place(x=890, y=140)
    para_add_eleve.place(x=890, y=170)
    para_suppr_eleve.place(x=890, y=200)
    para_suppr_enregistrer.place(x=635, y=680)
    para_listebox_chevaux.place(x=60, y=70)
    para_input_chevaux.place(x=220, y=140)
    para_add_chevaux.place(x=220, y=170)
    para_suppr_chevaux.place(x=220, y=200)
    para_input_ind_chevaux.place(x=360, y=140)
    # image1.place(x=1050, y=70)
    # image3.place(x=605, y=300)
    image3.place(x=170, y=500)

    image4.place(x=1070, y=70)


def lire_fichier_chevaux():
    liste = []
    fichier = open("liste_cheval.txt", "r")
    lignes = fichier.read()
    fichier.close()
    lignes = lignes.split("\n")
    for ligne in lignes:
        if ligne != '':
            if ligne[2] == '\t':
                liste.append([int(ligne[:2]), ligne[3:].strip()])
            else:
                liste.append([int(ligne[:2]), ligne[2:].strip()])
    return liste


def lire_fichier_cavalier(jour):
    liste = []
    liste_eleve = {}
    fichier = open("liste_cavalier_" + jour + ".txt", "r")
    lignes = fichier.read()
    fichier.close()
    para_visu_fichier.delete('1.0', END)
    para_visu_fichier.insert(END, lignes)
    lignes = lignes.split("\n")
    for ligne in lignes:
        if ligne != '':
            if "\\Fin fichier/" in ligne:
                return liste_eleve
            if "\\heure/" in ligne:
                liste_eleve[ligne[7:].strip()] = liste[:]
                liste.clear()
            else:
                nom, numero = ligne.strip().split('/')
                # if numero != "-1":
                liste.append([nom,int(numero)])
                # else:
                #     liste.append(nom)


def para_inserer_listebox(data):

    para_listebox_heure.delete(0, END)
    for i in list(data.keys()):
        para_listebox_heure.insert(END, i)


def mode_parametre():
    interface_paramete()


def mode_default():
    interface_default()
    # lire_fichier_cavalier(jour.j)


def visualiser_fichier_cavalier(data):
    txt = ""
    for heure in data.keys():
        for eleve in data[heure]:
            txt += eleve + "\r\n"
        txt += "\\heure/" + heure + "\r\n"
    txt += "\\Fin fichier/"
    para_visu_fichier.delete('1.0', END)
    para_visu_fichier.insert(END, txt)


def cmp_heure(d):
    if d[1].isdigit():
        heure = d[:2]
        if d[3] == '3':
            heure += '5'
        else:
            heure += '0'
    elif d[1] == 'H':
        heure = d[0]
        if d[2] == '3':
            heure += '5'
        else:
            heure += '0'
    else:
        heure = d[0]
    return (int(heure))


def ecrire_fichier_cavalier(data):
    txt = ""
    cle = list(data.keys())

    cle = sorted(cle, key=cmp_heure)

    for heure in cle:
        for eleve in data[heure]:
            txt += eleve + "\r"
        txt += "\\heure/" + heure + "\r"
    txt += "\\Fin fichier/"
    return txt


def prochain_jour(semaine, jour_actuel):
    jours_de_la_semaine = ["lundi", "mardi", "mercredi",
                           "jeudi", "vendredi", "samedi", "dimanche"]
    jour_cible = jours_de_la_semaine.index(semaine)

    jours_jusquau_prochain = (jour_cible - jour_actuel.weekday() + 7) % 7
    if jours_jusquau_prochain == 0:
        jours_jusquau_prochain = 7

    return jour_actuel + timedelta(days=jours_jusquau_prochain)


def add_centered_image(root, image_path, width, height):
    # Charger l'image
    original_image = Image.open(image_path)

    # Redimensionner l'image
    resized_image = original_image.resize((width, height), Image.NEAREST)
    photo = ImageTk.PhotoImage(resized_image)

    # Configurer l'image redimensionnée comme arrière-plan de la fenêtre
    image_label = tk.Label(root, borderwidth=0,
                           image=photo, highlightthickness=0, bg="#b4b4b4")
    image_label.image = photo
    image_label.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
    return image_label


def set_background(root, image_path):
    # Charger l'image
    original_image = Image.open(image_path)

    # Obtenir la taille de l'écran
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # Redimensionner l'image pour s'adapter à l'écran
    resized_image = original_image.resize((screen_width, screen_height), Image.NEAREST)

    photo = ImageTk.PhotoImage(resized_image)

    # Configurer l'image redimensionnée comme arrière-plan de la fenêtre
    background_label = tk.Label(root, image=photo, bg="#b4b4b4")
    background_label.image = photo
    background_label.place(x=0, y=0, relwidth=1, relheight=1)
    return background_label, photo

def nouveau_fichier():
    global pop
    pop = Toplevel(window)
    pop.title("Creation de excel")
    pop.geometry("350x150")

    date_actuelle = datetime.now()
    v = StringVar()
    date = StringVar()

    # Vérifier si aujourd'hui est déjà un mercredi
    if date_actuelle.weekday() == 2:  # Le code de semaine pour mercredi est 2 (0 pour lundi, 1 pour mardi, etc.)
        prochains_mercredis = [date_actuelle]
        prochain_mercredi = date_actuelle
    else:
        # Trouver le prochain mercredi
        prochain_mercredi = prochain_jour("mercredi", date_actuelle)

        # Ajouter les trois prochains mercredis au tableau
        prochains_mercredis = [prochain_mercredi]
    for _ in range(2):
        prochain_mercredi = prochain_jour("mercredi", prochain_mercredi)
        prochains_mercredis.append(prochain_mercredi)

    if date_actuelle.weekday() == 5:  # Le code de semaine pour mercredi est 2 (0 pour lundi, 1 pour mardi, etc.)
        prochains_samedis = [date_actuelle]
        prochain_samedi = date_actuelle
    else:
        # Trouver le prochain mercredi
        prochain_samedi = prochain_jour("samedi", date_actuelle)

        # Ajouter les trois prochains mercredis au tableau
        prochains_samedis = [prochain_samedi]
    for _ in range(2):
        prochain_samedi = prochain_jour("samedi", prochain_samedi)
        prochains_samedis.append(prochain_samedi)

    pop_r1_1 = Radiobutton(pop, text=prochains_samedis[0].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_samedis[0].strftime("%d-%m-%Y"))
    pop_r1_2 = Radiobutton(pop, text=prochains_samedis[1].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_samedis[1].strftime("%d-%m-%Y"))
    pop_r1_3 = Radiobutton(pop, text=prochains_samedis[2].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_samedis[2].strftime("%d-%m-%Y"))

    pop_r2_1 = Radiobutton(pop, text=prochains_mercredis[0].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_mercredis[0].strftime("%d-%m-%Y"))
    pop_r2_2 = Radiobutton(pop, text=prochains_mercredis[1].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_mercredis[1].strftime("%d-%m-%Y"))
    pop_r2_3 = Radiobutton(pop, text=prochains_mercredis[2].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_mercredis[2].strftime("%d-%m-%Y"))

    def affichejoursam():
        pop_r2_1.place_forget()
        pop_r2_2.place_forget()
        pop_r2_3.place_forget()

        pop_r1_1.place(y=40, x=80)
        pop_r1_2.place(y=60, x=80)
        pop_r1_3.place(y=80, x=80)

    def affichejoursmer():
        pop_r1_1.place_forget()
        pop_r1_2.place_forget()
        pop_r1_3.place_forget()
        # Vérifier si aujourd'hui est déjà un mercredi

        pop_r2_1.place(y=40, x=80)
        pop_r2_2.place(y=60, x=80)
        pop_r2_3.place(y=80, x=80)

    pop_label = Label(
        pop, text="voulez vous creer un fichier Mercredi ou Samedi")
    pop_label.place(x=40, y=20)

    v.set("Mercredi")  # initialiser
    pop_r1 = Radiobutton(pop, text="Mercredi", variable=v,
                         value="mercredi", command=affichejoursmer)
    pop_r1.place(x=10, y=40)
    pop_r2 = Radiobutton(pop, text="Samedi", variable=v,
                         value="samedi", command=affichejoursam)
    pop_r2.place(x=10, y=60)

    def choix_date():
        pop.destroy()
        workbook = Workbook()
        name = 'liste ' + v.get() + ' ' + date.get() + '.xlsx'
        workbook.save(name)
        ftp.telecharger_fichier_ftp(name)
        samedi_file_names, mercredi_file_names = ftp.download_files_from_ftp()
        if v.get() == "samedi":
            files = samedi_file_names
            menu = file_menu
        if v.get() == "mercredi":
            files = mercredi_file_names
            menu = edit_menu
        menu.delete(0, tk.END)
        for i, file_name in enumerate(files[:5]):
            menu.add_command(
                label=file_name, command=lambda file_name=file_name:  download_selected_and_recent_files(v.get(), name))

        download_selected_and_recent_files(v.get(), name)
    pop_valider = Button(pop, text="Valider", command=choix_date)
    pop_valider.place(x=300, y=120)


def image(root, image_path, width, height):
    original_image = Image.open(image_path)

    # Redimensionner l'image
    resized_image = original_image.resize((width, height), Image.NEAREST)
    photo = ImageTk.PhotoImage(resized_image)

    # Configurer l'image redimensionnée comme arrière-plan de la fenêtre
    image_label = tk.Label(root, borderwidth=0,
                           image=photo, highlightthickness=0, bg="#b4b4b4")
    image_label.image = photo
    return image_label


def ecrire_dans_fichier(tableau_tuples, nom_fichier='donnees.txt'):
    with open(nom_fichier, 'w') as fichier:
        for ligne in tableau_tuples:
            fichier.write(','.join(ligne) + '\n')


def lire_depuis_fichier(nom_fichier='donnees.txt'):
    tableau_tuples = []
    with open(nom_fichier, 'r') as fichier:
        lignes = fichier.readlines()
        for ligne in lignes:
            heure, cheval, personne = ligne.strip().split(',')
            tableau_tuples.append((heure, cheval, personne))
    return tableau_tuples

def online():

    # Importation des modules
    cellule = Cellule()  # Création d'une instance de la classe Cellule
    planning = Planning()  # Création d'une instance de la classe Planning
    jour = Jour()  # Création d'une instance de la classe Jour
    mode=1
    try:
        ftp = Ftp(adresse_serveur, nom_utilisateur, mot_de_passe)
        samedi_file_names, mercredi_file_names = ftp.download_files_from_ftp()
    except:
        mode = 0


    # Création de l'interface utilisateur
    window = tk.Tk()  # Création de la fenêtre principale
    window.title("Planning")  # Titre de la fenêtre
    window.attributes('-fullscreen', True)  # Affichage en mode plein écran
    # Permet de quitter en appuyant sur la touche "Échap"
    window.bind('<Escape>', lambda e: quitter())
    window_width = window.winfo_width()
    window_height = window.winfo_height()


    # window.wm_attributes('-alpha', 0)
    # window.wm_attributes('-transparentcolor', '#f0f0f0')


    # window.configure(bg='#b4b4b4')
    set_background(window, "image_fond.png")


    widgets_principaux = []

    widgets_parametre = []

    ancient_nom = ""
    ancient_nom2 = ""
    ancient_nom3 = ""

    image_label = add_centered_image(
        window, "0f382f680a13445c8e6484ecbbe2a2b5-transformed.png", 169*4, 166*4)

    para_image1 = image(
        window, "image1.png", int(2388/5), int(1668/5))
    image1 = image(window, "image1.png", int(2388/8.5), int(1668/8.5))
    image2 = image(window, "image2.png", int(2388/8.5), int(1668/8.5))
    image3 = image(window, "image3.png", int(2388/8.5), int(1668/8.5))
    image4 = image(window, "image4.png", int(2388/7), int(1668/7))

    # Création d'un cadre dans la fenêtre
    # frame = tk.Frame(master=window, width=300, height=100)
    # frame.pack()

    # Définition des variables de contrôle
    varheure = StringVar()
    varjour = StringVar()
    varcavalier = StringVar()
    varcheval = StringVar()
    varajout = StringVar()
    varheure_cheval = StringVar()
    varcavalier1 = StringVar()
    varcavalier2 = StringVar()
    varsemaine1 = StringVar()
    varsemaine2 = StringVar()
    varsemaine3 = StringVar()

    label_jour = tk.Label(window, textvariable=varjour, bg='#ffffff')

    label_heure = tk.Label(window, textvariable=varheure, bg='#ffffff')

    # Création d'une étiquette pour le titre
    title_label = tk.Label(
        window, text="GESTION PLANNING", font=("Comic Sans MS", 17), bg='#b4b4b4')

    # Boutons pour avancer et reculer dans les heures
    boutton_avancer_heure = tk.Button(
        window, width=8, bg='#8abd45', text="precedent", command=heure_precedant)

    boutton_reculer_heure = tk.Button(
        window, width=8, bg='#8abd45', text="suivant", command=heure_suivant)

    # Étiquettes pour afficher les informations du cavalier
    label_cavalier = tk.Label(
        window, text="INFOS CAVALIER", font=("Corbel", 14), bg='#8abd45')

    label_cavalier2 = tk.Label(
        window, textvariable=varsemaine1, font=("Corbel", 13), bg='#b4b4b4')

    label_cavalier3 = tk.Label(
        window, textvariable=varcavalier, font=("Corbel", 13), bg='#b4b4b4')

    label_cavalier6 = tk.Label(
        window, textvariable=varsemaine2, font=("Corbel", 13), bg='#b4b4b4')

    label_cavalier4 = tk.Label(
        window, textvariable=varcavalier1, font=("Corbel", 13), bg='#b4b4b4')

    label_cavalier7 = tk.Label(
        window, textvariable=varsemaine3, font=("Corbel", 13), bg='#b4b4b4')
    label_cavalier5 = tk.Label(
        window, textvariable=varcavalier2, font=("Corbel", 13), bg='#b4b4b4')


    def correction():
        workbook = load_workbook(ancient_nom)
        sheet = workbook.active
        if elevecarte == True and varcavalier.get() == "cheval":
            unesessionmoins(cellule.eleve)


        for ind in range(1, len(sheet["A"])+1):

            if sheet.cell(ind, 1).value == cellule.cheval:
                sheet.cell(ind,
                        planning.liste_heure[cellule.heure]).value = cellule.eleve
            if sheet.cell(ind, 1).value == dernier_cheval:

                sheet.cell(ind,
                        planning.liste_heure[cellule.heure]).value = None
        err = workbook.save(ancient_nom)
        if err == None:
            varcavalier.set(cellule.cheval)

    def absent():
        workbook = load_workbook(ancient_nom)
        sheet = workbook.active
        if elevecarte == True:
            unesessionplus(cellule.eleve)
        for ind in range(1, len(sheet["A"])+1):

            if sheet.cell(ind, 1).value == dernier_cheval:

                sheet.cell(ind,
                        planning.liste_heure[cellule.heure]).value = None
        err = workbook.save(ancient_nom)
        if err == None:
            varcavalier.set("cheval")


    dernier_cheval = ""

    boutton_absent = tk.Button(
        window, bg='#8abd45', height=1, width=4, text="ABS", command=absent, borderwidth=2)
    boutton_correction = tk.Button(
        window, bg='#8abd45', height=1, text="correction", command=correction)
    # Initialisation des variables de contrôle
    varcavalier.set("cheval")
    varcavalier1.set("cheval1")
    varcavalier2.set("cheval2")

    # Liste déroulante pour les élèves
    eleve_listbox = tk.Listbox(window, yscrollcommand=True)

    eleve_rattrapage = tk.Entry(window)

    # style = ttk.Style()
    # style.theme_use("clam")

    label_eleve_rattrapage = tk.Label(
        window, text="Ajouter un nom", font=("Corbel", 13), bg='#b4b4b4')
    boutton_eleve_rattrapage = tk.Button(
        window, width=8, bg='#8abd45', text="rattrapage", command=ajouter_rattrapage)

    # Fonction appelée lorsqu'un élément est sélectionné dans la liste des élèves


    def items_selected(event):
        global dernier_cheval,elevecarte
        # Indices des éléments sélectionnés
        selected_indices = eleve_listbox.curselection()
        eleve = eleve_listbox.get(selected_indices)

        cellule.set_eleve(eleve, selected_indices[0])
        elevecarte = False
        if isinstance(cellule.eleve[1], int):
            cellule.eleve = cellule.eleve[0]
            elevecarte = True
            
        ancient_cheval = planning.ancient_cheval_de(
                cellule.eleve, cellule.heure)

        # Mise à jour des étiquettes des chevaux associés
        if ancient_cheval[0][1] != "":
            varcavalier.set(ancient_cheval[0][0])
            dernier_cheval = ancient_cheval[0][0]
        else:
            varcavalier.set("cheval")
        if len(ancient_cheval) >= 2:
            varcavalier1.set(ancient_cheval[1][0])
        else:
            varcavalier1.set("cheval1")
        if len(ancient_cheval) >= 3:
            varcavalier2.set(ancient_cheval[2][0])
        else:
            varcavalier2.set("cheval2")
        colorier()
        colorier_ancient_chevaux(ancient_cheval)
        for tup in planning.planning:
            if (cellule.heure, eleve) == (tup[0], tup[2]):
                cavalier = []
                cellule.set_cheval(tup[1], planning.index_cheval(tup[1]))
                ancient_cavalier = planning.ancient_eleve_de(cellule.cheval)
                for i in ancient_cavalier:
                    cavalier.append(f"{i[0]} a {i[1]}")
                varheure_cheval.set(
                    f"HEURE DE TRAVAIL DE: {cellule.cheval}")
                varcheval.set(cavalier)
                inserer_liste_de_travaille()
        varajout.set(cellule.getCellule())


    # Association de la fonction à l'événement de relâchement du bouton de la souris
    eleve_listbox.bind('<ButtonRelease-1>', items_selected)

    # Liste déroulante pour les chevaux
    cheval_listbox = tk.Listbox(window, height=47)

    # Fonction appelée lorsqu'un élément est sélectionné dans la liste des chevaux


    def items_selected_cheval(event):
        # Indices des éléments sélectionnés
        selected_indices = cheval_listbox.curselection()
        cavalier = []
        cheval = cheval_listbox.get(selected_indices)
        cellule.set_cheval(cheval[1], selected_indices)
        ancient_cavalier = planning.ancient_eleve_de(
            cheval_listbox.get(selected_indices)[1])
        for i in ancient_cavalier:
            cavalier.append(f"{i[0]} a {i[1]}")
        varheure_cheval.set(f"HEURE DE TRAVAIL DE: {cellule.cheval}")
        inserer_liste_de_travaille()
        varcheval.set(cavalier)
        varajout.set(cellule.getCellule())


    # Association de la fonction à l'événement de relâchement du bouton de la souris
    cheval_listbox.bind('<ButtonRelease-1>', items_selected_cheval)

    # Zone de texte pour afficher le planning
    visu_fichier = tk.Text(window, width=70)

    label_visu_fichier = tk.Label(
        window, text="PREVISUALISATION", font=("Corbel", 14), bg='#8abd45')


    # Étiquette pour afficher des informations sur l'ajout
    label_ajout = tk.Label(window, textvariable=varajout,
                        font=20, bg='#ffffff')


    # Bouton pour ajouter une entrée
    boutton_ajouter = tk.Button(
        window, text="Ajouter", command=ajouter, width=11, height=2, bg='#8abd45')


    # Bouton pour supprimer une entrée
    boutton_supprimer = tk.Button(
        window, text="Supprimer", command=supprimer, width=11, height=2, bg='#8abd45')


    # Bouton pour enregistrer les modifications
    boutton_enregistrer = tk.Button(
        window, text="ENREGISTRER", command=ecrire_fichier, width=12, font=("Helvetica", 18, "bold"), bg='#000000', fg='#ffffff')


    # Étiquette pour afficher un message après l'enregistrement
    label_enregistrer = tk.Label(
        window, text="Le fichier a bien été enregistré", font=("Corbel", 13), bg='#b4b4b4')  # le fichier à bien été enregistré
    label_enregistrer.config(fg="#b4b4b4")

    # Étiquette pour afficher l'heure de travail du cheval
    label_heure_cheval = tk.Label(
        window, textvariable=varheure_cheval, font=("Corbel", 13), bg='#8abd45')


    # Liste déroulante pour les heures de travail
    heure_listebox = tk.Listbox(window, width=25, height=5)


    def ouvrir_excel():
        subprocess.Popen(['start', 'excel', planning.name_fichier], shell=True)


    def rafraichir():
        dict_planning, cheval, heure = recupperation_excel(
            "", planning.name_fichier)
        planning.set_planning(dict_planning)
        planning.set_cheval(cheval)
        planning.set_heure(heure)
        ajoutcheval()
        affichage_txt(jour, planning)


    bouton_ouvrir_excel = tk.Button(
        window, text="ouvrir", bg="#8abd45", command=ouvrir_excel)

    bouton_rafraichir = tk.Button(
        window, text="rafraichir", bg="#8abd45", command=rafraichir)

    # Fonction appelée lorsqu'un élément est sélectionné dans la liste des heures de travail


    def items_selected_heure_cheval(event):
        # Indices des éléments sélectionnés
        selected_indices = heure_listebox.curselection()
        (h, p) = heure_listebox.get(selected_indices)
        if h != cellule.heure:
            cellule.set_heure(h)
            cellule.set_eleve(p, -1)
        else:
            Nb = 0
            for i in range(0, eleve_listbox.size()):
                if p == eleve_listbox.get(i):
                    cellule.set_eleve(p, Nb)
                Nb += 1
        varajout.set(cellule.getCellule())


    # Association de la fonction à l'événement de relâchement du bouton de la souris
    heure_listebox.bind('<ButtonRelease-1>', items_selected_heure_cheval)

    # Étiquette pour afficher l'historique
    label_historique = tk.Label(
        window, text="HISTORIQUE", font=("Corbel", 13), bg='#8abd45')

    # Zone de texte pour afficher l'historique
    historique = tk.Text(window, width=60, height=13)


    # Création du menu
    menubar = Menu(window)

    # Création d'une liste déroulante pour sélectionner l'heure
    listeCombo = ttk.Combobox(window, height=10, width=40)


    # Fonction appelée lorsqu'un élément est sélectionné dans la liste déroulante


    def action(event):
        select = listeCombo.get()  # Élément sélectionné dans la liste déroulante
        changement_heure(select)


    listeCombo.bind("<<ComboboxSelected>>", action)


    def remplir_listecombo_heure():
        listeCombo.delete(0, "end")
        listeCombo['values'] = list(planning.liste_eleve)


    menubar = Menu(window)

    # Ajout des éléments au menu
    sousmenu = Menu(menubar, tearoff=0)
    sousmenu.add_command(label="parametre", command=mode_parametre)
    sousmenu.add_command(label="principal", command=mode_default)

    if mode == 1:
        def download_selected_and_recent_files(day, selected_file):
            files = ftp.download_selected_and_recent_files(day, selected_file)

            recup_donne2(files)


        # Ajout des éléments au menu
        menubar.add_command(label="Nouveau", command=nouveau_fichier)
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Samedi", menu=file_menu)

        for i, file_name in enumerate(samedi_file_names[:5]):
            file_menu.add_command(
                label=file_name, command=lambda file_name=file_name:  download_selected_and_recent_files("Samedi", file_name))

        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Mercredi", menu=edit_menu)

        for i, file_name in enumerate(mercredi_file_names[:5]):
            edit_menu.add_command(
                label=file_name, command=lambda file_name=file_name:  download_selected_and_recent_files("Mercredi", file_name))


        def upload_les_excel():
            fichier = []
            if planning.name_fichier:
                fichier.append(planning.name_fichier)
            if ancient_nom:
                fichier.append(ancient_nom)
            ftp.telecharger_fichier_ftp(fichier)
            
        def quitter():
            upload_les_excel()
            suppr_excel()
            window.quit()
        
        def suppr_excel():
            if planning.name_fichier:
                os.remove(planning.name_fichier)
            if ancient_nom:
                os.remove(ancient_nom)
            if ancient_nom2:
                os.remove(ancient_nom2)
            if ancient_nom3:
                os.remove(ancient_nom3)

    def sauvegarder_liste_eleve():
        fichier = open("liste_cavalier_" + jour.j + ".txt", "w")
        
        txt = ""
        cle = list(planning.liste_eleve.keys())
        cle = sorted(cle, key=cmp_heure)

        for heure in cle:
            for eleve in planning.liste_eleve[heure]:
                txt += eleve[0]+"/"+ str(eleve[1]) + "\r"
            txt += "\\heure/" + heure + "\r"
        txt += "\\Fin fichier/"
        
        fichier.write(txt)
        fichier.close()


    # menubar.add_command(label="Jour", command=recup_donne)
    menubar.add_cascade(label="Mode", menu=sousmenu)
    menubar.add_command(label="Quitter!", command=quitter)

    # Affichage du menu dans la fenêtre
    window.config(menu=menubar)

    data = {}

    # interface_paramete()

    para_visu_fichier = tk.Text(window, width=70)

    # Étiquette pour afficher l'historique
    para_label_historique = tk.Label(
        window, text="historique", font=("Corbel", 13), bg='#b4b4b4')

    # Zone de texte pour afficher l'historique
    para_historique = tk.Text(window, width=60, height=13)

    para_listebox_heure = tk.Listbox(window, width=25, height=45)

    heure = ""


    def items_selected_heure(event):
        global heure
        # Indices des éléments sélectionnés
        selected_indices = para_listebox_heure.curselection()
        heure = para_listebox_heure.get(selected_indices)
        remplirlisteboxeleve()


    # Association de la fonction à l'événement de relâchement du bouton de la souris
    para_listebox_heure.bind('<ButtonRelease-1>', items_selected_heure)


    para_listebox_eleve = tk.Listbox(window, width=25, height=12)


    eleve = ""


    def items_selected_eleve(event):
        # Indices des éléments sélectionnés
        global eleve
        selected_indices = para_listebox_eleve.curselection()
        eleve = para_listebox_eleve.get(selected_indices)


    # Association de la fonction à l'événement de relâchement du bouton de la souris
    para_listebox_eleve.bind('<ButtonRelease-1>', items_selected_eleve)


    para_listebox_chevaux = tk.Listbox(window, width=25, height=45)

    chevaux = []
    cheval = ""


    def items_selected_cheval(event):
        global cheval
        # Indices des éléments sélectionnés
        selected_indices = para_listebox_chevaux.curselection()
        cheval = list(para_listebox_chevaux.get(selected_indices))


    # Association de la fonction à l'événement de relâchement du bouton de la souris
    para_listebox_chevaux.bind('<ButtonRelease-1>', items_selected_cheval)

    # Création d'une liste déroulante pour sélectionner l'heure
    para_listeCombo = ttk.Combobox(window)
    para_listeCombo['values'] = ["mercredi", "samedi"]

    parajour = ""


    def action(event):
        global parajour
        global data
        global chevaux
        parajour = para_listeCombo.get()  # Élément sélectionné dans la liste déroulante
        data = lire_fichier_cavalier(parajour)
        chevaux = lire_fichier_chevaux()

        remplirlisteboxcheval(chevaux)
        para_inserer_listebox(data)
        return data


    para_input_chevaux = tk.Entry(window)
    para_input_ind_chevaux = tk.Entry(window, width=3)
    para_add_chevaux = tk.Button(
        window, text="ajouter cheval", command=add_cheval, width=18, bg='#8abd45')
    para_suppr_chevaux = tk.Button(
        window, text="supprimer cheval", command=suppr_cheval, width=18, bg='#8abd45')

    para_listeCombo.bind("<<ComboboxSelected>>", action)

    para_input_heure = tk.Entry(window)

    para_add_heure = tk.Button(
        window, text="creer heure", command=add_heure, width=18, bg='#8abd45')

    para_suppr_heure = tk.Button(
        window, text="supprimer heure", command=suppr_heure, width=18, bg='#8abd45')


    para_input_eleve = tk.Entry(window)

    para_add_eleve = tk.Button(
        window, text="creer eleve", command=add_eleve, width=18, bg='#8abd45')

    para_suppr_eleve = tk.Button(
        window, text="supprimer eleve", command=suppr_eleve, width=18, bg='#8abd45')

    para_suppr_enregistrer = tk.Button(
        window, text="ENREGISTRER", command=para_enregistrer, width=12, font=("Helvetica", 18, "bold"), bg='#000000', fg='#ffffff')

    widgets_parametre.extend(
        [para_visu_fichier, para_label_historique, para_historique, para_listebox_eleve, para_listebox_heure, para_listeCombo, para_input_heure, para_input_eleve, para_suppr_enregistrer, para_suppr_eleve, para_add_eleve, para_suppr_heure, para_add_heure, para_listebox_chevaux, para_input_chevaux, para_add_chevaux, para_suppr_chevaux, para_input_ind_chevaux, image3, image4, para_image1])

    widgets_principaux.extend([label_jour, label_heure, title_label, boutton_avancer_heure, boutton_reculer_heure, label_cavalier, label_cavalier2, label_cavalier3, label_cavalier6, label_cavalier4,
                            label_cavalier5, label_cavalier7, eleve_listbox, cheval_listbox, label_ajout, boutton_ajouter, boutton_supprimer, visu_fichier, label_visu_fichier,
                            boutton_enregistrer, bouton_rafraichir, bouton_ouvrir_excel, label_enregistrer, image1, image2, label_heure_cheval, heure_listebox, label_historique, historique, listeCombo, boutton_absent, boutton_correction, eleve_rattrapage, label_eleve_rattrapage, boutton_eleve_rattrapage])

    # Lancement de la boucle principale de l'application
    window.mainloop()
