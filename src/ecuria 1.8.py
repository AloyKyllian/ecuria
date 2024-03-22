from Planning import *
from Jour import *
from Word import *
import Parametre as param
from Mail import *
from Zip import *
from Maj import *
from Ftp import *
import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter.filedialog import askopenfilename, askdirectory
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import os
from datetime import datetime, timedelta
from PIL import Image, ImageTk
import subprocess
import shutil
from tkinter import messagebox

path_parametre = "parametre/"
path_cavalier_mercredi =path_parametre+ "liste_cavalier_mercredi.txt"
path_cavalier_samedi =path_parametre+ "liste_cavalier_samedi.txt"
path_cavalier_semaine =path_parametre+ "liste_cavalier_semaine.txt"
path_cheval =path_parametre+ "liste_cheval.txt"
path_cheval_semaine =path_parametre+ "liste_cheval_semaine.txt"
path_mail =path_parametre+ "mail.txt"
path_user = "user.txt"
path_Mercredi =path_parametre+ "Mercredi.xlsx"
path_Samedi =path_parametre+ "Samedi.xlsx"
path_semaine = path_parametre+ "Semaine.xlsx"


def remplir_cheval(dict_chevaux):
    for cheval in dict_chevaux:
        print(planning.planning)
        print(cheval, planning.nb_heure(cheval))
        dict_chevaux[cheval] = [dict_chevaux[cheval][0], planning.nb_heure(cheval)]
    return dict_chevaux
    

def get_personne():
    with open(path_user, "r") as file:
        return file.read()

def get_mail():
    with open(path_mail, 'r') as file:
        lines = file.readlines()
        tableau = [line.strip() for line in lines]
    print(tableau)
    return tableau
    
def unesessionmoins(eleve,heure):
    for i in range(len(planning.liste_eleve[heure])):
        if planning.liste_eleve[heure][i][0] == eleve:
            planning.liste_eleve[heure][i][1] -= 1
            if planning.liste_eleve[heure][i][1] == 0:
                planning.liste_eleve[heure][i][1] = 10
            break
    ajouteleve()
    
def unesessionplus(eleve,heure):
    for i in range(len(planning.liste_eleve[heure])):
        if planning.liste_eleve[heure][i][0] == eleve:
            planning.liste_eleve[heure][i][1] += 1
            if planning.liste_eleve[heure][i][1] > 10:
                planning.liste_eleve[heure][i][1] = 1
            break
    ajouteleve()

def ecrire_excel_ref(jour):
    global planning_theme
    #ouverture du fichier reference
    workbook = load_workbook(path_parametre+jour+'.xlsx')
    feuille = workbook.active

    # recuperer la mise en forme du fichier reference
    taillecellule = feuille.column_dimensions['B'].width
    hauteurcellule = feuille.row_dimensions[4].height
    taille_police = feuille.cell(4, 2).font.size

    # creation d'un nouveau fichier excel
    workbook = Workbook()
    sheet = workbook.active
    
    #lecture des fichiers parametres (a revoir pour les supprimer)
    # if jour != 'semaine':
    #     path_che = path_cheval
    # else:
    #     path_che = path_cheval_semaine
    
    # for i in lire_fichier_chevaux(path_che):
    #     dict_cheval[i[1]] = [0, i[0]+3]
        
    # 
    # data2 = lire_fichier_cavalier(jour)
    print(dict_cheval[jour])
    liste_cheval = list(dict_cheval[jour].keys())
    data2 = dict_eleve[jour]
    cle2 = list(data2.keys())
    
    #tri des heures (a revoir pour les supprimer)
    if jour != "Semaine":
        cle2 = sorted(cle2, key=cmp_heure)
    elif jour == "Semaine":
        cle2 = sorted(cle2, key=cmp_heure_semaine)
    lignes = len(liste_cheval)+4
    colonnes = len(cle2)+2
    
    dico_numeros = {
        1: 'A',
        2: 'B',
        3: 'C',
        4: 'D',
        5: 'E',
        6: 'F',
        7: 'G',
        8: 'H',
        9: 'I',
        10: 'J',
        11: 'K',
        12: 'L',
        13: 'M',
        14: 'N',
        15: 'O',
        16: 'P',
        17: 'Q',
        18: 'R',
        19: 'S',
        20: 'T',
        21: 'U',
        22: 'V',
        23: 'W',
        24: 'X',
        25: 'Y',
        26: 'Z'
    }

    double = Side(border_style="thin", color="000000")
    
    ind_ligne = 0
    for ligne in range(3, lignes):
        ind_colonne = 0
        #mise en forme de la heuteur des lignes
        for colonne in range(1, colonnes):
            print(colonne)
            #mise en forme des cellules
            sheet.cell(ligne, colonne).font = Font(size=taille_police)
            sheet.cell(ligne, colonne).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(ligne, colonne).border = Border(left=double, top=double, right=double, bottom=double)
            #mise en forme des heures
            if ligne == 3 :
                #mise en forme de la largeur des lignes
                sheet.column_dimensions[dico_numeros[colonne]].width = taillecellule
                if colonne != 1:
                    #ajout des heures
                    sheet.cell(ligne, colonne).value = cle2[ind_colonne]
                    ind_colonne+=1
                if colonne % 2 == 0:
                    #creation des heures vertes
                    sheet.cell(ligne, colonne).fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            #creation des lignes vertes        
            if ligne % 2 == 0:
                sheet.cell(ligne, colonne).fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
            #mise en forme des chevaux
            if colonne == 1 :
                sheet.row_dimensions[ligne].height = hauteurcellule
                if ligne != 3:
                    #ajout des chevaux
                    sheet.cell(ligne, colonne).value = liste_cheval[ind_ligne]
                    ind_ligne+=1
    
    #ajout du titre theme
    sheet.cell(lignes, 1).value = "theme"

    #sauvegarde du fichier reference
    workbook.save(path_parametre+jour+'.xlsx')

def heure_precedant():
    """
    Définit l'heure précédente pour la cellule et appelle la fonction changer_heure.

    Cette fonction met à jour l'heure de la cellule en la définissant comme étant l'heure
    précédente par rapport à la liste des heures de travail du planning. Elle utilise la
    fonction changer_heure pour mettre à jour l'interface utilisateur.

    Args:
        Aucun.

    Returns:
        Aucun.
    """

    liste_heure = list(planning.liste_eleve)

    for i in range(1, len(liste_heure)):
        if liste_heure[i] == cellule.heure:
            cellule.set_heure(liste_heure[i-1])
            changer_heure()


def heure_suivant():
    """
    Définit l'heure suivante pour la cellule et appelle la fonction changer_heure.

    Cette fonction met à jour l'heure de la cellule en la définissant comme étant l'heure
    suivante par rapport à la liste des heures de travail du planning. Elle utilise la fonction
    changer_heure pour mettre à jour l'interface utilisateur.

    Args:
        Aucun.

    Returns:
        int: 0 si l'heure suivante est définie avec succès, sinon rien.
    """
    liste_heure = list(planning.liste_eleve)
    for i in range(0, len(liste_heure)-1):
        if liste_heure[i] == cellule.heure:
            cellule.set_heure(liste_heure[i+1])
            changer_heure()
            return 0


def ajouter_rattrapage():
    planning.liste_eleve[cellule.heure].append((eleve_rattrapage.get().upper(),-1))
    ajouteleve()
    
def ecrire_word():
    err=False
    try:
        theme_t = [planning_theme,planning_theme1,planning_theme2,planning_theme3]
        eleves = lire_fichier_cavalier(jour.j)#remplacer par planning.liste_eleve
        word(jour.j, nom_fichier, planning,theme_t,user,eleves)
    except Exception as e:
        err = True
        messagebox.showerror("Erreur", f"Erreur lors de la création des fichiers Word : {e}")
    if not err:
        messagebox.showinfo("Information", "Les fichiers Word ont été créé avec succès")

def ajouter_theme():
    global planning_theme
    theme.set(theme_entry.get())
    planning_theme[cellule.heure] = theme.get()

def ajouter():
    """
    Ajoute une cellule au planning avec des vérifications et met à jour l'interface.

    Cette fonction ajoute une cellule au planning en utilisant la méthode ajout de la classe
    Planning. Elle effectue des vérifications préliminaires et met à jour l'interface utilisateur
    en conséquence. Elle ajoute également une entrée à l'historique du planning en cas de succès.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    err = planning.ajout(cellule)
    if err == None or err == -4:
        if elevecarte == True:
            unesessionmoins(cellule.eleve,cellule.heure)
        if cellule.ind_eleve != -1:
            colorier_eleve(cellule.ind_eleve)
        ajoutuncheval(cellule.cheval, cellule.ind_cheval)
        affichage_txt(jour, planning)
        label_enregistrer.config(fg="#b4b4b4")
        colorier_chevaux()
        inserer_liste_de_travaille()
        ajout_historique(
            "ajout", (cellule.heure, cellule.cheval, cellule.eleve))
    elif err == -1:
        msgbox = tk.messagebox.showerror(
            title="creation de fichier", message="vous n'avez pas selectionné toutes les informations necessaire à l'ajout")
    elif err == -2:
        msgbox = tk.messagebox.showerror(
            title="creation de fichier", message="Ne peux etre ajouter car ce cheval travaille deja durant cette heure")
    elif err == -3:
        msgbox = tk.messagebox.showerror(
            title="creation de fichier", message="Ne peux etre ajouter car ce cheval travaille deja 4 heure dans la journée")
    elif err == -5:
        print(f"Erreur numéro {err}, ajout annulé.")


def supprimer():
    """
    Supprime une cellule du planning avec des vérifications et met à jour l'interface.

    Cette fonction supprime une cellule du planning en utilisant la méthode supprime de la classe
    Planning. Elle effectue des vérifications préliminaires et met à jour l'interface utilisateur
    en conséquence. Elle ajoute également une entrée à l'historique du planning en cas de succès.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    err = planning.supprime(cellule)
    if err == None:
        if elevecarte == True:
            unesessionplus(cellule.eleve,cellule.heure)
        if cellule.ind_eleve != -1:
            decolorier_eleve(cellule.ind_eleve)
        ajoutuncheval(cellule.cheval, cellule.ind_cheval)
        affichage_txt(jour, planning)
        label_enregistrer.config(fg="#b4b4b4")
        colorier_chevaux()
        inserer_liste_de_travaille()
        ajout_historique("suppression", (cellule.heure,
                         cellule.cheval, cellule.eleve))
    else:
        msgbox = tk.messagebox.showerror(
            title="creation de fichier", message="suppression impossible vous voulez supprimez un element qui n'existe pas")


def inserer_liste_de_travaille():
    """
    Insère les heures de travail dans la listebox et met en évidence les heures travaillées.

    Cette fonction vide d'abord la listebox des heures de travail, puis insère les heures de
    travail associées au cheval de la cellule dans la listebox. Elle met également en évidence
    les heures travaillées dans la listebox.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    vider_listebox(heure_listebox)
    liste = planning.heure_travailler(cellule.cheval)
    for i in liste:
        heure_listebox.insert(tk.END, i)


def colorier_eleve(ind):
    """
    Change la couleur de fond de l'élément d'index donné dans la listebox des élèves en rouge.

    Cette fonction prend en argument l'index d'un élément dans la listebox des élèves et change
    la couleur de fond de cet élément en rouge.

    Args:
        ind (int): L'index de l'élément dans la listebox des élèves.

    Returns:
        Aucun.
    """
    eleve_listbox.itemconfig(ind, {'bg': 'red'})


def decolorier_eleve(ind):
    """
    Change la couleur de fond de l'élément d'index donné dans la listebox des élèves en blanc.

    Cette fonction prend en argument l'index d'un élément dans la listebox des élèves et change
    la couleur de fond de cet élément en blanc.

    Args:
        ind (int): L'index de l'élément dans la listebox des élèves.

    Returns:
        Aucun.
    """
    eleve_listbox.itemconfig(ind, {'bg': 'white'})


def ajout_historique(type, element):
    """
    Ajoute un élément à l'historique du planning et met à jour l'affichage dans la zone de texte.

    Cette fonction prend en argument un type d'action (ajout ou suppression) et un élément à ajouter
    à l'historique du planning. Elle met à jour l'affichage de l'historique dans la zone de texte
    correspondante.

    Args:
        type (str): Le type d'action (ajout ou suppression).
        element (tuple): L'élément à ajouter à l'historique.

    Returns:
        Aucun.
    """
    historique.config(state='normal')
    planning.append_historique(type, element)
    historique.delete('1.0', END)
    for i in planning.historique:
        historique.insert("1.0", f"{i}\r\n")
    historique.config(state='disabled')


def affichage_txt(jour, planning):
    """
    Affiche le planning du jour dans la zone de texte correspondante.

    Cette fonction prend en argument un objet "jour" et un objet "planning", puis affiche le planning
    du jour dans la zone de texte correspondante.

    Args:
        jour (object): L'objet jour correspondant au jour de la semaine.
        planning (object): L'objet planning contenant les données du planning.

    Returns:
        Aucun.
    """
    visu_fichier.config(state='normal')
    visu_fichier.delete('1.0', END)
    visu_fichier.insert(END, planning.fichier(jour.j))
    visu_fichier.config(state='disabled')


def vider_listebox(listebox):
    """
    Vide une listebox en supprimant tous ses éléments.

    Cette fonction prend en argument une listebox et supprime tous les éléments qu'elle contient.

    Args:
        listebox: La listebox à vider.

    Returns:
        Aucun.
    """
    if listebox.size() > 0:
        listebox.delete(0, listebox.size())


def inserer_listbox(i):
    """
    Insère un élément dans la listebox des élèves et le met en évidence s'il est présent dans le planning.

    Cette fonction prend en argument un élément "i" et l'insère dans la listebox des élèves. Si l'élément
    est présent dans le planning, il est mis en évidence en rouge.

    Args:
        i (str): L'élément à insérer dans la listebox des élèves.

    Returns:
        Aucun.
    """
    if i[1] == -1:
        i = i[0]
    eleve_listbox.insert(tk.END, i)
    try:
        entier_valeur = int(i[1])
        i = i[0]
    except ValueError:
        pass
    # if int(i[1]) <= 10 and int(i[1]) >= 0:
    #     i = i[0]
    if len(planning.planning) != 0:
        present = any((cellule.heure, i) == (tup[0], tup[2])
                      for tup in planning.planning)
        if present:
            eleve_listbox.itemconfig(tk.END, {'bg': 'red'})


def ajouteleve():
    """
    Vide la listebox des élèves et insère les éléments correspondant à l'heure de la cellule.

    Cette fonction vide d'abord la listebox des élèves, puis insère les éléments correspondant à
    l'heure de la cellule. Elle met également en évidence les élèves qui sont présents dans le planning.

    Args:
        Aucun.

    Returns:
        Aucun.
    """

    vider_listebox(eleve_listbox)
    if cellule.heure != 'heure':
        for i in planning.liste_eleve[cellule.heure]:
            inserer_listbox(i)


def ajoutuncheval(cheval, ind):
    """
    Supprime un élément de la listebox des chevaux et insère un nouvel élément à la position donnée.

    Cette fonction prend en argument le nom d'un cheval et sa position dans la listebox des chevaux.
    Elle supprime l'élément existant à cette position et insère un nouvel élément représentant le cheval.

    Args:
        cheval (str): Le nom du cheval.
        ind (int): La position dans la listebox.

    Returns:
        Aucun.
    """
    cheval_listbox.delete(ind)
    cheval_listbox.insert(
        ind, (planning.cheval[cheval][1], cheval))
    cheval_listbox.itemconfig(ind, {'bg': 'green'})


def ajoutcheval():
    """
    Vide la listebox des chevaux et insère les éléments correspondant à tous les chevaux du planning.

    Cette fonction vide d'abord la listebox des chevaux, puis insère les éléments correspondant à tous
    les chevaux du planning. Elle met en évidence les chevaux en vert s'ils sont associés à des élèves
    présents à l'heure de la cellule.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    cheval_listbox.delete(0, END)
    for i in planning.cheval:
        cheval_listbox.insert(tk.END, (planning.cheval[i][1], i))
    if cellule.heure != "heure" and cellule.heure in planning.liste_heure:
        colorier()


def colorier():
    """
    Change la couleur de fond des éléments de la listebox des chevaux en fonction de leur disponibilité.

    Cette fonction met en évidence les chevaux en vert s'ils sont associés à des élèves présents à
    l'heure de la cellule. Les autres chevaux ont un fond blanc.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    for i in range(0, len(planning.cheval)):
        cheval_listbox.itemconfig(
            i, {'bg': 'white'})
    setcheval = set()
    for i in planning.liste_eleve[cellule.heure]:
        if len(i)==2:
            ancient = planning.ancient_cheval_de(i[0], cellule.heure)
        else:
            ancient = planning.ancient_cheval_de(i, cellule.heure)
        print("ancien",ancient)
        for y in ancient:
            if y[1] != "":
                setcheval.add(y[1])
    print(setcheval)
    for i in setcheval:
        print(i)
        cheval_listbox.itemconfig(
            i, {'bg': 'green'})


def colorier_ancient_chevaux(ancient_cheval_eleve):
    """
    Change la couleur de fond des chevaux anciens en fonction de leur disponibilité.

    Cette fonction prend en argument une liste d'éléments représentant des chevaux anciens associés à
    un élève. Elle change la couleur de fond des chevaux en fonction de leur disponibilité, en utilisant
    des couleurs telles que jaune, orange et rouge en fonction du nombre d'anciens chevaux.

    Args:
        ancient_cheval_eleve (list): Une liste d'éléments représentant des chevaux anciens.

    Returns:
        Aucun.
    """

    if len(ancient_cheval_eleve) >= 3 and ancient_cheval_eleve[2][1] != "":
        cheval_listbox.itemconfig(
            ancient_cheval_eleve[2][1], {'bg': 'yellow'})
    if len(ancient_cheval_eleve) >= 2 and ancient_cheval_eleve[1][1] != "":
        cheval_listbox.itemconfig(
            ancient_cheval_eleve[1][1], {'bg': 'orange'})
    if len(ancient_cheval_eleve) >= 1 and ancient_cheval_eleve[0][1] != "":
        cheval_listbox.itemconfig(
            ancient_cheval_eleve[0][1], {'bg': 'red'})

def colorier_chevaux():
    if jour.j != "Semaine":
        for cell in planning.planning:
            if cell[0][0:2] in cellule.heure and cell[1] in planning.cheval:
                cheval_listbox.itemconfig(planning.cheval[cell[1]][0], {'bg': 'violet'})
    elif jour.j == "Semaine":
        
        for cell in planning.planning:
            print( cell[0][-4:-1] , cellule.heure)
            if cell[0][0:2] in cellule.heure and cell[1] in planning.cheval and cell[0][-4:-1] in cellule.heure:
                cheval_listbox.itemconfig(planning.cheval[cell[1]][0], {'bg': 'violet'})

def changer_heure():
    """
    Met à jour l'affichage en fonction de l'heure de la cellule.

    Cette fonction met à jour l'affichage en fonction de l'heure de la cellule, en vidant la listebox
    des élèves, en ajoutant les élèves correspondant à l'heure, en coloriant les chevaux disponibles,
    et en mettant à jour les variables d'interface utilisateur.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    ajouteleve()
    colorier()
    colorier_chevaux()
    varheure.set(f"{cellule.heure}")
    varajout.set(cellule.getCellule())
    if cellule.heure in planning_theme:
        theme.set(planning_theme[cellule.heure])
    else:
        theme.set("theme")
    if cellule.heure in planning_theme1:
        theme1.set(planning_theme1[cellule.heure])
    else:
        theme1.set("theme1")
    if cellule.heure in planning_theme2:
        theme2.set(planning_theme2[cellule.heure])
    else:
        theme2.set("theme2")
    if cellule.heure in planning_theme3:
        theme3.set(planning_theme3[cellule.heure])
    else:
        theme3.set("theme3")



def changement_heure(i):
    """
    Change l'heure de la cellule et met à jour l'affichage.

    Cette fonction prend en argument une nouvelle heure "i", change l'heure de la cellule en conséquence
    et met à jour l'affichage.

    Args:
        i (str): La nouvelle heure pour la cellule.

    Returns:
        Aucun.
    """
    cellule.set_heure(i)
    changer_heure()


def cmp_dates(d):
    """
    Compare deux dates au format "jour-mois-année".

    Cette fonction prend en argument une date au format "jour-mois-année" et retourne un tuple (année, mois, jour)
    pour permettre une comparaison correcte des dates.

    Args:
        d (str): Une date au format "jour-mois-année".

    Returns:
        tuple: Un tuple (année, mois, jour).
    """

    j, m, a = d[0].split("-")

    return (int(a), int(m), int(j))


def recupperation_excel(name):
    """
    Récupère les données depuis un fichier Excel et les organise en listes et dictionnaires.

    Cette fonction prend en argument une liste "listeself" (non utilisée dans la fonction) et le nom d'un
    fichier Excel "name". Elle extrait les données de ce fichier, les organise en listes et dictionnaires
    (liste, dict_cheval, dict_heure) et les renvoie.

    Args:
        listeself: Liste (non utilisée dans la fonction).
        name (str): Le nom du fichier Excel à lire.

    Returns:
        tuple: Un tuple contenant trois éléments :
            - Une liste de tuples (heure, cheval, élève).
            - Un dictionnaire des chevaux avec leur indice et ligne.
            - Un dictionnaire des heures avec leur libellé.
    """
    workbook = load_workbook(name)
    sheet = workbook.active
    liste = []
    planning_theme = {}
    dict_heure = {}
    dict_cheval = {}
    heure_temp = {}
    Nb = 0
    print(len(sheet["A"]))
    for i in range(3, len(sheet["A"])+1):
        print(i,str(sheet.cell(row=i, column=1).value).strip())
        for j in list(range(1, len(sheet[3])+1))[::-1]:
            valeur_case = str(sheet.cell(row=i, column=j).value).strip()
            
            if valeur_case != 'None':
                if i == 3:
                    dict_heure[valeur_case] = j
                    heure_temp[j] = valeur_case
                elif j == 1:
                    dict_cheval[valeur_case] = [Nb,i]
                    Nb = 0
                elif j > 1 and valeur_case != "MERCREDI" and valeur_case != "SAMEDI" and sheet.cell(row=3, column=j).value != None and sheet.cell(row=i, column=1).value != None and str(sheet.cell(row=i, column=1).value).strip() != "theme":
                    Nb += 1
                    liste.append(
                        (sheet.cell(row=3, column=j).value.strip(), sheet.cell(row=i, column=1).value.strip(), valeur_case))
                elif j>1 and str(sheet.cell(row=i, column=1).value).strip() == "theme" :
                    planning_theme[heure_temp[j].upper()] = valeur_case
                    print(i,j,heure_temp[j],valeur_case)
    return liste, dict_cheval, dict_heure,planning_theme

def sort_files_by_date(files):
        # Trie les fichiers par date (les plus récents en premier)
        sorted_files = sorted(files, key=lambda x: x[1], reverse=True)
        return sorted_files

def extract_date_from_filename(filename):
    # Extrait la date du nom du fichier
    date_str = filename.split()[-1].split('.')[0]
    file_date = datetime.strptime(date_str, "%d-%m-%Y")
    return file_date
    
def recup_donne():
    """
    Récupère les données depuis un fichier Excel et initialise le planning.

    Cette fonction affiche une boîte de dialogue pour sélectionner un fichier Excel, puis utilise la
    fonction "recupperation_excel" pour extraire les données du fichier, les initialise dans l'objet
    "planning" et met à jour l'affichage.

    Args:
        Aucun.

    Returns:
        Aucun.
    """

    global ancient_nom,nom_fichier,planning_theme,planning_theme1,planning_theme2,planning_theme3
    interface_default()
    msgbox = tk.messagebox.showinfo(
        title="Sélection de fichier", message="Veuillez sélectionner le fichier que vous souhaitez remplir")
    chemin = tk.Tk()
    chemin.withdraw()                 # pour ne pas afficher la fenêtre Tk
    path = askopenfilename()
    nom_fichier=[]
    folder = path.split('/')
    name = folder[-1]
    nom_fichier.append(name)
    print(name)
    if "mercredi" in name.lower():
        jour.set_mercredi()
    elif "samedi" in name.lower():
        jour.set_samedi()
    elif "semaine" in name.lower():
        jour.set_semaine()
    else :
        msgbox = tk.messagebox.showerror(
            title="Erreur de fichier", message="Le fichier n'est pas un fichier de planning")
        return
    planning_theme={}
    planning_theme1={}
    planning_theme2={}
    planning_theme3={}
    planning.cheval.clear()
    planning.liste_heure.clear()
    
    # temp_cheval = {}
    # for cheval in dict_cheval[jour.j]:
    #     temp_cheval[cheval[1]] = 0
    # planning.set_cheval(temp_cheval)

    planning.set_liste_eleve(dict_eleve[jour.j])
    print(planning.liste_eleve)
    varjour.set(jour.j)
    
    planning.set_nom_fichier(path)

    dict_planning, cheval, heure,planning_theme = recupperation_excel(path)
    # for i in planning.cheval.keys():
    #     if i in cheval:
    #         #mise a jour des valeur des chevaux
    #         dict_cheval[i] = cheval[i][0]
            
    
    planning.set_heure(heure)
    planning.set_planning(dict_planning)
    planning.set_cheval(remplir_cheval(dict_cheval[jour.j]))
    ajoutcheval()
    affichage_txt(jour, planning)
    
    liste = []
    path = path.replace(name, "")
    files = []
    for file in os.listdir(path):
        if jour.j.lower() in file.lower() and "~$" not in file.lower() and ".xlsx" in file.lower():
            files.append((file, extract_date_from_filename(file)))
    liste = sort_files_by_date(files)
    for i in range(len(liste)):
            if liste[i][0] == name:
                selected_ind = i
    nb_fichier=4
    if i-selected_ind < 3:
        nb_fichier = i-selected_ind
    liste = [item[0] for item in liste[selected_ind+1:selected_ind+nb_fichier]]
    
    if len(liste) > 0:
        varsemaine1.set(liste[0].replace(".xlsx", ""))
        ancient_nom = path + liste[0]
        nom_fichier.append(liste[0])
        ancient_planning, x, y,planning_theme1 = recupperation_excel( path + liste[0])
        planning.set_ancien_planning(ancient_planning)
    if len(liste) > 1:
        varsemaine2.set(liste[1].replace(".xlsx", ""))
        nom_fichier.append(liste[1])
        ancient_planning2, x, y,planning_theme2 = recupperation_excel( path +liste[1])
        planning.set_ancien_planning2(ancient_planning2)
    if len(liste) > 2:
        varsemaine3.set(liste[2].replace(".xlsx", ""))
        nom_fichier.append(liste[2])
        ancient_planning3, x, y,planning_theme3 = recupperation_excel(path +liste[2])
        planning.set_ancien_planning3(ancient_planning3)
    
    msgbox = tk.messagebox.showinfo(
        title="Création de fichier", message="Tous les fichiers ont été récupérés")
    ajout_des_commande_lena()
    liste_heure = list(planning.liste_eleve.keys())
    changement_heure(liste_heure[0])

def ecrire_fichier():
    """
    Enregistre les données dans un fichier Excel.

    Cette fonction ouvre le fichier Excel actuellement en cours de modification, efface les données des
    cellules correspondant au planning, puis insère les nouvelles données du planning. Elle enregistre
    ensuite le fichier Excel.

    Args:
        Aucun.

    Returns:
        Aucun.
    """
    global planning_theme

    wb = load_workbook(path_parametre+jour.j+'.xlsx')

    # Accéder à la feuille de calcul souhaitée
    feuille = wb.active

    # Parcourir les lignes et les colonnes de la feuille de calcul

    largeur = feuille.column_dimensions['B'].width
    hauteur = feuille.row_dimensions[4].height
    taille_police = feuille.cell(4, 2).font.size

    print(largeur,hauteur)
    wb.save(jour.j+'.xlsx')
    workbook = Workbook()
    sheet = workbook.active
    
    # dictio_cheval = {}
    # for i in dict_cheval[jour.j]:
    #     dictio_cheval[i[1]] = [0, i[0]+3]
    print(dict_cheval[jour.j])
    liste_cheval = list(dict_cheval[jour.j].keys())
    data2 = dict_eleve[jour.j]
    cle2 = list(data2.keys())
    
    if jour.j != "Semaine":
        cle2 = sorted(cle2, key=cmp_heure)
    elif jour.j == "Semaine":
        cle2 = sorted(cle2, key=cmp_heure_semaine)
    
    lignes = len(liste_cheval)+4
    colonnes = len(cle2)+2

    dico_numeros = {
        1: 'A',
        2: 'B',
        3: 'C',
        4: 'D',
        5: 'E',
        6: 'F',
        7: 'G',
        8: 'H',
        9: 'I',
        10: 'J',
        11: 'K',
        12: 'L',
        13: 'M',
        14: 'N',
        15: 'O',
        16: 'P',
        17: 'Q',
        18: 'R',
        19: 'S',
        20: 'T',
        21: 'U',
        22: 'V',
        23: 'W',
        24: 'X',
        25: 'Y',
        26: 'Z'
    }
    # effacer excel
    # vert ligne = #A9D08E
    # vert heure = #70AD47
    # {'thin', 'slantDashDot', 'dotted', 'mediumDashDotDot', 'double', 'medium', 'thick', 'mediumDashed', 'dashed', 'dashDotDot', 'hair', 'mediumDashDot', 'dashDot'}
    taillecellule = largeur
    hauteurcellule = hauteur
    double = Side(border_style="thin", color="000000")
    sans = Side(border_style=None, color='FF000000')
    ind_ligne = 0
    for ligne in range(3, lignes):
        ind_colonne = 0
        for colonne in range(1, colonnes):
            #mise en forme des cellules
            sheet.cell(ligne, colonne).font = Font(size=taille_police)
            sheet.cell(ligne, colonne).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(ligne, colonne).border = Border(left=double, top=double, right=double, bottom=double)
            #mise en forme des heures
            if ligne == 3 :
                #mise en forme de la largeur des lignes
                sheet.column_dimensions[dico_numeros[colonne]].width = taillecellule
                if colonne != 1:
                    #ajout des heures
                    sheet.cell(ligne, colonne).value = cle2[ind_colonne]
                    ind_colonne+=1
                if colonne % 2 == 0:
                    #creation des heures vertes
                    sheet.cell(ligne, colonne).fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            #creation des lignes vertes        
            if ligne % 2 == 0:
                sheet.cell(ligne, colonne).fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
            #mise en forme des chevaux
            if colonne == 1 :
                sheet.row_dimensions[ligne].height = hauteurcellule
                if ligne != 3:
                    #ajout des chevaux
                    sheet.cell(ligne, colonne).value = liste_cheval[ind_ligne]
                    ind_ligne+=1

    dict_heure = {}
    Nb = 0
    for i in cle2:
        dict_heure[i] = Nb+2
        Nb += 1
    planning.set_heure(dict_heure)

    for i in planning.planning:
        if i[1] in planning.cheval and i[0] in planning.liste_heure:
            # if sheet.cell(planning.cheval[i[1]][1], planning.liste_heure[i[0]]).value == None:
            sheet.cell(planning.cheval[i[1]][0]+4,
                       planning.liste_heure[i[0]]).value = i[2]
    for i in planning_theme:
        sheet.cell(len(liste_cheval)+4, dict_heure[i]).value = planning_theme[i]
        # sheet.cell(len(liste_cheval)+4, dict_heure[i]).font = Font(size=taille_police)
    sheet.cell(len(liste_cheval)+4, 1).value = "theme"
    # sheet.cell(len(liste_cheval)+4, 1).font = Font(size=taille_police)

    err = workbook.save(planning.name_fichier)
    if err == None:
        label_enregistrer.config(fg="#ffffff")
        cavalier = lire_fichier_cavalier(jour.j)
        # print(cavalier)
        fichier = open(path_parametre+"liste_cavalier_" + jour.j + ".txt", "w")
        fichier.write(ecrire_fichier_cavalier(cavalier,carte=True))
        fichier.close()


def add_heure():
    global dict_eleve
    dict_eleve[parajour][para_input_heure.get().upper()] = []
    para_inserer_listebox(dict_eleve[parajour])
    visualiser_fichier_cavalier(dict_eleve[parajour])

def suppr_heure(dict_eleve,heure):
    dict_eleve[parajour] = param.suppr_heure(dict_eleve[parajour], heure)
    para_inserer_listebox(dict_eleve[parajour])
    visualiser_fichier_cavalier(dict_eleve[parajour])

def add_eleve():
    global dict_eleve
    global heure
    nb_carte = -1
    if v.get() == 1:
        nb_carte = para_nbcarte.get()
    dict_eleve[parajour][heure].append((para_input_eleve.get().upper().strip(),nb_carte))
    remplirlisteboxeleve()
    visualiser_fichier_cavalier(dict_eleve[parajour])

def add_cheval():
    liste_cheval = []
    for cheval in dict_cheval[parajour]:
        liste_cheval.append([dict_cheval[parajour][cheval][0],cheval])
    
    if not para_input_ind_chevaux.get():
        ind_cheval = len(liste_cheval)
    else:
        ind_cheval = int(para_input_ind_chevaux.get())
    
    present = False
    if para_input_chevaux.get() in dict_cheval[parajour]:
        messagebox.showerror(
            "Erreur", "Le cheval est déjà présent dans la liste")
        present = True
    
    print(dict_cheval[parajour])
    print(liste_cheval,ind_cheval)
    print(present)
    if not present and len(liste_cheval) >= ind_cheval:
        for cheval in liste_cheval:
            if cheval[0] >= ind_cheval:
                cheval[0] += 1
        liste_cheval.append([ind_cheval,
                        para_input_chevaux.get().upper()])
        liste_cheval.sort()
        print(liste_cheval)
        for i in liste_cheval:
            if i[1] in dict_cheval[parajour]:
                dict_cheval[parajour][i[1]] = [i[0], dict_cheval[parajour][i[1]][1]]
            else:
                dict_cheval[parajour][i[1]] = [i[0], planning.nb_heure(i[1])]
        remplirlisteboxcheval(liste_cheval)

def suppr_cheval():
    print(cheval)
    liste_cheval = []
    for chevali in dict_cheval[parajour]:
        liste_cheval.append([dict_cheval[parajour][chevali][0],chevali])
    print(dict_cheval[parajour])
    if cheval in liste_cheval:
        liste_cheval.remove(cheval)
        for che in liste_cheval:
            if che[0] >= cheval[0]:
                che[0] -= 1
        liste_cheval.sort()
        print(liste_cheval)
        for i in liste_cheval:
            dict_cheval[parajour][i[1]] = [i[0], dict_cheval[parajour][i[1]][1]]
        del dict_cheval[parajour][cheval[1]]
        print(dict_cheval[parajour])
        remplirlisteboxcheval(liste_cheval)
    
def suppr_eleve():
    liste = []
    for eleves in dict_eleve[parajour][heure]:
        print(eleves, eleve)
        if (eleves[0],eleves[1]) != (eleve[0],eleve[1]):
            print(eleves, eleve)
            liste.append(eleves)
    dict_eleve[parajour][heure] = liste[:]
    remplirlisteboxeleve()
    visualiser_fichier_cavalier(dict_eleve[parajour])
        
def ecrire_fichier_cavalier(liste_eleve={},carte =False):
    txt = ""
    if liste_eleve == {}:
        cle = list(planning.liste_eleve.keys())
        eleves = planning.liste_eleve
        if jour.j != "semaine":
            cle = sorted(cle, key=cmp_heure)
        elif jour.j == "semaine":
            cle = sorted(cle, key=cmp_heure_semaine)
    else:
        cle = list(liste_eleve.keys())
        eleves = liste_eleve
        if parajour != "Semaine":
            cle = sorted(cle, key=cmp_heure)
        elif parajour == "Semaine":
            cle = sorted(cle, key=cmp_heure_semaine)

    print("cle",cle)
    for heure in cle:
        ind=0
        for eleve in eleves[heure]:
            if carte and eleve[1] != -1:
                txt += planning.liste_eleve[heure][ind][0]+"/"+ str(planning.liste_eleve[heure][ind][1]) + "\r"
            else:
                txt += eleve[0]+"/"+ str(eleve[1]) + "\r"
            ind+=1
        txt += "\\heure/" + heure + "\r"
    txt += "\\Fin fichier/"
    return txt
        


def ecrire_fichier_cheval(dict_chevaux):
    liste_cheval = []
    for cheval in dict_chevaux:
        liste_cheval.append([dict_chevaux[cheval][0],cheval])
    txt = ""
    for cheval in liste_cheval:
        txt += str(cheval[0]) + '\t' + cheval[1] + "\r"
    return txt

def para_enregistrer():
    global parajour,user,dict_cheval,dict_eleve
    err=False
    # try:
    if para_listeCombo_user.get() != "" and para_listeCombo_user.get() == 'Lena' or para_listeCombo_user.get() == 'Karine':
        with open(path_user, "w") as fichier:
            fichier.write(para_listeCombo_user.get())
            user = para_listeCombo_user.get()
            user_var.set(user)
        
    if parajour != '' and dict_eleve[parajour]:
        fichier = open(path_parametre+"liste_cavalier_" + parajour + ".txt", "w")
        fichier.write(ecrire_fichier_cavalier(dict_eleve[parajour]))
        fichier.close()
    if dict_cheval[parajour] and parajour!='Semaine':
        fichier = open(path_cheval, "w")
        fichier.write(ecrire_fichier_cheval(dict_cheval[parajour]))
        fichier.close()
    elif dict_cheval[parajour]:
        fichier = open(path_cheval_semaine, "w")
        fichier.write(ecrire_fichier_cheval(dict_cheval[parajour]))
        fichier.close()
    # dict_eleve, dict_cheval = lire_parametre()
    visualiser_fichier_cavalier(dict_eleve[parajour])
    print("enregistrer excel ref")
    ecrire_excel_ref("Mercredi")
    ecrire_excel_ref("Samedi")
    ecrire_excel_ref("Semaine")
    
    sauvegarder_mail()
        
    # except Exception as e:
        # err = True
        # messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement des paramètres : {e}")
    if not err:
        messagebox.showinfo("Enregistrement", "Les paramètres ont été enregistrés avec succès!")


def remplirlisteboxcheval(chevaux):
    para_listebox_chevaux.delete(0, END)
    for cheval in chevaux:
        para_listebox_chevaux.insert(END, cheval)


def remplirlisteboxeleve():
    para_listebox_eleve.delete(0, END)
    for eleve in dict_eleve[parajour][heure]:
        para_listebox_eleve.insert(END, eleve)


def interface_default():
    global parajour
    for widget in widgets_parametre:
        widget.place_forget()

    image_label.destroy()

    label_jour.place(x=133, y=140)
    label_heure.place(x=183, y=140)
    title_label.place(x=60, y=35)
    boutton_avancer_heure.place(x=65, y=140)
    boutton_reculer_heure.place(x=260, y=140)
    label_cavalier.place(x=460, y=70)
    label_cavalier2.place(x=460, y=100)
    label_cavalier3.place(x=650, y=100)
    label_cavalier6.place(x=460, y=150)
    label_cavalier4.place(x=650, y=150)
    label_cavalier7.place(x=460, y=200)
    label_cavalier5.place(x=650, y=200)
    boutton_absent.place(x=755, y=100)
    boutton_correction.place(x=810, y=100)
    eleve_listbox.place(x=133, y=160)
    eleve_rattrapage.place(x=133, y=360)
    label_eleve_rattrapage.place(x=137, y=330)
    boutton_eleve_rattrapage.place(x=160, y=390)
    cheval_listbox.place(x=330, y=35)
    visu_fichier.place(x=900, y=400)
    label_visu_fichier.place(x=900, y=370)
    label_ajout.place(x=460, y=400)
    boutton_ajouter.place(x=570, y=480)
    boutton_supprimer.place(x=670, y=480)
    boutton_enregistrer.place(x=570, y=530)
    label_enregistrer.place(x=560, y=585)
    label_heure_cheval.place(x=460, y=250)
    heure_listebox.place(x=460, y=280)
    historique.place(x=900, y=70)
    label_historique.place(x=900, y=40)
    label_user.place(x=60, y=70)
    listeCombo.place(x=65, y=100)
    image1.place(x=535, y=606)
    image2.place(x=70, y=606)
    image3.place(x=680, y=220)
    bouton_ouvrir_excel.place(x=1400, y=60)
    bouton_rafraichir.place(x=1400, y=100)
    
    label_theme.place(x=133, y=440)
    theme_entry.place(x=133, y=470)
    boutton_theme.place(x=140, y=500)
    label_theme_actuelle.place(x=160, y=530)
    label_theme_avant1.place(x=650, y=125)
    label_theme_avant2.place(x=650, y=175)
    label_theme_avant3.place(x=650, y=225)
    
    bouton_word.place(x=1400, y=140)
    bouton_mail.place(x=1400, y=180)

    if parajour == jour.j and jour.j != '':
        planning.set_liste_eleve(dict_eleve[jour.j])
        planning.set_cheval(dict_cheval[jour.j])
        ajoutcheval()
        ajouteleve()
        

    print(planning.liste_eleve)
    print(planning.cheval)

    # dict_cheval = {}

    # if parajour != "":
    #     for i in chevaux:
    #         if i[1] in planning.cheval:
    #             if i[0] == planning.cheval[i[1]][1]-3:
    #                 dict_cheval[i[1]] = planning.cheval[i[1]]
    #             else:
    #                 dict_cheval[i[1]] = [planning.nb_heure(i[1]), i[0]+3]
    #         else:

    #             dict_cheval[i[1]] = [planning.nb_heure(i[1]), i[0]+3]

    #     planning.set_cheval(dict_cheval)

    # if parajour == jour.j:
    #     planning.set_heure(dict_eleve[parajour])
    #     planning.set_liste_eleve(dict_eleve[parajour])
    # elif jour.j != '':
    #     planning.set_liste_eleve(lire_fichier_cavalier(jour.j))
    # else:
    #     planning.set_liste_eleve({})
    # if planning.liste_eleve != {}:
    #     ajouteleve()
    # if planning.cheval != []:
    #     ajoutcheval()
    #     para_enregistrer()


    # ajout_des_commande_lena()


def interface_paramete():
    for widget in widgets_principaux:
        widget.place_forget()

    image_label.destroy()

    para_image1.place(relx=0.48, rely=0.6, anchor=tk.CENTER)
    para_visu_fichier.place(x=900, y=400)
    # para_label_historique.place(x=1000, y=40)
    # para_historique.place(x=1000, y=70)
    para_listebox_heure.place(x=400, y=70)
    para_listebox_eleve.place(x=730, y=70)
    para_listeCombo.place(x=65, y=40)
    para_listeCombo_user.place(x=170, y=40)
    para_input_heure.place(x=560, y=140)
    para_add_heure.place(x=560, y=170)
    para_suppr_heure.place(x=560, y=200)
    para_input_eleve.place(x=890, y=140)
    para_add_eleve.place(x=890, y=170)
    para_suppr_eleve.place(x=890, y=200)
    para_boutton_enregistrer.place(x=635, y=680)
    para_listebox_chevaux.place(x=60, y=70)
    para_input_chevaux.place(x=220, y=140)
    para_add_chevaux.place(x=220, y=170)
    para_suppr_chevaux.place(x=220, y=200)
    para_input_ind_chevaux.place(x=360, y=140)
    para_case.place(x=890, y=230)
    para_nbcarte.place(x=890, y=260)
    # image1.place(x=1050, y=70)
    # image3.place(x=605, y=300)
    image3.place(x=170, y=500)

    image4.place(x=1070, y=70)
    
    posymail = 290
    para_label_mail_karine.place(x=730, y=posymail)
    para_entry_karine.place(x=820, y=posymail)
    para_label_mail_lena.place(x=730, y=posymail+30)
    para_entry_lena.place(x=820, y=posymail+30)
    # para_boutton_mail.place(x=835, y=posymail+60)
    
    
    para_bouton_importer_param.place(x=1300,y=290)
    para_bouton_exporter_param.place(x=1300,y=325)
    para_bouton_ouvrir_excel.place(x=1295, y=360)
    


def lire_fichier_chevaux(path):
    liste = {}
    fichier = open(path, "r")
    lignes = fichier.read()
    fichier.close()
    lignes = lignes.split("\n")
    for ligne in lignes:
        if ligne != '':
            if ligne[2] == '\t':
                cheval = ligne[3:].strip()
            else:
                cheval = ligne[2:].strip()
            liste[cheval] = ([int(ligne[:2]), planning.nb_heure(cheval)])
    return liste

def lire_fichier_cavalier(jour):
    liste = []
    liste_eleve = {}
    fichier = open(path_parametre+"liste_cavalier_" + jour + ".txt", "r")
    lignes = fichier.read()
    fichier.close()
    # para_visu_fichier.config(state='normal')
    # para_visu_fichier.delete('1.0', END)
    # para_visu_fichier.insert(END, lignes)
    # para_visu_fichier.config(state='disabled')
    
    lignes = lignes.split("\n")
    for ligne in lignes:
        if ligne != '':
            if "\\Fin fichier/" in ligne:
                return liste_eleve
            if "\\heure/" in ligne:
                liste_eleve[ligne[7:].strip()] = liste[:]
                liste.clear()
            else:
                nom, numero = ligne.strip().split('/')
                liste.append([nom,int(numero)])


def para_inserer_listebox(data):
    para_listebox_heure.delete(0, END)
    for i in list(data.keys()):
        para_listebox_heure.insert(END, i)


def mode_parametre():
    interface_paramete()


def mode_default():
    interface_default()


def visualiser_fichier_cavalier(data):
    cle = list(data.keys())
    if parajour != "Semaine":
        cle = sorted(cle, key=cmp_heure)
    elif parajour == "Semaine":
        cle = sorted(cle, key=cmp_heure_semaine)
    
    txt = ""
    for heure in data.keys():
        for eleve in data[heure]:
            txt += eleve[0]+"/"+ str(eleve[1]) + "\r\n"
        txt += "\\heure/" + heure + "\r\n"
    txt += "\\Fin fichier/"
    para_visu_fichier.config(state='normal')
    para_visu_fichier.delete('1.0', END)
    para_visu_fichier.insert(END, txt)
    para_visu_fichier.config(state='disabled')


def cmp_heure(d):
    if d[1].isdigit():
        heure = d[:2]
        if d[3] == '3':
            heure += '5'
        else:
            heure += '0'
    elif d[1] == 'H':
        heure = d[0]
        if d[2] == '3':
            heure += '5'
        else:
            heure += '0'
    else:
        heure = d[0]
    return (int(heure))

def cmp_heure_semaine(d):
    heure, jour = d.split()
    print(heure, jour)
    jours = {"LUNDI": 1, "MARDI": 2, "MERCREDI": 3, "JEUDI": 4, "VENDREDI": 5, "SAMEDI": 6, "DIMANCHE": 7}
    print(jours[jour], heure)
    return int(jours[jour]), heure


def prochain_jour(semaine, jour_actuel):
    jours_de_la_semaine = ["lundi", "mardi", "mercredi",
                           "jeudi", "vendredi", "samedi", "dimanche"]
    jour_cible = jours_de_la_semaine.index(semaine)

    jours_jusquau_prochain = (jour_cible - jour_actuel.weekday() + 7) % 7
    if jours_jusquau_prochain == 0:
        jours_jusquau_prochain = 7

    return jour_actuel + timedelta(days=jours_jusquau_prochain)


def add_centered_image(root, image_path, width, height):
    # Charger l'image
    original_image = Image.open(image_path)

    # Redimensionner l'image
    resized_image = original_image.resize((width, height), Image.NEAREST)
    photo = ImageTk.PhotoImage(resized_image)

    # Configurer l'image redimensionnée comme arrière-plan de la fenêtre
    image_label = tk.Label(root, borderwidth=0,
                           image=photo, highlightthickness=0, bg="#b4b4b4")
    image_label.image = photo
    image_label.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
    return image_label


def set_background(root, image_path):
    # Charger l'image
    original_image = Image.open(image_path)

    # Obtenir la taille de l'écran
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # Redimensionner l'image pour s'adapter à l'écran
    resized_image = original_image.resize((screen_width, screen_height), Image.NEAREST)

    photo = ImageTk.PhotoImage(resized_image)

    # Configurer l'image redimensionnée comme arrière-plan de la fenêtre
    background_label = tk.Label(root, image=photo, bg="#b4b4b4")
    background_label.image = photo
    background_label.place(x=0, y=0, relwidth=1, relheight=1)
    return background_label, photo


def nouveau_fichier():
    global pop
    pop = Toplevel(window)
    pop.title("Creation de excel")
    pop.geometry("350x150")

    date_actuelle = datetime.now()
    v = StringVar()
    date = StringVar()
    
    # Vérifier si aujourd'hui est déjà un mercredi
    if date_actuelle.weekday() == 1:  # Le code de semaine pour mercredi est 2 (0 pour lundi, 1 pour mardi, etc.)
        prochains_mardis = [date_actuelle]
        prochain_mardi = date_actuelle
    else:
        # Trouver le prochain mardi
        prochain_mardi = prochain_jour("mardi", date_actuelle)

        # Ajouter les trois prochains mardis au tableau
        prochains_mardis = [prochain_mardi]
    for _ in range(2):
        prochain_mardi = prochain_jour("mardi", prochain_mardi)
        prochains_mardis.append(prochain_mardi)



  

    # Vérifier si aujourd'hui est déjà un mercredi
    if date_actuelle.weekday() == 2:  # Le code de semaine pour mercredi est 2 (0 pour lundi, 1 pour mardi, etc.)
        prochains_mercredis = [date_actuelle]
        prochain_mercredi = date_actuelle
    else:
        # Trouver le prochain mercredi
        prochain_mercredi = prochain_jour("mercredi", date_actuelle)

        # Ajouter les trois prochains mercredis au tableau
        prochains_mercredis = [prochain_mercredi]
    for _ in range(2):
        prochain_mercredi = prochain_jour("mercredi", prochain_mercredi)
        prochains_mercredis.append(prochain_mercredi)

    if date_actuelle.weekday() == 5:  # Le code de semaine pour mercredi est 2 (0 pour lundi, 1 pour mardi, etc.)
        prochains_samedis = [date_actuelle]
        prochain_samedi = date_actuelle
    else:
        # Trouver le prochain mercredi
        prochain_samedi = prochain_jour("samedi", date_actuelle)

        # Ajouter les trois prochains mercredis au tableau
        prochains_samedis = [prochain_samedi]
    for _ in range(2):
        prochain_samedi = prochain_jour("samedi", prochain_samedi)
        prochains_samedis.append(prochain_samedi)

    pop_r1_1 = Radiobutton(pop, text=prochains_samedis[0].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_samedis[0].strftime("%d-%m-%Y"))
    pop_r1_2 = Radiobutton(pop, text=prochains_samedis[1].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_samedis[1].strftime("%d-%m-%Y"))
    pop_r1_3 = Radiobutton(pop, text=prochains_samedis[2].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_samedis[2].strftime("%d-%m-%Y"))

    pop_r2_1 = Radiobutton(pop, text=prochains_mercredis[0].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_mercredis[0].strftime("%d-%m-%Y"))
    pop_r2_2 = Radiobutton(pop, text=prochains_mercredis[1].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_mercredis[1].strftime("%d-%m-%Y"))
    pop_r2_3 = Radiobutton(pop, text=prochains_mercredis[2].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_mercredis[2].strftime("%d-%m-%Y"))

    pop_r3_1 = Radiobutton(pop, text=prochains_mardis[0].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_mardis[0].strftime("%d-%m-%Y"))
    pop_r3_2 = Radiobutton(pop, text=prochains_mardis[1].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_mardis[1].strftime("%d-%m-%Y"))
    pop_r3_3 = Radiobutton(pop, text=prochains_mardis[2].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_mardis[2].strftime("%d-%m-%Y"))
    pop_r3_2 = Radiobutton(pop, text=prochains_mardis[1].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_mardis[1].strftime("%d-%m-%Y"))
    pop_r3_3 = Radiobutton(pop, text=prochains_mardis[2].strftime("%d-%m-%Y"), variable=date,
                           value=prochains_mardis[2].strftime("%d-%m-%Y"))

    def affichejoursam():
        pop_r2_1.place_forget()
        pop_r2_2.place_forget()
        pop_r2_3.place_forget()
        pop_r3_1.place_forget()
        pop_r3_2.place_forget()
        pop_r3_3.place_forget()

        pop_r1_1.place(y=40, x=80)
        pop_r1_2.place(y=60, x=80)
        pop_r1_3.place(y=80, x=80)

    def affichejoursmer():
        pop_r1_1.place_forget()
        pop_r1_2.place_forget()
        pop_r1_3.place_forget()
        pop_r3_1.place_forget()
        pop_r3_2.place_forget()
        pop_r3_3.place_forget()
        # Vérifier si aujourd'hui est déjà un mercredi

        pop_r2_1.place(y=40, x=80)
        pop_r2_2.place(y=60, x=80)
        pop_r2_3.place(y=80, x=80)
        
    def affichejourmar():
        pop_r2_1.place_forget()
        pop_r2_2.place_forget()
        pop_r2_3.place_forget()
        pop_r1_1.place_forget()
        pop_r1_2.place_forget()
        pop_r1_3.place_forget()

        pop_r3_1.place(y=40, x=80)
        pop_r3_2.place(y=60, x=80)
        pop_r3_3.place(y=80, x=80)

    pop_label = Label(
        pop, text="voulez vous creer un fichier Mercredi ou Samedi")
    pop_label.place(x=40, y=20)

    v.set("Mercredi")  # initialiser
    pop_r1 = Radiobutton(pop, text="Mercredi", variable=v,
                         value="mercredi", command=affichejoursmer)
    pop_r1.place(x=10, y=40)
    pop_r2 = Radiobutton(pop, text="Samedi", variable=v,
                         value="samedi", command=affichejoursam)
    pop_r2.place(x=10, y=60)

    pop_r3 = Radiobutton(pop, text="Semaine", variable=v,
                         value="semaine", command=affichejourmar)
    pop_r3.place(x=10, y=80)
    
    def choix_date():
        pop.destroy()
        workbook = Workbook()
        name = askdirectory()
        workbook.save(name + '/liste ' + v.get() +
                      ' ' + date.get() + '.xlsx')

        recup_donne()
    pop_valider = Button(pop, text="Valider", command=choix_date)
    pop_valider.place(x=300, y=120)


def image(root, image_path, width, height):
    original_image = Image.open(image_path)

    # Redimensionner l'image
    resized_image = original_image.resize((width, height), Image.NEAREST)
    photo = ImageTk.PhotoImage(resized_image)

    # Configurer l'image redimensionnée comme arrière-plan de la fenêtre
    image_label = tk.Label(root, borderwidth=0,
                           image=photo, highlightthickness=0, bg="#b4b4b4")
    image_label.image = photo
    return image_label

def lire_parametre():
    
    dict_eleve = {"Semaine":lire_fichier_cavalier("semaine"), "Mercredi":lire_fichier_cavalier("mercredi"), "Samedi":lire_fichier_cavalier("samedi")}
    dict_cheval = {"Semaine":lire_fichier_chevaux(path_cheval_semaine), "Mercredi":lire_fichier_chevaux(path_cheval), "Samedi":lire_fichier_chevaux(path_cheval)}

    print("eleve : ",dict_eleve)
    print("cheval : ",dict_cheval)
    
    return dict_eleve, dict_cheval
    # global mail,dict_eleve[parajour],chevaux
    # if parajour != 'Semaine':
    #     path_che = path_cheval
    # else:
    #     path_che = path_cheval_semaine
    # dict_eleve[parajour] = lire_fichier_cavalier(parajour)
    # chevaux = lire_fichier_chevaux(path_che)
    # mail = get_mail()
    # remplirlisteboxcheval(chevaux)
    # para_inserer_listebox(dict_eleve[parajour])

def mettre_a_jour():
    global dict_eleve, dict_cheval,mail
    dict_eleve, dict_cheval = lire_parametre()

    mail = get_mail()
    if len(mail) > 0:
        para_entry_karine.delete(0, END)
        para_entry_karine.insert(0, mail[0])
    if len(mail) > 1:
        para_entry_lena.delete(0, END)
        para_entry_lena.insert(0, mail[1])
    return dict_eleve, dict_cheval

def importer_param():
    err = False
    try :
        chemin = askopenfilename()
        dezipper(chemin,path_parametre ,suppr_rep_destination=False)
        mettre_a_jour()
    except Exception as e:
        err = True
        messagebox.showerror("Erreur", f"Erreur lors de l'importation des paramètres : {e}")
    if not err:        
        messagebox.showinfo("importation de parametre", "Les paramètres ont été importés avec succès!")


    
def exporter_param():
    err = False
    try:
        nom_zip = 'parametre.zip'
        chemin = zip_fichiers(path_parametre, nom_zip)
        erreur =envoyer_email(user, chemin,nom_zip,"exportation des parametres de la version "+str(version),mail)
        if erreur != {}:
            msgbox = tk.messagebox.showerror(
            title="envoie des parametre par mail", message=erreur)
    except Exception as e:
        err = True
        messagebox.showerror("Erreur", f"Erreur lors de l'exportation des paramètres : {e}")
    if not err:
        messagebox.showinfo("exportation de parametre", "Les paramètres ont été exportés avec succès!")



# Importation des modules
cellule = Cellule()  # Création d'une instance de la classe Cellule
planning = Planning()  # Création d'une instance de la classe Planning
jour = Jour()  # Création d'une instance de la classe Jour
dict_eleve, dict_cheval = lire_parametre()

version = 1.8  # Version actuelle du programme
user = get_personne()
print(user)
mail = get_mail()

# Création de l'interface utilisateur
window = tk.Tk()  # Création de la fenêtre principale
window.title("Planning")  # Titre de la fenêtre
window.attributes('-fullscreen', True)  # Affichage en mode plein écran
# Permet de quitter en appuyant sur la touche "Échap"
window.bind('<Escape>', lambda e: window.destroy())

set_background(window, "image_fond.png")

widgets_principaux = []

widgets_parametre = []

ancient_nom = ""

image_label = add_centered_image(
    window, "0f382f680a13445c8e6484ecbbe2a2b5-transformed.png", 169*4, 166*4)

para_image1 = image(
    window, "image1.png", int(2388/5), int(1668/5))
image1 = image(window, "image1.png", int(2388/8.5), int(1668/8.5))
image2 = image(window, "image2.png", int(2388/8.5), int(1668/8.5))
image3 = image(window, "image3.png", int(2388/8.5), int(1668/8.5))
image4 = image(window, "image4.png", int(2388/7), int(1668/7))

label_version = tk.Label(window, text="Version " + str(version), bg='#b4b4b4')
label_version.place(x=1400, y=785)
# Création d'un cadre dans la fenêtre
# frame = tk.Frame(master=window, width=300, height=100)
# frame.pack()

# Définition des variables de contrôle
varheure = StringVar()
varjour = StringVar()
varcavalier = StringVar()
varcheval = StringVar()
varajout = StringVar()
varheure_cheval = StringVar()
varcavalier1 = StringVar()
varcavalier2 = StringVar()
varsemaine1 = StringVar()
varsemaine2 = StringVar()
varsemaine3 = StringVar()
theme1 = StringVar()
theme2 = StringVar()
theme3 = StringVar()
theme = StringVar()
v = IntVar ()
user_var = StringVar()
user_var.set(user)

label_jour = tk.Label(window, textvariable=varjour, bg='#b4b4b4')

label_heure = tk.Label(window, textvariable=varheure, bg='#b4b4b4')

label_user = tk.Label(window, textvariable=user_var,font=("Comic Sans MS", 15), bg='#b4b4b4')

# Création d'une étiquette pour le titre
title_label = tk.Label(
    window, text="GESTION PLANNING", font=("Comic Sans MS", 17), bg='#b4b4b4')

# Boutons pour avancer et reculer dans les heures
boutton_avancer_heure = tk.Button(
    window, width=8, bg='#8abd45', text="precedent", command=heure_precedant)

boutton_reculer_heure = tk.Button(
    window, width=8, bg='#8abd45', text="suivant", command=heure_suivant)

# Étiquettes pour afficher les informations du cavalier
label_cavalier = tk.Label(
    window, text="INFOS CAVALIER", font=("Corbel", 14), bg='#8abd45')

label_cavalier2 = tk.Label(
    window, textvariable=varsemaine1, font=("Corbel", 13), bg='#b4b4b4')

label_cavalier3 = tk.Label(
    window, textvariable=varcavalier, font=("Corbel", 13), bg='#b4b4b4')

label_cavalier6 = tk.Label(
    window, textvariable=varsemaine2, font=("Corbel", 13), bg='#b4b4b4')

label_cavalier4 = tk.Label(
    window, textvariable=varcavalier1, font=("Corbel", 13), bg='#b4b4b4')

label_cavalier7 = tk.Label(
    window, textvariable=varsemaine3, font=("Corbel", 13), bg='#b4b4b4')
label_cavalier5 = tk.Label(
    window, textvariable=varcavalier2, font=("Corbel", 13), bg='#b4b4b4')



def correction():
    global dernier_cheval
    plan = []
    plan.append((cellule.heure, cellule.cheval, cellule.eleve))
    for cel in planning.ancien_planning:
        if cel[0] == cellule.heure and cel[2] == cellule.eleve:
            pass
        else:
            plan.append(cel)
    planning.set_ancien_planning(plan)
    print(planning.ancien_planning)
    print(planning.liste_heure)
    
    workbook = load_workbook(ancient_nom)
    sheet = workbook.active
    if elevecarte == True and varcavalier.get() == "cheval":
        unesessionmoins(cellule.eleve,cellule.heure)


    data2 = dict_eleve[jour.j]
    cle2 = list(data2.keys())
    colonnes = len(cle2)+2

    if cellule.heure not in planning.liste_heure:
        planning.set_liste_eleve(dict_eleve[jour.j])

    for ind in range(1, len(sheet["A"])+1):
        for colonne in range(1,colonnes):
            if sheet.cell(3, colonne).value == cellule.heure:
                if sheet.cell(ind, 1).value == dernier_cheval:
                    sheet.cell(ind, colonne).value = None
                if sheet.cell(ind, 1).value == cellule.cheval:
                    print(ind,colonne,cellule.eleve)
                    sheet.cell(ind, colonne).value = cellule.eleve
                    dernier_cheval = cellule.cheval
                
    err = workbook.save(ancient_nom)
    if err == None:
        ajout_historique(
            "correction", (cellule.heure,cellule.cheval, cellule.eleve))
        varcavalier.set(cellule.cheval)

def absent():
    global dernier_cheval
    workbook = load_workbook(ancient_nom)
    sheet = workbook.active
    err = True
    for cel in planning.ancien_planning:
        if cel[0] == cellule.heure and cel[2] == cellule.eleve:
            planning.ancien_planning.remove(cel)
            err = False
    if not err:
        print(planning.ancien_planning)
        data2 = dict_eleve[jour.j]
        cle2 = list(data2.keys())
        colonnes = len(cle2)+2
        print("dercheval",dernier_cheval)
        if elevecarte == True:
            unesessionplus(cellule.eleve,cellule.heure)
        for ind in range(1, len(sheet["A"])+1):
            if sheet.cell(ind, 1).value == dernier_cheval:
                for colonne in range(1,colonnes):
                    if sheet.cell(3, colonne).value == cellule.heure:
                        sheet.cell(ind,
                                colonne).value = None
                        dernier_cheval = ""
        err = workbook.save(ancient_nom)
        if err == None:
            ajout_historique(
                "absence", (cellule.heure, cellule.eleve))
            varcavalier.set("cheval")
    else:
        messagebox.showerror("Erreur", "cette eleve avec ce chaval a cette heure ci n'est pas dans la liste du" + ancient_nom)



dernier_cheval = ""

boutton_absent = tk.Button(
    window, bg='#8abd45', height=1, width=4, text="ABS", command=absent, borderwidth=2)
boutton_correction = tk.Button(
    window, bg='#8abd45', height=1, text="correction", command=correction)
# Initialisation des variables de contrôle
varcavalier.set("cheval")
varcavalier1.set("cheval1")
varcavalier2.set("cheval2")

# Liste déroulante pour les élèves
eleve_listbox = tk.Listbox(window, yscrollcommand=True)

eleve_rattrapage = tk.Entry(window)

# style = ttk.Style()
# style.theme_use("clam")

label_eleve_rattrapage = tk.Label(
    window, text="Ajouter un nom", font=("Corbel", 13), bg='#b4b4b4')
boutton_eleve_rattrapage = tk.Button(
    window, width=8, bg='#8abd45', text="rattrapage", command=ajouter_rattrapage)

theme_entry = tk.Entry(window)

# style = ttk.Style()
# style.theme_use("clam")

label_theme = tk.Label(
    window, text="Ajouter un theme", font=("Corbel", 13), bg='#b4b4b4')
boutton_theme = tk.Button(
    window, width=12, bg='#8abd45', text="ajout du theme", command=ajouter_theme)
label_theme_actuelle = tk.Label(
    window, textvariable=theme, font=("Corbel", 13), bg='#b4b4b4')
label_theme_avant1 = tk.Label(
    window, textvariable=theme1, font=("Corbel", 13), bg='#b4b4b4')
label_theme_avant2 = tk.Label(
    window, textvariable=theme2, font=("Corbel", 13), bg='#b4b4b4')
label_theme_avant3 = tk.Label(
    window, textvariable=theme3, font=("Corbel", 13), bg='#b4b4b4')
theme1.set("theme1")
theme2.set("theme2")
theme3.set("theme3")

# Fonction appelée lorsqu'un élément est sélectionné dans la liste des élèves


def items_selected(event):
    global dernier_cheval,elevecarte
    # Indices des éléments sélectionnés
    selected_indices = eleve_listbox.curselection()
    eleve = eleve_listbox.get(selected_indices)

    cellule.set_eleve(eleve, selected_indices[0])
    elevecarte = False
    if isinstance(cellule.eleve[1], int):
        cellule.eleve = cellule.eleve[0]
        elevecarte = True
        
    ancient_cheval = planning.ancient_cheval_de(
            cellule.eleve, cellule.heure)

    # Mise à jour des étiquettes des chevaux associés
    if ancient_cheval[0][1] != "":
        varcavalier.set(ancient_cheval[0][0])
        dernier_cheval = ancient_cheval[0][0]
    else:
        varcavalier.set("cheval")
    if len(ancient_cheval) >= 2:
        varcavalier1.set(ancient_cheval[1][0])
    else:
        varcavalier1.set("cheval1")
    if len(ancient_cheval) >= 3:
        varcavalier2.set(ancient_cheval[2][0])
    else:
        varcavalier2.set("cheval2")
    colorier()
    colorier_ancient_chevaux(ancient_cheval)
    colorier_chevaux()
    for tup in planning.planning:
        if (cellule.heure, cellule.eleve) == (tup[0], tup[2]) and tup[1] in planning.cheval:
            cellule.set_cheval(tup[1], planning.index_cheval(tup[1]))
            varheure_cheval.set(
                f"HEURE DE TRAVAIL DE: {cellule.cheval}")
            inserer_liste_de_travaille()
    varajout.set(cellule.getCellule())


# Association de la fonction à l'événement de relâchement du bouton de la souris
eleve_listbox.bind('<ButtonRelease-1>', items_selected)

# Liste déroulante pour les chevaux
cheval_listbox = tk.Listbox(window, height=47)

# Fonction appelée lorsqu'un élément est sélectionné dans la liste des chevaux


def items_selected_cheval(event):
    # Indices des éléments sélectionnés
    selected_indices = cheval_listbox.curselection()
    cheval = cheval_listbox.get(selected_indices)
    cellule.set_cheval(cheval[1], selected_indices)
    varheure_cheval.set(f"HEURE DE TRAVAIL DE: {cellule.cheval}")
    inserer_liste_de_travaille()
    varajout.set(cellule.getCellule())


# Association de la fonction à l'événement de relâchement du bouton de la souris
cheval_listbox.bind('<ButtonRelease-1>', items_selected_cheval)

# Zone de texte pour afficher le planning
visu_fichier = tk.Text(window, width=70)
visu_fichier.config(state='disabled')

label_visu_fichier = tk.Label(
    window, text="PREVISUALISATION", font=("Corbel", 14), bg='#8abd45')


# Étiquette pour afficher des informations sur l'ajout
label_ajout = tk.Label(window, textvariable=varajout,
                       font=20, bg='#ffffff')


# Bouton pour ajouter une entrée
boutton_ajouter = tk.Button(
    window, text="Ajouter", command=ajouter, width=11, height=2, bg='#8abd45')


# Bouton pour supprimer une entrée
boutton_supprimer = tk.Button(
    window, text="Supprimer", command=supprimer, width=11, height=2, bg='#8abd45')


# Bouton pour enregistrer les modifications
boutton_enregistrer = tk.Button(
    window, text="ENREGISTRER", command=ecrire_fichier, width=12, font=("Helvetica", 18, "bold"), bg='#000000', fg='#ffffff')


# Étiquette pour afficher un message après l'enregistrement
label_enregistrer = tk.Label(
    window, text="Le fichier a bien été ebregistré", font=("Corbel", 13), bg='#b4b4b4')  # le fichier à bien été enregistré
label_enregistrer.config(fg="#b4b4b4")

# Étiquette pour afficher l'heure de travail du cheval
label_heure_cheval = tk.Label(
    window, textvariable=varheure_cheval, font=("Corbel", 13), bg='#8abd45')


# Liste déroulante pour les heures de travail
heure_listebox = tk.Listbox(window, width=25, height=5)



def rafraichir():
    global planning_theme
    dict_planning, cheval, heures,planning_theme = recupperation_excel(planning.name_fichier)
    elements_ajout = [element for element in dict_planning if element not in planning.planning]
    print(elements_ajout)
    elements_suppr = [element for element in planning.planning if element not in dict_planning]
    print(elements_suppr)
    eleves_carte=[]
    print("heure",heures)
    for heure in planning.liste_eleve:
        for eleve in planning.liste_eleve[heure]:
            if eleve[1] != -1:
                eleves_carte.append((heure,eleve[0]))
    print(eleves_carte)
    for cell in elements_suppr:
        if (cell[0],cell[2]) in eleves_carte:
            print('cessionplus',cell[2])
            unesessionplus(cell[2],cell[0])
    for cell in elements_ajout:
        if (cell[0],cell[2]) in eleves_carte:
            print('cessionMOINS',cell[2])
            unesessionmoins(cell[2],cell[0])

    planning.set_planning(dict_planning)
    planning.set_cheval(remplir_cheval(dict_cheval[jour.j]))
    planning.set_heure(heures)
    ajoutcheval()
    changer_heure()
    affichage_txt(jour, planning)

def ouvrir_excel():
    subprocess.Popen(['start', 'excel', planning.name_fichier], shell=True)

def ecrire_mail():
    err=False
    try:
        erreur = envoyer_email(user,planning.name_fichier,nom_fichier[0],"envoie du planning du "+str(extract_date_from_filename(planning.name_fichier))[0:11],mail)
        if erreur != {}:
            msgbox = tk.messagebox.showerror(
            title="envoie des parametre par mail", message=erreur)
        else:
            messagebox.showinfo("envoie du planning", "Le planning a été envoyé avec succès!")
    except Exception as e:
        err = True
        messagebox.showerror("Erreur", f"Erreur lors de l'envoie du planning : {e}")
    if not err:
        messagebox.showinfo("envoie du planning", "Le planning a été envoyé avec succès!")

bouton_ouvrir_excel = tk.Button(
    window, text="ouvrir", bg="#8abd45", command=ouvrir_excel)

bouton_rafraichir = tk.Button(
    window, text="rafraichir", bg="#8abd45", command=rafraichir)

bouton_word = tk.Button(
    window, text="word", bg="#8abd45", command=ecrire_word)

bouton_mail = tk.Button(
    window, text="mail", bg="#8abd45", command=ecrire_mail)

# Fonction appelée lorsqu'un élément est sélectionné dans la liste des heures de travail


def items_selected_heure_cheval(event):
    # Indices des éléments sélectionnés
    selected_indices = heure_listebox.curselection()
    (h, p) = heure_listebox.get(selected_indices)
    if h != cellule.heure:
        cellule.set_heure(h)
        changer_heure()
        cellule.set_eleve(p, -1)
    else:
        Nb = 0
        for i in range(0, eleve_listbox.size()):
            eleve = eleve_listbox.get(i)
            if isinstance(eleve[1], int):
                eleve = eleve[0]
            if p == eleve:
                cellule.set_eleve(p, Nb)
            Nb += 1
    varajout.set(cellule.getCellule())


# Association de la fonction à l'événement de relâchement du bouton de la souris
heure_listebox.bind('<ButtonRelease-1>', items_selected_heure_cheval)

# Étiquette pour afficher l'historique
label_historique = tk.Label(
    window, text="HISTORIQUE", font=("Corbel", 13), bg='#8abd45')

# Zone de texte pour afficher l'historique
historique = tk.Text(window, width=60, height=13)
historique.config(state='disabled')

# Création du menu
menubar = Menu(window)

# Création d'une liste déroulante pour sélectionner l'heure
listeCombo = ttk.Combobox(window, height=10, width=40)


# Fonction appelée lorsqu'un élément est sélectionné dans la liste déroulante


def action(event):
    select = listeCombo.get()  # Élément sélectionné dans la liste déroulante
    changement_heure(select)


listeCombo.bind("<<ComboboxSelected>>", action)


def ajout_des_commande_lena():
    listeCombo.delete(0, "end")
    listeCombo['values'] = list(planning.liste_eleve)


menubar = Menu(window)

# Ajout des éléments au menu
sousmenu = Menu(menubar, tearoff=0)
sousmenu.add_command(label="parametre", command=mode_parametre)
sousmenu.add_command(label="principal", command=mode_default)


# Ajout des éléments au menu
menubar.add_command(label="Nouveau", command=nouveau_fichier)
menubar.add_command(label="Jour", command=recup_donne)
menubar.add_cascade(label="Mode", menu=sousmenu)
menubar.add_command(label="Quitter!", command=window.quit)

# Affichage du menu dans la fenêtre
window.config(menu=menubar)



# interface_paramete()

para_visu_fichier = tk.Text(window, width=70)
para_visu_fichier.config(state='disabled')

# Étiquette pour afficher l'historique
para_label_historique = tk.Label(
    window, text="historique", font=("Corbel", 13), bg='#b4b4b4')

# Zone de texte pour afficher l'historique
para_historique = tk.Text(window, width=60, height=13)
para_historique.config(state='disabled')

para_listebox_heure = tk.Listbox(window, width=25, height=45)

heure = ""


def items_selected_heure(event):
    global heure
    # Indices des éléments sélectionnés
    selected_indices = para_listebox_heure.curselection()
    heure = para_listebox_heure.get(selected_indices)
    remplirlisteboxeleve()


# Association de la fonction à l'événement de relâchement du bouton de la souris
para_listebox_heure.bind('<ButtonRelease-1>', items_selected_heure)


para_listebox_eleve = tk.Listbox(window, width=25, height=12)


eleve = ""


def items_selected_eleve(event):
    # Indices des éléments sélectionnés
    global eleve
    selected_indices = para_listebox_eleve.curselection()
    eleve = para_listebox_eleve.get(selected_indices)


# Association de la fonction à l'événement de relâchement du bouton de la souris
para_listebox_eleve.bind('<ButtonRelease-1>', items_selected_eleve)


para_listebox_chevaux = tk.Listbox(window, width=25, height=45)

cheval = ""


def items_selected_cheval(event):
    global cheval
    # Indices des éléments sélectionnés
    selected_indices = para_listebox_chevaux.curselection()
    cheval = list(para_listebox_chevaux.get(selected_indices))


# Association de la fonction à l'événement de relâchement du bouton de la souris
para_listebox_chevaux.bind('<ButtonRelease-1>', items_selected_cheval)

# Création d'une liste déroulante pour sélectionner l'heure
para_listeCombo = ttk.Combobox(window,width=10)
para_listeCombo['values'] = ["Mercredi", "Samedi","Semaine"]

para_listeCombo_user = ttk.Combobox(window,width=10)
if user == "Lena":
    para_listeCombo_user['values'] = ["Lena", "Karine"]
elif user == "Karine":
    para_listeCombo_user['values'] = ["Karine", "Lena"]

parajour = ""


def action(event):
    global parajour
    parajour = para_listeCombo.get()  # Élément sélectionné dans la liste déroulante
    liste_cheval = []
    for cheval in dict_cheval[parajour]:
        liste_cheval.append([dict_cheval[parajour][cheval][0],cheval])
    remplirlisteboxcheval(liste_cheval)
    para_inserer_listebox(dict_eleve[parajour])
    visualiser_fichier_cavalier(dict_eleve[parajour])


para_input_chevaux = tk.Entry(window)
para_input_ind_chevaux = tk.Entry(window, width=3)
para_add_chevaux = tk.Button(
    window, text="ajouter cheval", command=add_cheval, width=18, bg='#8abd45')
para_suppr_chevaux = tk.Button(
    window, text="supprimer cheval", command=suppr_cheval, width=18, bg='#8abd45')

para_listeCombo.bind("<<ComboboxSelected>>", action)

para_input_heure = tk.Entry(window)

para_add_heure = tk.Button(
    window, text="creer heure", command=add_heure, width=18, bg='#8abd45')

para_suppr_heure = tk.Button(
    window, text="supprimer heure", command=lambda:suppr_heure(dict_eleve,heure), width=18, bg='#8abd45')


para_input_eleve = tk.Entry(window)

para_add_eleve = tk.Button(
    window, text="creer eleve", command=add_eleve, width=18, bg='#8abd45')

para_suppr_eleve = tk.Button(
    window, text="supprimer eleve", command=suppr_eleve, width=18, bg='#8abd45')

para_boutton_enregistrer = tk.Button(
    window, text="ENREGISTRER", command=para_enregistrer, width=12, font=("Helvetica", 18, "bold"), bg='#000000', fg='#ffffff')


def sauvegarder_mail():
    if para_entry_karine.get() and len(mail) > 0:
        mail[0] = para_entry_karine.get()
    elif para_entry_karine.get() and len(mail) == 0:
        mail.append(para_entry_karine.get())
    if para_entry_lena.get() and len(mail) > 1:
        mail[1] = para_entry_lena.get()
    elif para_entry_lena.get() and len(mail) == 1:
        mail.append(para_entry_lena.get())
        
    with open(path_mail, "w") as f:
        for i in mail:
            f.write(i + "\n")
        

para_label_mail_karine = tk.Label(window, text="mail karine", font=("Corbel", 13), bg='#b4b4b4')
para_entry_karine = tk.Entry(window, width=25)
if len(mail) > 0:
    para_entry_karine.insert(0, mail[0])

para_label_mail_lena = tk.Label(window, text="mail lena", font=("Corbel", 13), bg='#b4b4b4')
para_entry_lena = tk.Entry(window, width=25)
if len(mail) > 1:
    para_entry_lena.insert(0, mail[1])

# para_boutton_mail = tk.Button(window, bg='#8abd45', text="sauvegarder mail", command=sauvegarder_mail)

def ouvrir_excel():
    print("parajour",parajour)
    if parajour == "Mercredi":
        subprocess.Popen(['start', 'excel', path_Mercredi], shell=True)
    elif parajour == "Samedi":
        subprocess.Popen(['start', 'excel', path_Samedi], shell=True)
    elif parajour == "Semaine":
        subprocess.Popen(['start', 'excel', path_semaine], shell=True)

para_bouton_ouvrir_excel = tk.Button(
    window, text="ouvrir excel reference", bg="#8abd45", command=ouvrir_excel)

def toggle_entry_nbcarte():
    if v.get() == 1:
        para_nbcarte.config(state=NORMAL)
    else:
        para_nbcarte.config(state=DISABLED)

para_case = Checkbutton (variable = v,bg='#b4b4b4',text= "eleve à la carte",command=toggle_entry_nbcarte)

para_nbcarte = Entry(window)
para_nbcarte.insert(0, "nombre de seances")
para_nbcarte.config(state=DISABLED)

def on_para_nbcarte_click(event):
    if para_nbcarte.get() == 'nombre de seances':
        para_nbcarte.delete(0, tk.END)
para_nbcarte.bind('<FocusIn>', on_para_nbcarte_click)

para_bouton_importer_param = tk.Button(
    window, text="importer parametre",width=15, bg="#8abd45", command=importer_param)

para_bouton_exporter_param = tk.Button(
    window, text="exporter parametre",width=15, bg="#8abd45", command=exporter_param)


widgets_parametre.extend(
    [para_visu_fichier,para_bouton_importer_param,para_bouton_exporter_param,para_bouton_ouvrir_excel,para_case,para_label_mail_karine,para_entry_lena,para_label_mail_lena,para_entry_karine,para_nbcarte,para_listeCombo_user, para_label_historique, para_historique, para_listebox_eleve, para_listebox_heure, para_listeCombo, para_input_heure, para_input_eleve, para_boutton_enregistrer, para_suppr_eleve, para_add_eleve, para_suppr_heure, para_add_heure, para_listebox_chevaux, para_input_chevaux, para_add_chevaux, para_suppr_chevaux, para_input_ind_chevaux, image3, image4, para_image1])

widgets_principaux.extend([label_jour, label_heure, title_label, boutton_avancer_heure, boutton_reculer_heure, label_cavalier, label_cavalier2, label_cavalier3, label_cavalier6, label_cavalier4,
                           label_cavalier5, label_cavalier7, eleve_listbox, cheval_listbox, label_ajout, boutton_ajouter, boutton_supprimer, visu_fichier, label_visu_fichier,
                           boutton_enregistrer,bouton_rafraichir,bouton_ouvrir_excel, label_enregistrer, image1, image2, label_heure_cheval, heure_listebox, label_historique, historique, listeCombo, boutton_absent, boutton_correction, eleve_rattrapage, label_eleve_rattrapage, boutton_eleve_rattrapage,
                           theme_entry,bouton_mail,label_user,bouton_word,label_theme,boutton_theme,label_theme_actuelle,label_theme_avant1,label_theme_avant2,label_theme_avant3])
try:
    ftp = Ftp("83.113.54.154","lena","1234")
    connexion=True
except:
    print("pas de connexion internet")
    connexion =False
if connexion:
    fichiers = ftp.liste_fichiers()
    if fichiers:
        print(fichiers)
        version_fichiers=[]
        for fiche in fichiers:
            version_fiche  = fiche.replace(".zip","")
            version_fiche = version_fiche.split(" ")[-1]
            version_fichiers.append(float(version_fiche))
            
        version_fiche = max(version_fichiers)
        fiche = fichiers[version_fichiers.index(version_fiche)]
        print(version_fiche)
        print(fiche)
        if float(version_fiche) > version:
            import tkinter.messagebox as messagebox
            response = messagebox.askyesno("Mise à jour disponible", "Une nouvelle version est disponible. Voulez-vous la télécharger ?")
            if response:
                current_path = os.getcwd()#C:\Users\33621\Documents\cheval_python\ecuria
                no_current_path = current_path.rsplit('\\', 1)[0]#C:\Users\33621\Documents\cheval_python
                path_nv_version = os.path.join(no_current_path, fiche.replace(".zip",""))
                nom_appli_act = current_path.rsplit('\\', 1)[1]#ecuria
                print(current_path)
                def telecharger_et_mettre_a_jour():
                    ftp.telecharger_fichier_zip(fiche)
                    dezipper(fiche, no_current_path, suppr_rep_destination=False)
                    os.remove(fiche)

                def valider():
                    telecharger_et_mettre_a_jour()
                    if raccourci_bureau_var.get():
                        print("raccourci")
                        print("la",no_current_path,fiche.replace(".zip",".exe"))
                        raccourci(no_current_path,fiche.replace(".zip",".exe"))
                        # Code pour créer un raccourci sur le bureau
                        
                    if garder_parametre_var.get():
                        print("garder")
                        # Code pour garder les paramètres
                        
                        path_parametre= os.path.join(path_nv_version, "parametre")
                        if os.path.exists(path_parametre):
                            liste_fichier_parametre = os.listdir(os.path.join(current_path, "parametre"))
                            for fichier in os.listdir(path_parametre):
                                if fichier in liste_fichier_parametre:
                                    os.remove(os.path.join(path_nv_version, "parametre", fichier))
                                    shutil.copy(os.path.join(current_path, "parametre", fichier), os.path.join(path_nv_version, "parametre"))
                        else:
                            shutil.copytree("parametre", path_parametre)
                    
                    if supprimer_ancienne_version_var.get():
                        print("supprimer")
                        # os.remove(current_path)
                        bureau = os.path.join(os.path.expanduser("~"), "Desktop")
                        try:
                            os.remove(bureau + "\\" + nom_appli_act + ".lnk")
                        except:
                            pass
                        os.startfile(path_nv_version +"\\"+fiche.replace(".zip",".exe"))
                        top.destroy()
                        window.destroy()
                        # Code pour supprimer l'ancienne version
                    else:
                        top.destroy()

                def annuler():
                    top.destroy()

                # Création de la fenêtre top
                top = tk.Toplevel(window)
                top.title("Mise à jour")
                top.geometry("300x200")
                # Création des cases à cocher
                raccourci_bureau_var = tk.BooleanVar()
                raccourci_bureau_var.set(True)
                raccourci_bureau_checkbox = tk.Checkbutton(top, text="Raccourci bureau", variable=raccourci_bureau_var)

                supprimer_ancienne_version_var = tk.BooleanVar()
                supprimer_ancienne_version_checkbox = tk.Checkbutton(top, text="Supprimer ancienne version", variable=supprimer_ancienne_version_var)

                garder_parametre_var = tk.BooleanVar()
                garder_parametre_checkbox = tk.Checkbutton(top, text="Garder paramètre", variable=garder_parametre_var)

                # Création des boutons
                valider_button = tk.Button(top, text="Valider", command=valider)
                annuler_button = tk.Button(top, text="Annuler", command=annuler)

                # Placement des éléments dans la fenêtre top
                raccourci_bureau_checkbox.place(x=10, y=10)
                supprimer_ancienne_version_checkbox.place(x=10, y=40)
                garder_parametre_checkbox.place(x=10, y=70)
                valider_button.place(x=250, y=170)
                annuler_button.place(x=200, y=170)

                # Lancement de la boucle principale de la fenêtre top
                top.mainloop()

# Lancement de la boucle principale de l'application
window.mainloop()
